# -*- coding: utf-8 -*-
"""
Serviço de Exportação para o Gestor de Processos
Implementa exportação para PDF, Excel e backup de banco de dados
"""

import os
import shutil
import sqlite3
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from tkinter import filedialog, messagebox

from config import EXPORT_CONFIG
from utils.date_utils import DateUtils
from utils.string_utils import StringUtils
from logger_config import log_operation, log_error, export_logger

class ExportService:
    """Serviço para exportação de dados em diferentes formatos"""
    
    def __init__(self):
        self.export_config = EXPORT_CONFIG
        self.logger = export_logger
    
    def exportar_pdf(self, dados: List[Dict[str, Any]], 
                    filtros: Dict[str, str] = None,
                    arquivo_destino: str = None) -> bool:
        """Exporta dados para PDF formatado"""
        try:
            # Importa as bibliotecas necessárias
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
        except ImportError as e:
            error_msg = f"Biblioteca ReportLab não encontrada: {e}\n\nInstale com: pip install reportlab"
            messagebox.showerror("Erro", error_msg)
            log_error(f"ImportError na exportação PDF: {e}")
            return False
        
        try:
            if not dados:
                messagebox.showwarning("Aviso", "Nenhum dado para exportar")
                return False
            
            # Determina o nome do arquivo se não fornecido
            if not arquivo_destino:
                arquivo_destino = self._gerar_nome_arquivo_pdf(dados, filtros)
                arquivo_destino = filedialog.asksaveasfilename(
                    defaultextension=".pdf",
                    filetypes=[("PDF files", "*.pdf")],
                    title="Salvar relatório como",
                    initialfile=arquivo_destino
                )
                
                if not arquivo_destino:
                    return False
            
            # Cria o documento PDF
            doc = SimpleDocTemplate(
                arquivo_destino, 
                pagesize=A4, 
                rightMargin=36, 
                leftMargin=36, 
                topMargin=36, 
                bottomMargin=36
            )
            
            story = []
            styles = getSampleStyleSheet()
            
            # Adiciona título
            titulo = self._gerar_titulo_relatorio(dados, filtros)
            titulo_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Centralizado
            )
            story.append(Paragraph(titulo, titulo_style))
            
            # Adiciona informações do período
            info_periodo = self._gerar_info_periodo(dados)
            if info_periodo:
                info_style = ParagraphStyle(
                    'InfoStyle',
                    parent=styles['Normal'],
                    fontSize=10,
                    spaceAfter=20,
                    alignment=1
                )
                story.append(Paragraph(info_periodo, info_style))
            
            # Prepara dados da tabela
            dados_tabela = self._preparar_dados_tabela_pdf(dados)
            
            # Configuração das colunas otimizada para A4
            largura_total = 520
            larguras = [25, 60, 60, 75, 75, 85, 75, 85]  # Proporcionais
            soma = sum(larguras)
            larguras = [largura_total * (w / soma) for w in larguras]
            
            # Cria a tabela principal
            tabela_pdf = Table(dados_tabela, repeatRows=1, colWidths=larguras)
            
            # Define o estilo da tabela
            estilo_tabela = [
                # Cabeçalho
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#B4C6E7")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ]
            
            # Aplica cores alternadas para processos em andamento
            for i, processo in enumerate(dados, 1):
                if processo.get('situacao') == 'Em Andamento':
                    estilo_tabela.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor("#FFE699")))
            
            tabela_pdf.setStyle(TableStyle(estilo_tabela))
            story.append(tabela_pdf)
            
            # Adiciona seção de observações para processos em andamento
            processos_andamento = [p for p in dados if p.get('situacao') == 'Em Andamento']
            if processos_andamento:
                story.append(Spacer(1, 20))
                
                obs_title = ParagraphStyle(
                    'ObsTitle',
                    parent=styles['Heading2'],
                    fontSize=12,
                    spaceAfter=10
                )
                story.append(Paragraph("Observações - Processos em Andamento", obs_title))
                
                # Cria tabela de observações
                dados_obs = self._preparar_dados_observacoes(processos_andamento)
                if dados_obs:
                    tabela_obs = Table(dados_obs, colWidths=[100, 420])
                    estilo_obs = [
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#E7E6E6")),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 9),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ]
                    
                    # Cores alternadas para observações
                    for i in range(1, len(dados_obs)):
                        if (i - 1) % 2 == 0:
                            estilo_obs.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor("#FFE699")))
                    
                    tabela_obs.setStyle(TableStyle(estilo_obs))
                    story.append(tabela_obs)
            
            # Gera o PDF
            doc.build(story)
            
            log_operation(f"PDF exportado com sucesso: {arquivo_destino}")
            
            # Pergunta se deseja abrir o arquivo
            resposta = messagebox.askyesno(
                "Exportação Concluída",
                f"Relatório PDF exportado com sucesso:\n{arquivo_destino}\n\nDeseja abrir o arquivo agora?"
            )
            
            if resposta:
                os.startfile(arquivo_destino)
            
            return True
            
        except Exception as e:
            error_msg = f"Falha ao exportar PDF: {e}"
            messagebox.showerror("Erro", error_msg)
            log_error(f"Erro na exportação PDF: {e}")
            return False
    
    def exportar_excel(self, dados: List[Dict[str, Any]], 
                      filtros: Dict[str, str] = None,
                      arquivo_destino: str = None) -> bool:
        """Exporta dados para Excel formatado"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
            
        except ImportError as e:
            error_msg = f"Biblioteca OpenPyXL não encontrada: {e}\n\nInstale com: pip install openpyxl"
            messagebox.showerror("Erro", error_msg)
            log_error(f"ImportError na exportação Excel: {e}")
            return False
        
        try:
            if not dados:
                messagebox.showwarning("Aviso", "Nenhum dado para exportar")
                return False
            
            # Determina o nome do arquivo se não fornecido
            if not arquivo_destino:
                arquivo_destino = self._gerar_nome_arquivo_excel(dados, filtros)
                arquivo_destino = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                    title="Salvar planilha como",
                    initialfile=arquivo_destino
                )
                
                if not arquivo_destino:
                    return False
            
            # Cria o workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório de Processos"
            
            # Estilos
            fonte_cabecalho = Font(bold=True, size=11)
            fonte_dados = Font(size=10)
            preenchimento_cabecalho = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
            preenchimento_andamento = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            borda = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            alinhamento_centro = Alignment(horizontal='center', vertical='center')
            
            # Adiciona título
            titulo = self._gerar_titulo_relatorio(dados, filtros)
            ws['A1'] = titulo
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:H1')
            
            # Adiciona informações do período
            info_periodo = self._gerar_info_periodo(dados)
            if info_periodo:
                ws['A2'] = info_periodo
                ws['A2'].font = Font(size=10)
                ws['A2'].alignment = Alignment(horizontal='center')
                ws.merge_cells('A2:H2')
                linha_inicio = 4
            else:
                linha_inicio = 3
            
            # Cabeçalhos
            cabecalhos = ['Nº', 'Recebimento', 'Devolução', 'Contrato', 'Licitação', 'Modalidade', 'Situação', 'Contratado']
            for col, cabecalho in enumerate(cabecalhos, 1):
                celula = ws.cell(row=linha_inicio, column=col, value=cabecalho)
                celula.font = fonte_cabecalho
                celula.fill = preenchimento_cabecalho
                celula.border = borda
                celula.alignment = alinhamento_centro
            
            # Dados
            linha_atual = linha_inicio + 1
            for idx, processo in enumerate(dados, 1):
                # Formata dados do processo
                dados_linha = [
                    idx,
                    DateUtils.para_exibicao(processo.get('data_inicio', '')),
                    DateUtils.para_exibicao(processo.get('data_entrega', '')),
                    str(processo.get('numero_processo', '')),
                    str(processo.get('numero_licitacao', '')),
                    processo.get('modalidade', ''),
                    processo.get('situacao', ''),
                    processo.get('contratado', '')
                ]
                
                for col, valor in enumerate(dados_linha, 1):
                    celula = ws.cell(row=linha_atual, column=col, value=valor)
                    celula.font = fonte_dados
                    celula.border = borda
                    celula.alignment = alinhamento_centro
                    
                    # Destaca processos em andamento
                    if processo.get('situacao') == 'Em Andamento':
                        celula.fill = preenchimento_andamento
                
                linha_atual += 1
            
            # Adiciona seção de observações
            processos_andamento = [p for p in dados if p.get('situacao') == 'Em Andamento']
            if processos_andamento:
                linha_atual += 2
                
                # Título da seção
                ws.cell(row=linha_atual, column=1, value="Observações - Processos em Andamento").font = Font(bold=True, size=12)
                ws.merge_cells(f'A{linha_atual}:H{linha_atual}')
                linha_atual += 2
                
                # Cabeçalhos das observações
                ws.cell(row=linha_atual, column=1, value="Contrato").font = fonte_cabecalho
                ws.cell(row=linha_atual, column=2, value="Descrição/Observações").font = fonte_cabecalho
                
                for col in [1, 2]:
                    celula = ws.cell(row=linha_atual, column=col)
                    celula.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                    celula.border = borda
                    celula.alignment = alinhamento_centro
                
                linha_atual += 1
                
                # Dados das observações
                for processo in processos_andamento:
                    ws.cell(row=linha_atual, column=1, value=str(processo.get('numero_processo', ''))).border = borda
                    ws.cell(row=linha_atual, column=2, value=processo.get('descricao', '')).border = borda
                    
                    # Destaca linha
                    for col in [1, 2]:
                        ws.cell(row=linha_atual, column=col).fill = preenchimento_andamento
                    
                    linha_atual += 1
            
            # Ajusta largura das colunas
            larguras_colunas = [8, 12, 12, 15, 15, 20, 12, 25]
            for col, largura in enumerate(larguras_colunas, 1):
                ws.column_dimensions[get_column_letter(col)].width = largura
            
            # Configurações de página
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            
            # Margens
            ws.page_margins.left = 0.25
            ws.page_margins.right = 0.25
            ws.page_margins.top = 0.75
            ws.page_margins.bottom = 0.75
            
            # Configurações de impressão
            ws.print_options.horizontalCentered = True
            
            # Define área de impressão
            if ws.max_row > 0 and ws.max_column > 0:
                ws.print_area = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'
            
            # Salva o arquivo
            wb.save(arquivo_destino)
            
            log_operation(f"Excel exportado com sucesso: {arquivo_destino}")
            
            # Pergunta se deseja abrir o arquivo
            resposta = messagebox.askyesno(
                "Exportação Concluída",
                f"Planilha Excel exportada com sucesso:\n{arquivo_destino}\n\nDeseja abrir o arquivo agora?"
            )
            
            if resposta:
                os.startfile(arquivo_destino)
            
            return True
            
        except Exception as e:
            error_msg = f"Falha ao exportar Excel: {e}"
            messagebox.showerror("Erro", error_msg)
            log_error(f"Erro na exportação Excel: {e}")
            return False
    
    def exportar_banco(self, caminho_banco_origem: str, arquivo_destino: str = None) -> bool:
        """Exporta backup do banco de dados"""
        try:
            if not arquivo_destino:
                data_hora_atual = datetime.now()
                nome_padrao = f"banco_dados_{data_hora_atual.strftime('%H%M%S_%d%m%Y')}.db"
                arquivo_destino = filedialog.asksaveasfilename(
                    defaultextension=".db",
                    filetypes=[("Banco de Dados SQLite", "*.db")],
                    title="Exportar banco de dados",
                    initialfile=nome_padrao
                )
                
                if not arquivo_destino:
                    return False
            
            # Copia o arquivo do banco
            shutil.copy2(caminho_banco_origem, arquivo_destino)
            
            log_operation(f"Banco de dados exportado: {arquivo_destino}")
            
            # Pergunta se deseja abrir o arquivo
            resposta = messagebox.askyesno(
                "Exportação Concluída",
                f"Banco de dados exportado com sucesso:\n{arquivo_destino}\n\nDeseja abrir o arquivo agora?"
            )
            
            if resposta:
                os.startfile(arquivo_destino)
            
            return True
            
        except Exception as e:
            error_msg = f"Erro ao exportar banco: {e}"
            messagebox.showerror("Erro", error_msg)
            log_error(f"Erro na exportação do banco: {e}")
            return False
    
    def _gerar_nome_arquivo_pdf(self, dados: List[Dict[str, Any]], 
                               filtros: Dict[str, str] = None) -> str:
        """Gera nome do arquivo PDF baseado nos dados e filtros"""
        data_hora_atual = datetime.now()
        # Padrão solicitado: Relatório_PDF_DDMMYYYY_HHMMSS
        nome = f"Relatório_PDF_{data_hora_atual.strftime('%d%m%Y_%H%M%S')}.pdf"
        
        # Adiciona filtro de secretaria se disponível
        if filtros and filtros.get('secretaria'):
            secretaria = StringUtils.gerar_nome_arquivo_seguro(filtros['secretaria'])
            nome = f"{secretaria}_{nome}"
        
        return nome
    
    def _gerar_nome_arquivo_excel(self, dados: List[Dict[str, Any]], 
                                 filtros: Dict[str, str] = None) -> str:
        """Gera nome do arquivo Excel baseado nos dados e filtros"""
        data_hora_atual = datetime.now()
        # Padrão solicitado: Relatório_EXCEL_DDMMYYYY_HHMMSS
        nome = f"Relatório_EXCEL_{data_hora_atual.strftime('%d%m%Y_%H%M%S')}.xlsx"
        
        # Adiciona filtro de secretaria se disponível
        if filtros and filtros.get('secretaria'):
            secretaria = StringUtils.gerar_nome_arquivo_seguro(filtros['secretaria'])
            nome = f"{secretaria}_{nome}"
        
        return nome
    
    def _determinar_tipo_relatorio(self, dados: List[Dict[str, Any]]) -> str:
        """Determina o tipo de relatório baseado nas datas dos dados"""
        if not dados:
            return "Vazio"
        
        # Coleta datas de recebimento
        datas_recebimento = []
        for processo in dados:
            data_str = processo.get('data_inicio', '')
            if data_str:
                try:
                    if '/' in data_str:
                        data_obj = datetime.strptime(data_str, "%d/%m/%Y")
                    else:
                        data_obj = datetime.strptime(data_str, "%Y-%m-%d")
                    datas_recebimento.append(data_obj)
                except ValueError:
                    continue
        
        if not datas_recebimento:
            return "Personalizado"
        
        # Determina o período
        data_inicio = min(datas_recebimento)
        data_fim = max(datas_recebimento)
        diferenca = (data_fim - data_inicio).days
        
        if (data_inicio.month == data_fim.month and 
            data_inicio.year == data_fim.year):
            return "Mensal"
        elif data_inicio.year == data_fim.year:
            return "Anual"
        elif diferenca == 0:
            return "Diário"
        elif diferenca <= 7:
            return "Semanal"
        elif diferenca <= 15:
            return "Quinzenal"
        elif diferenca <= 31:
            return "Mensal"
        elif diferenca <= 62:
            return "Bimestral"
        elif diferenca <= 93:
            return "Trimestral"
        elif diferenca <= 186:
            return "Semestral"
        else:
            return "Personalizado"
    
    def _gerar_titulo_relatorio(self, dados: List[Dict[str, Any]], 
                               filtros: Dict[str, str] = None) -> str:
        """Gera título do relatório"""
        tipo_relatorio = self._determinar_tipo_relatorio(dados)
        titulo = f"Relatório {tipo_relatorio} de Processos"
        
        if filtros and filtros.get('secretaria'):
            titulo += f" - {filtros['secretaria']}"
        
        return titulo
    
    def _gerar_info_periodo(self, dados: List[Dict[str, Any]]) -> str:
        """Gera informações do período do relatório"""
        if not dados:
            return ""
        
        # Coleta datas
        datas_recebimento = []
        for processo in dados:
            data_str = processo.get('data_inicio', '')
            if data_str:
                try:
                    if '/' in data_str:
                        data_obj = datetime.strptime(data_str, "%d/%m/%Y")
                    else:
                        data_obj = datetime.strptime(data_str, "%Y-%m-%d")
                    datas_recebimento.append(data_obj)
                except ValueError:
                    continue
        
        if not datas_recebimento:
            return ""
        
        data_inicio = min(datas_recebimento)
        data_fim = max(datas_recebimento)
        
        if data_inicio == data_fim:
            return f"Data: {data_inicio.strftime('%d/%m/%Y')}"
        else:
            return f"Período: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
    
    def _preparar_dados_tabela_pdf(self, dados: List[Dict[str, Any]]) -> List[List[str]]:
        """Prepara dados para tabela do PDF"""
        # Cabeçalho
        tabela_dados = [['Nº', 'Recebimento', 'Devolução', 'Contrato', 'Licitação', 'Modalidade', 'Situação', 'Contratado']]
        
        # Dados
        for idx, processo in enumerate(dados, 1):
            linha = [
                str(idx),
                DateUtils.para_exibicao(processo.get('data_inicio', '')),
                DateUtils.para_exibicao(processo.get('data_entrega', '')),
                str(processo.get('numero_processo', '')),
                str(processo.get('numero_licitacao', '')),
                processo.get('modalidade', ''),
                processo.get('situacao', ''),
                processo.get('contratado', '')
            ]
            tabela_dados.append(linha)
        
        return tabela_dados
    
    def _preparar_dados_observacoes(self, processos_andamento: List[Dict[str, Any]]) -> List[List[str]]:
        """Prepara dados para tabela de observações"""
        if not processos_andamento:
            return []
        
        # Cabeçalho
        dados_obs = [['Contrato', 'Descrição/Observações']]
        
        # Dados
        for processo in processos_andamento:
            linha = [
                str(processo.get('numero_processo', '')),
                processo.get('descricao', '')
            ]
            dados_obs.append(linha)
        
        return dados_obs

# Instância global do serviço
_export_service = ExportService()

# Funções de conveniência
def exportar_pdf(dados: List[Dict[str, Any]], filtros: Dict[str, str] = None) -> bool:
    """Exporta dados para PDF"""
    return _export_service.exportar_pdf(dados, filtros)

def exportar_excel(dados: List[Dict[str, Any]], filtros: Dict[str, str] = None) -> bool:
    """Exporta dados para Excel"""
    return _export_service.exportar_excel(dados, filtros)

def exportar_banco(caminho_banco: str) -> bool:
    """Exporta backup do banco de dados"""
    return _export_service.exportar_banco(caminho_banco)

def get_export_service() -> ExportService:
    """Retorna a instância do serviço de exportação"""
    return _export_service