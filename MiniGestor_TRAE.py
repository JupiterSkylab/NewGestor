# --- Bibliotecas padrão --- #
import ctypes
import json
import logging
import math
import os
import re
import shutil
import sqlite3
import subprocess
import sys
import threading
import time
# --- Bibliotecas da interface Tkinter ---
import tkinter as tk
import tkinter.font
import unicodedata
from contextlib import closing
from ctypes import wintypes
from datetime import datetime, timedelta
from tkinter import (END, Button, Listbox, Menu, Toplevel, filedialog,
                     messagebox, ttk)
from typing import Any, Dict, List, Optional, Tuple, Union

# --- Bibliotecas de terceiros ---
import pyperclip
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

# --- Constantes e Configurações --- #
# Caminhos de arquivos
CAMINHO_BANCO = 'meus_trabalhos.db'
CONFIG_COLUNAS_PATH = 'config_colunas.json'
NOMES_PERSISTENTES_PATH = 'nomes_persistentes.json'

# Configurações de cache
CACHE_TTL_DEFAULT = 300  # segundos (5 minutos)

# Configurações de UI
COR_FEEDBACK_POSITIVO = "#e0f7fa"
COR_FEEDBACK_NEGATIVO = "#ffebee"
COR_DESTAQUE = "#0078D7"
TOOLTIP_BG = "#2b2b2b"
TOOLTIP_SELECTED_BG = "#394857"
TOOLTIP_HOVER_BG = "#D6EAF8"
TOOLTIP_HOVER_FG = "#1B4F72"
ROW_TOOLTIP_HOVER_BG = "#a4b7d7"
BUTTON_PRIMARY_BG = "#394857"
BUTTON_PRIMARY_FG = "#66c0f4"
BUTTON_PRIMARY_ACTIVE_BG = "#23262e"
BUTTON_PRIMARY_ACTIVE_FG = "#66c0f4"
BUTTON_PRIMARY_HIGHLIGHT = "#394857"
BUTTON_DANGER_BG = "#e74c3c"
BUTTON_DANGER_FG = "white"
BUTTON_DANGER_ACTIVE_BG = "#c0392b"
BUTTON_DANGER_ACTIVE_FG = "white"
BUTTON_DANGER_HIGHLIGHT = "#e74c3c"

# Configuração do encoding
sys.stdout.reconfigure(encoding='utf-8')


def center_window(window, width=860, height=650):
    try:
        screen_width = window.winfo_screenwidth()
        screen_height = window.winfo_screenheight()
        width = min(width, max(1, screen_width - 20))
        height = min(height, max(1, screen_height - 40))
        x = (screen_width - width) // 2
        y = (screen_height - height) // 6
        x = max(0, x)
        y = max(0, y)
        window.geometry(f"{width}x{height}+{x}+{y}")
        try:
            window.maxsize(screen_width, screen_height)
        except Exception:
            pass
    except Exception:
        try:
            window.geometry("800x600+10+10")
        except Exception:
            pass


class DateUtils:
    FORMATO_EXIBICAO = "%d/%m/%Y"
    FORMATO_BANCO = "%Y-%m-%d"
    FORMATO_DATETIME_BANCO = "%Y-%m-%d %H:%M:%S"
    FORMATO_DATETIME_EXIBICAO = "%d/%m/%Y %H:%M"

    @staticmethod
    def para_exibicao(data_str: str) -> str:
        if not data_str:
            return ""
        if "/" in data_str:
            return data_str
        try:
            from datetime import datetime
            d = datetime.strptime(str(data_str).strip(), DateUtils.FORMATO_BANCO)
            return d.strftime(DateUtils.FORMATO_EXIBICAO)
        except Exception:
            return str(data_str)

    @staticmethod
    def formatar_data_hora(data_str: str) -> str:
        if not data_str:
            return ""
        from datetime import datetime
        formatos = [
            DateUtils.FORMATO_DATETIME_BANCO,
            DateUtils.FORMATO_BANCO,
            DateUtils.FORMATO_DATETIME_EXIBICAO,
            DateUtils.FORMATO_EXIBICAO,
            "%Y-%m-%d %H:%M:%S.%f",
            "%d/%m/%Y %H:%M:%S",
        ]
        s = str(data_str).strip()
        for f in formatos:
            try:
                dt = datetime.strptime(s, f)
                if dt.hour or dt.minute or dt.second:
                    return dt.strftime(DateUtils.FORMATO_DATETIME_EXIBICAO)
                return dt.strftime(DateUtils.FORMATO_EXIBICAO)
            except Exception:
                pass
        return s

    @staticmethod
    def obter_data_atual() -> str:
        from datetime import datetime
        return datetime.now().strftime(DateUtils.FORMATO_EXIBICAO)

    @staticmethod
    def obter_data_hora_atual() -> str:
        from datetime import datetime
        return datetime.now().strftime(DateUtils.FORMATO_DATETIME_EXIBICAO)

    @staticmethod
    def obter_data_hora_atual_banco() -> str:
        from datetime import datetime
        return datetime.now().strftime(DateUtils.FORMATO_DATETIME_BANCO)

    @staticmethod
    def _parse_data_flexivel(data_str: str):
        if not data_str:
            return None
        from datetime import datetime
        formatos = [
            DateUtils.FORMATO_EXIBICAO,
            DateUtils.FORMATO_BANCO,
            DateUtils.FORMATO_DATETIME_EXIBICAO,
            DateUtils.FORMATO_DATETIME_BANCO,
            "%Y-%m-%d %H:%M:%S.%f",
            "%d/%m/%Y %H:%M:%S",
        ]
        s = str(data_str).strip()
        for f in formatos:
            try:
                return datetime.strptime(s, f)
            except Exception:
                pass
        return None


def get_promessa_by_id(cursor, rowid: int):
    cursor.execute(
        "SELECT descricao, pessoa, data_criacao FROM promessas WHERE rowid = ?",
        (rowid,)
    )
    return cursor.fetchone()


def update_promessa_descricao(cursor, rowid: int, nova_descricao: str) -> bool:
    cursor.execute(
        "UPDATE promessas SET descricao = ? WHERE rowid = ?",
        (nova_descricao, rowid)
    )
    return cursor.rowcount > 0


def delete_promessa(cursor, rowid: int) -> bool:
    cursor.execute(
        "DELETE FROM promessas WHERE rowid = ?",
        (rowid,)
    )
    return cursor.rowcount > 0


class LembreteDetailsWindow:
    def __init__(self, parent: tk.Toplevel, mapping: dict, cursor, conn, on_reload, start_index: int = 0):
        self.parent = parent
        self.mapping = mapping
        self.cursor = cursor
        self.conn = conn
        self.on_reload = on_reload
        self.current_rowid = None
        self.index_var = tk.IntVar(value=start_index)
        self.total = len(mapping)

        self.win = tk.Toplevel(parent)
        self.win.withdraw()
        self.win.title("📋 Detalhes do Lembrete")
        self.win.resizable(True, True)
        self.win.configure(bg="#F5F7FA")
        self.win.transient(parent)
        self.win.grab_set()

        try:
            w, h, xi, yi = carregar_tamanho_janela('lembrete_detalhes', 425, 420)
            if xi is not None and yi is not None:
                self.win.geometry(f"{w}x{h}+{xi}+{yi}")
            else:
                center_window(self.win, w, h)
        except Exception:
            center_window(self.win, 425, 420)
        try:
            self.win.protocol("WM_DELETE_WINDOW", lambda: self.close())
        except Exception:
            pass

        nav = tk.Frame(self.win, bg="#F5F7FA")
        nav.pack(fill=tk.X, padx=16, pady=(4, 0))
        self.btn_prev = tk.Button(nav, text="◀ Anterior")
        self.btn_prev.pack(side=tk.LEFT, padx=5)
        self.lbl_count = tk.Label(nav, text="", font=("Segoe UI", 10, "bold"), bg="#F5F7FA")
        self.lbl_count.pack(side=tk.LEFT, expand=True)
        self.btn_next = tk.Button(nav, text="Próximo ▶")
        self.btn_next.pack(side=tk.RIGHT, padx=5)

        frame = tk.Frame(self.win, bg="#F5F7FA", padx=25, pady=10)
        frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(frame, text="📋 Detalhes do Lembrete", font=("Segoe UI", 14, "bold"), bg="#F5F7FA", fg="#2C3E50").pack(
            pady=(0, 12))

        self.info = tk.Frame(frame, bg="#F5F7FA")
        self.info.pack(fill=tk.X, pady=(0, 12))
        tk.Label(self.info, text="🕒 Criado às:", font=("Segoe UI", 10, "bold"), bg="#F5F7FA", fg="#34495E",
                 anchor="w").grid(row=0, column=0, sticky=tk.W, pady=(6, 0), padx=(0, 0))
        self.lbl_created = tk.Label(self.info, text="", font=("Segoe UI", 10), bg="#F5F7FA", fg="#2C3E50", anchor="w")
        self.lbl_created.grid(row=0, column=1, sticky=tk.W + tk.E, pady=(6, 0), padx=(1, 0))
        self.info.grid_columnconfigure(0, weight=0)
        self.info.grid_columnconfigure(1, weight=1, minsize=330)

        tk.Label(frame, text="📝 Descrição:", font=("Segoe UI", 10, "bold"), bg="#F5F7FA", fg="#34495E",
                 anchor="w").pack(anchor="w", pady=(6, 4))
        desc_frame = tk.Frame(frame, bg="white", relief="solid", bd=1)
        desc_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 2))
        self.text = tk.Text(desc_frame, wrap=tk.WORD, font=("Segoe UI", 10), bg="white", fg="#2C3E50", padx=8, pady=3,
                            height=10, relief="flat")
        self.text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll = tk.Scrollbar(desc_frame, orient=tk.VERTICAL, command=self.text.yview)
        self.text.configure(yscrollcommand=scroll.set)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)

        try:
            self.text.configure(undo=True, maxundo=-1)
            self.text.bind('<Control-z>', lambda e: (e.widget.edit_undo(), 'break'))
            self.text.bind('<Control-BackSpace>', self._delete_prev_word)
        except Exception:
            pass

        try:
            self.text.bind('<Double-Button-1>', lambda e: self.enable_edit())
        except Exception:
            pass

        bottom = tk.Frame(frame, bg="#F5F7FA")
        estilo_padrao = {
            "bg": BUTTON_PRIMARY_BG,
            "fg": BUTTON_PRIMARY_FG,
            "activebackground": BUTTON_PRIMARY_ACTIVE_BG,
            "activeforeground": BUTTON_PRIMARY_ACTIVE_FG,
            "font": ("Segoe UI", 10, "bold"),
            "relief": tk.FLAT,
            "bd": 0,
            "highlightthickness": 1,
            "highlightbackground": BUTTON_PRIMARY_HIGHLIGHT,
        }
        estilo_excluir = dict(estilo_padrao)
        estilo_excluir.update({
            "bg": BUTTON_DANGER_BG,
            "fg": BUTTON_DANGER_FG,
            "activebackground": BUTTON_DANGER_ACTIVE_BG,
            "activeforeground": BUTTON_DANGER_ACTIVE_FG,
            "highlightbackground": BUTTON_DANGER_HIGHLIGHT,
            "highlightthickness": 2
        })
        self.btn_delete = tk.Button(bottom, text="Excluir Lembrete", width=12)
        self.btn_delete.configure(**estilo_excluir)
        self.btn_delete.config(padx=11, pady=9, height=2)
        self.btn_edit = tk.Button(bottom, text="Editar Lembrete", width=12)
        self.btn_edit.configure(**estilo_padrao)
        self.btn_edit.config(padx=11, pady=9, height=2)
        bottom.pack(side=tk.BOTTOM, fill=tk.X, padx=25, pady=(14, 0))
        bottom.configure(height=38)
        bottom.pack_propagate(False)
        self.btn_delete.pack(side=tk.LEFT, padx=8, pady=2, ipady=2)
        self.btn_edit.pack(side=tk.RIGHT, padx=8, pady=2, ipady=2)

        self.btn_prev.config(command=self.prev)
        self.btn_next.config(command=self.next)
        self.btn_edit.config(command=self.enable_edit)
        self.btn_delete.config(command=self.delete_current)
        self.win.bind("<Escape>", lambda e: self.close())
        self.win.bind("<Left>", lambda e: self.prev())
        self.win.bind("<Right>", lambda e: self.next())

        self.load_index(self.index_var.get())
        self.update_nav()
        self.win.deiconify()
        self.win.focus_set()

    def update_nav(self):
        idx = self.index_var.get()
        self.btn_prev.config(state="normal" if idx > 0 else "disabled")
        self.btn_next.config(state="normal" if idx < self.total - 1 else "disabled")
        self.lbl_count.config(text=f"Lembrete {idx + 1} de {self.total}")
        self.win.title(f"📋 Detalhes do Lembrete ({idx + 1}/{self.total})")

    def load_index(self, idx: int):
        rowid = self.mapping.get(idx)
        if rowid is None:
            return
        self.current_rowid = rowid
        dados = get_promessa_by_id(self.cursor, rowid)
        if not dados:
            return
        descricao, pessoa, data_criacao = dados
        try:
            dt = DateUtils._parse_data_flexivel(str(data_criacao))
            if dt:
                dias = ["segunda-feira", "terça-feira", "quarta-feira", "quinta-feira", "sexta-feira", "sábado",
                        "domingo"]
                dia = dias[dt.weekday()]
                created_fmt = f"{dt.strftime('%H:%M')} do dia {dt.strftime('%d/%m/%Y')}, de uma {dia}"
            else:
                created_fmt = str(data_criacao)
        except Exception:
            created_fmt = str(data_criacao)
        self.lbl_created.config(text=created_fmt)
        try:
            self.win.update_idletasks()
        except Exception:
            pass
        try:
            base_h = 8
            max_h = base_h * 2
            self.text.config(state="normal")
            try:
                display_lines = int(
                    self.text.tk.call(self.text._w, 'count', '-update', '-displaylines', '1.0', 'end-1c'))
            except Exception:
                display_lines = self.text.get("1.0", "end-1c").count("\n") + 1
            desired_h = max(base_h, min(display_lines + 1, max_h))
            self.text.config(height=desired_h)
            self.text.config(state="disabled")
            self.win.update_idletasks()
        except Exception:
            pass
        self.text.config(state="normal")
        self.text.delete("1.0", tk.END)
        self.text.insert("1.0", descricao or "")
        self.text.config(state="disabled")

    def enable_edit(self):
        self.text.config(state="normal")
        self.text.focus_set()
        self.btn_edit.config(text="Atualizar Lembrete", command=self.update_current)

    def update_current(self):
        nova_desc = self.text.get("1.0", "end-1c").strip()
        if not nova_desc:
            messagebox.showwarning("Aviso", "Digite a descrição.", parent=self.win)
            return
        try:
            ok = update_promessa_descricao(self.cursor, self.current_rowid, nova_desc)
            if ok:
                self.conn.commit()
                try:
                    c = globals().get('cache')
                    f = globals().get('atualizar_cor_botao_lembrete')
                    if c:
                        c.invalidate('count_promessas')
                    if callable(f):
                        f()
                except Exception:
                    pass
                messagebox.showinfo("Sucesso", "Lembrete atualizado.", parent=self.win)
            else:
                messagebox.showwarning("Aviso", "Lembrete não encontrado.", parent=self.win)
        except Exception:
            messagebox.showerror("Erro", "Falha ao atualizar lembrete.", parent=self.win)

    def delete_current(self):
        if self.current_rowid is None:
            return
        try:
            ok = delete_promessa(self.cursor, self.current_rowid)
            if ok:
                self.conn.commit()
                messagebox.showinfo("Sucesso", "Lembrete excluído.", parent=self.win)
                if callable(self.on_reload):
                    self.on_reload()
                self.close()
            else:
                messagebox.showwarning("Aviso", "Lembrete não encontrado.", parent=self.win)
        except Exception:
            messagebox.showerror("Erro", "Falha ao excluir lembrete.", parent=self.win)

    def prev(self):
        idx = self.index_var.get()
        if idx <= 0:
            return
        self.index_var.set(idx - 1)
        self.load_index(self.index_var.get())
        self.update_nav()

    def next(self):
        idx = self.index_var.get()
        if idx >= self.total - 1:
            return
        self.index_var.set(idx + 1)
        self.load_index(self.index_var.get())
        self.update_nav()

    def _delete_prev_word(self, event):
        try:
            event.widget.delete('insert wordstart', 'insert')
            return 'break'
        except Exception:
            return None

    def close(self):
        try:
            try:
                self.win.update_idletasks()
                salvar_tamanho_janela('lembrete_detalhes', self.win.winfo_width(), self.win.winfo_height(),
                                      self.win.winfo_x(), self.win.winfo_y())
            except Exception:
                pass
            self.win.destroy()
        except Exception:
            pass


# --- Variáveis globais para controle de janelas ---
janela_visualizar_ativa = None
janela_visualizar_excluida_ativa = None


# --- Função para mostrar o guia de atalhos de teclado ---
def mostrar_guia_atalhos():
    """Mostra um guia com todos os atalhos de teclado disponíveis."""
    janela_atalhos = tk.Toplevel(janela)
    janela_atalhos.title("Guia de Atalhos de Teclado")
    try:
        w, h, xi, yi = carregar_tamanho_janela('guia_atalhos', 600, 500)
        if xi is not None and yi is not None:
            janela_atalhos.geometry(f"{w}x{h}+{xi}+{yi}")
        else:
            center_window(janela_atalhos, w, h)
    except Exception:
        janela_atalhos.geometry("600x500")
    janela_atalhos.resizable(True, True)
    janela_atalhos.transient(janela)
    try:
        def _close():
            try:
                janela_atalhos.update_idletasks()
                salvar_tamanho_janela('guia_atalhos', janela_atalhos.winfo_width(), janela_atalhos.winfo_height(),
                                      janela_atalhos.winfo_x(), janela_atalhos.winfo_y())
            except Exception:
                pass
            janela_atalhos.destroy()

        janela_atalhos.protocol("WM_DELETE_WINDOW", _close)
    except Exception:
        pass
    janela_atalhos.grab_set()


# Controles aprimorados para caixas de texto
def enhance_text_controls(widget: tk.Text):
    try:
        widget.configure(undo=True)
        widget.bind('<Control-z>', lambda e: (e.widget.edit_undo(), 'break'))
        widget.bind('<Control-BackSpace>', lambda e: (e.widget.delete('insert wordstart', 'insert'), 'break'))
    except Exception:
        pass


def bind_entry_word_delete(root: tk.Tk):
    def _handler(event):
        try:
            entry = event.widget
            idx = entry.index('insert')
            text = entry.get()
            import re
            before = text[:idx]
            m = re.search(r"\s*\w+\s*$", before)
            if m:
                start = idx - len(m.group(0))
                entry.delete(start, idx)
                return 'break'
        except Exception:
            return None

    try:
        root.bind_class('Entry', '<Control-BackSpace>', _handler)
    except Exception:
        pass


def bind_text_word_delete(root: tk.Tk):
    def _handler(event):
        try:
            event.widget.delete('insert wordstart', 'insert')
            return 'break'
        except Exception:
            return None

    try:
        root.bind_class('Text', '<Control-BackSpace>', _handler)
    except Exception:
        pass


def setup_global_text_shortcuts(root: tk.Tk):
    def _copy(e):
        try:
            e.widget.event_generate('<<Copy>>')
            return 'break'
        except Exception:
            return None

    def _cut(e):
        try:
            e.widget.event_generate('<<Cut>>')
            return 'break'
        except Exception:
            return None

    def _paste(e):
        try:
            e.widget.event_generate('<<Paste>>')
            return 'break'
        except Exception:
            return None

    def _select_all(e):
        try:
            e.widget.tag_add('sel', '1.0', 'end-1c')
            return 'break'
        except Exception:
            return None

    def _undo(e):
        try:
            e.widget.edit_undo()
            return 'break'
        except Exception:
            return None

    def _redo(e):
        try:
            e.widget.edit_redo()
            return 'break'
        except Exception:
            return None

    def _del_prev_word(e):
        try:
            e.widget.delete('insert -1c wordstart', 'insert')
            return 'break'
        except Exception:
            return None

    try:
        root.bind_class('Text', '<Control-c>', _copy)
        root.bind_class('Text', '<Control-x>', _cut)
        root.bind_class('Text', '<Control-v>', _paste)
        root.bind_class('Text', '<Control-a>', _select_all)
        root.bind_class('Text', '<Control-z>', _undo)
        root.bind_class('Text', '<Control-y>', _redo)
        root.bind_class('Text', '<Control-BackSpace>', _del_prev_word)
    except Exception:
        pass


def setup_global_entry_shortcuts(root: tk.Tk):
    def _copy(e):
        try:
            e.widget.event_generate('<<Copy>>')
            return 'break'
        except Exception:
            return None

    def _cut(e):
        try:
            e.widget.event_generate('<<Cut>>')
            return 'break'
        except Exception:
            return None

    def _paste(e):
        try:
            e.widget.event_generate('<<Paste>>')
            return 'break'
        except Exception:
            return None

    def _select_all(e):
        try:
            e.widget.select_range(0, 'end')
            e.widget.icursor('end')
            return 'break'
        except Exception:
            return None

    def _del_prev_word(e):
        try:
            entry = e.widget
            idx = entry.index('insert')
            text = entry.get()
            import re
            before = text[:idx]
            m = re.search(r"\s*\w+$", before)
            if m:
                start = idx - len(m.group(0))
                entry.delete(start, idx)
                return 'break'
        except Exception:
            return None

    try:
        root.bind_class('Entry', '<Control-c>', _copy)
        root.bind_class('Entry', '<Control-x>', _cut)
        root.bind_class('Entry', '<Control-v>', _paste)
        root.bind_class('Entry', '<Control-a>', _select_all)
        root.bind_class('Entry', '<Control-BackSpace>', _del_prev_word)
    except Exception:
        pass


def setup_context_menu(root: tk.Tk):
    def _show_menu(e):
        try:
            w = e.widget
            menu = tk.Menu(w, tearoff=0)

            def _do(cmd):
                try:
                    if cmd == 'copy':
                        w.event_generate('<<Copy>>')
                    elif cmd == 'cut':
                        w.event_generate('<<Cut>>')
                    elif cmd == 'paste':
                        w.event_generate('<<Paste>>')
                    elif cmd == 'undo':
                        w.edit_undo()
                    elif cmd == 'redo':
                        w.edit_redo()
                    elif cmd == 'select_all':
                        if isinstance(w, tk.Text):
                            w.tag_add('sel', '1.0', 'end-1c')
                        else:
                            w.select_range(0, 'end')
                            w.icursor('end')
                except Exception:
                    pass

            menu.add_command(label='Copiar', command=lambda: _do('copy'))
            menu.add_command(label='Recortar', command=lambda: _do('cut'))
            menu.add_command(label='Colar', command=lambda: _do('paste'))
            menu.add_separator()
            menu.add_command(label='Desfazer', command=lambda: _do('undo'))
            menu.add_command(label='Refazer', command=lambda: _do('redo'))
            menu.add_separator()
            menu.add_command(label='Selecionar tudo', command=lambda: _do('select_all'))
            try:
                menu.tk_popup(e.x_root, e.y_root)
            finally:
                menu.grab_release()
            return 'break'
        except Exception:
            return None

    try:
        root.bind_class('Text', '<Button-3>', _show_menu)
        root.bind_class('Entry', '<Button-3>', _show_menu)
    except Exception:
        pass

    # Estilo para os títulos das categorias
    estilo_titulo = {"font": ("Arial", 12, "bold"), "pady": 10, "anchor": "w"}

    # Estilo para os atalhos
    estilo_atalho = {"font": ("Arial", 10), "pady": 2, "padx": 20, "anchor": "w"}

    # Frame principal com scroll
    main_frame = tk.Frame(janela_atalhos)
    main_frame.pack(fill="both", expand=True, padx=10, pady=4)

    # Canvas com scrollbar
    canvas = tk.Canvas(main_frame)
    scrollbar = tk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
    scrollable_frame = tk.Frame(canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # Categorias de atalhos
    categorias = {
        "Navegação": [
            ("Setas ↑/↓", "Navegar entre registros na tabela"),
            ("Tab / Shift+Tab", "Navegar entre campos"),
            ("Esc", "Limpar campos ou fechar janelas"),
        ],
        "Seleção": [
            ("Ctrl+A", "Selecionar todos os registros"),
            ("Ctrl+T", "Selecionar todos os registros (alternativo)"),
            ("Clique + Shift+Clique", "Selecionar intervalo de registros"),
            ("Ctrl+Clique", "Selecionar múltiplos registros"),
        ],
        "Edição": [
            ("Ctrl+C", "Copiar texto selecionado"),
            ("Ctrl+X", "Recortar texto selecionado"),
            ("Ctrl+V", "Colar texto"),
            ("Ctrl+BackSpace", "Apagar palavra anterior"),
            ("Ctrl+Delete", "Apagar palavra seguinte"),
        ],
        "Busca e Salvamento": [
            ("Ctrl+F", "Buscar"),
            ("Ctrl+S", "Salvar alterações"),
            ("F5", "Atualizar dados"),
            ("F1", "Mostrar este guia de atalhos"),
        ],
        "Tabela": [
            ("Duplo clique", "Editar registro"),
            ("Botão direito", "Abrir menu de contexto"),
        ]
    }

    row = 0
    for categoria, atalhos in categorias.items():
        # Título da categoria
        tk.Label(scrollable_frame, text=categoria, **estilo_titulo).grid(
            row=row, column=0, sticky="w", columnspan=2
        )
        row += 1

        # Linha separadora
        separator = tk.Frame(scrollable_frame, height=1, bg="gray")
        separator.grid(row=row, column=0, sticky="ew", columnspan=2, padx=5, pady=5)
        row += 1

        # Atalhos da categoria
        for atalho, descricao in atalhos:
            tk.Label(scrollable_frame, text=atalho, width=15, **estilo_atalho).grid(
                row=row, column=0, sticky="w"
            )
            tk.Label(scrollable_frame, text=descricao, **estilo_atalho).grid(
                row=row, column=1, sticky="w"
            )
            row += 1

        # Espaço entre categorias
        tk.Label(scrollable_frame, text="", pady=5).grid(row=row, column=0)
        row += 1

    # Botão fechar
    btn_fechar = tk.Button(janela_atalhos, text="Fechar", command=janela_atalhos.destroy)
    btn_fechar.pack(pady=10)

    # Centraliza a janela
    janela_atalhos.update_idletasks()
    largura = janela_atalhos.winfo_width()
    altura = janela_atalhos.winfo_height()
    x = (janela_atalhos.winfo_screenwidth() // 2) - (largura // 2)
    y = (janela_atalhos.winfo_screenheight() // 2) - (altura // 2)
    janela_atalhos.geometry(f"{largura}x{altura}+{x}+{y}")

    # Atalhos de teclado
    janela_atalhos.bind("<Escape>", lambda e: janela_atalhos.destroy())


janela_visualizar_dia_ativa = None
ultimo_clique_visualizar = 0


# --- Classes Utilitárias --- #

class Logger:
    """Sistema de logging estruturado."""

    @staticmethod
    def configurar():
        """Configura o sistema de logging."""
        logging.basicConfig(
            filename='app.log',
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )

    @staticmethod
    def info(mensagem, contexto=None):
        """Registra mensagem informativa."""
        contexto_str = f"[{contexto}]" if contexto else ""
        print(f"[INFO] {contexto_str} {mensagem}")
        logging.info(f"{contexto_str} {mensagem}")

    @staticmethod
    def erro(mensagem, erro=None, contexto=None):
        """Registra mensagem de erro."""
        contexto_str = f"[{contexto}]" if contexto else ""
        erro_str = f": {str(erro)}" if erro else ""
        print(f"[ERRO] {contexto_str} {mensagem}{erro_str}")
        logging.error(f"{contexto_str} {mensagem}{erro_str}")

    @staticmethod
    def aviso(mensagem, contexto=None):
        """Registra mensagem de aviso."""
        contexto_str = f"[{contexto}]" if contexto else ""
        print(f"[AVISO] {contexto_str} {mensagem}")
        logging.warning(f"{contexto_str} {mensagem}")


class DatabaseManager:
    """Gerencia consultas SQLite sem manter conexão global."""

    def __init__(self, caminho_banco):
        """Inicializa o gerenciador de banco de dados.

        Args:
            caminho_banco: Caminho para o arquivo do banco de dados SQLite.
        """
        self.caminho_banco = caminho_banco

    def conectar(self):
        """Cria e retorna uma nova conexão com o banco de dados."""
        try:
            return sqlite3.connect(
                self.caminho_banco,
                detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES
            )
        except sqlite3.Error as e:
            Logger.erro(f"Erro ao conectar ao banco de dados", e, "DatabaseManager")
            raise

    def executar(self, query, params=(), fetch=False):
        """Executa query SQL com commit automático.

        Args:
            query: String SQL a ser executada.
            params: Parâmetros para a query (opcional).
            fetch: Se True, retorna os resultados da query.

        Returns:
            Resultados da query se fetch=True, None caso contrário.
        """
        try:
            with closing(self.conectar()) as conn:
                cursor = conn.cursor()
                cursor.execute(query, params)
                conn.commit()
                if fetch:
                    return cursor.fetchall()
        except sqlite3.Error as e:
            Logger.erro(f"Erro ao executar query: {query}", e, "DatabaseManager")
            raise

    def executar_muitos(self, query, params_list):
        """Executa múltiplas queries em uma transação.

        Args:
            query: String SQL a ser executada.
            params_list: Lista de tuplas de parâmetros.
        """
        try:
            with closing(self.conectar()) as conn:
                cursor = conn.cursor()
                cursor.executemany(query, params_list)
                conn.commit()
        except sqlite3.Error as e:
            Logger.erro(f"Erro ao executar queries em lote", e, "DatabaseManager")
            raise

    def obter_um(self, query, params=()):
        """Executa query e retorna apenas o primeiro resultado.

        Args:
            query: String SQL a ser executada.
            params: Parâmetros para a query (opcional).

        Returns:
            Primeiro resultado da query ou None se não houver resultados.
        """
        resultados = self.executar(query, params, fetch=True)
        return resultados[0] if resultados else None


def verificar_proximo_foco(event, campo_desejado):
    """Verifica se o próximo campo a receber o foco é o desejado, caso contrário, força o foco para ele."""
    # Obtém o widget que está recebendo o foco após o evento FocusOut
    widget_atual = event.widget

    # Se o próximo widget não for o campo desejado, força o foco para o campo desejado
    janela.after(1, lambda: campo_desejado.focus_set())
    return "break"


def on_tab_listbox(event):
    """Fallback para Listboxes que usam <Tab>.
    Se for uma listbox de Autocomplete, delega para o entry; caso contrário avança o foco."""
    lb = event.widget
    try:
        tl = lb.winfo_toplevel()
        parent = getattr(tl, "master", None)  # tenta achar o Entry dono
        if isinstance(parent, AutocompleteEntry):
            return parent.on_tab(event)
    except Exception:
        pass

    try:
        event.widget.tk_focusNext().focus()
    except Exception:
        pass
    return "break"


class SimpleCache:
    """Sistema de cache com TTL configurável."""

    def __init__(self):
        """Inicializa o sistema de cache."""
        self.data = {}
        self.timestamps = {}

    def get(self, key):
        """Obtém um valor do cache se ainda for válido.

        Args:
            key: Chave do valor no cache.

        Returns:
            Valor armazenado ou None se não existir ou estiver expirado.
        """
        if key in self.data:
            # Verifica se o cache ainda é válido
            if time.time() - self.timestamps[key] < CACHE_TTL_DEFAULT:
                return self.data[key]
            else:
                # Cache expirado, remove
                self.invalidate(key)
        return None

    def set(self, key, value, ttl=CACHE_TTL_DEFAULT):
        """Armazena um valor no cache.

        Args:
            key: Chave para armazenar o valor.
            value: Valor a ser armazenado.
            ttl: Tempo de vida em segundos (padrão: CACHE_TTL_DEFAULT).
        """
        self.data[key] = value
        self.timestamps[key] = time.time()

    def invalidate(self, key):
        """Remove um item específico do cache.

        Args:
            key: Chave do item a ser removido.
        """
        if key in self.data:
            del self.data[key]
        if key in self.timestamps:
            del self.timestamps[key]

    def invalidate_pattern(self, pattern):
        """Invalida todas as chaves que correspondem ao padrão regex.

        Args:
            pattern: Padrão regex para corresponder às chaves.
        """
        keys_to_invalidate = []
        for key in list(self.data.keys()):
            if re.search(pattern, str(key)):
                keys_to_invalidate.append(key)

        for key in keys_to_invalidate:
            self.invalidate(key)

    def clear(self):
        """Limpa todo o cache."""
        self.data.clear()
        self.timestamps.clear()


# Instância global do cache
cache = SimpleCache()


class AutocompleteEntry(tk.Entry):
    """Widget de entrada com autocompletar."""

    def __init__(self, master, completion_list,
                 listbox_x_offset=-14, listbox_y_offset=-2,
                 listbox_width=300, listbox_max_height=10,
                 *args, **kwargs):
        """Inicializa o widget de entrada com autocompletar.

        Args:
            master: Widget pai.
            completion_list: Lista de itens para autocompletar.
            listbox_x_offset: Deslocamento horizontal da listbox.
            listbox_y_offset: Deslocamento vertical da listbox.
            listbox_width: Largura da listbox.
            listbox_max_height: Altura máxima da listbox em número de itens.
        """
        super().__init__(master, *args, **kwargs)
        self.completion_list = sorted(completion_list, key=lambda x: str(x).lower())

        self.listbox = None  # referência ao Listbox (widget)
        self.listbox_window = None  # Toplevel que contém o Listbox (popup)
        self.listbox_max_height = listbox_max_height
        self._custom_listbox_settings = {
            'x_offset': listbox_x_offset,
            'y_offset': listbox_y_offset,
            'width': listbox_width,
            'max_height': listbox_max_height
        }

        # Configuração de eventos
        self._configurar_eventos()

    def _configurar_eventos(self):
        """Configura todos os eventos do widget."""
        self.bind('<KeyRelease>', self.on_keyrelease)
        self.bind('<FocusOut>', self._on_focus_out)
        self.bind('<Up>', self.on_arrow_up)
        self.bind('<Down>', self.on_arrow_down)
        self.bind('<Escape>', self.limpar_campo)
        self.bind('<Return>', self.on_enter)  # Enter
        self.bind('<Tab>', self.on_tab)  # Tab

    def _on_focus_out(self, event):
        """Espera um pouquinho e só esconde se o foco tk.NÃO for a popup (listbox)."""

        def check_focus():
            focused = self.focus_get()
            try:
                if focused and self.listbox_window and hasattr(focused,
                                                               "winfo_toplevel") and focused.winfo_toplevel() == self.listbox_window:
                    # foco foi para a popup -> tk.NÃO fechar
                    return
            except Exception as e:
                Logger.erro("Erro ao verificar foco", e, "AutocompleteEntry")
            # caso contrário, esconde
            self.hide_suggestions()

        # pequena espera para permitir que o click no listbox ou a tecla seja processada
        self.after(50, check_focus)

    def set_listbox_properties(self, x_offset=None, y_offset=None, width=None, max_height=None):
        """Configura as propriedades da listbox.

        Args:
            x_offset: Deslocamento horizontal da listbox.
            y_offset: Deslocamento vertical da listbox.
            width: Largura da listbox.
            max_height: Altura máxima da listbox em número de itens.
        """
        if not hasattr(self, '_custom_listbox_settings'):
            self._custom_listbox_settings = {}
        if x_offset is not None:
            self._custom_listbox_settings['x_offset'] = x_offset
        if y_offset is not None:
            self._custom_listbox_settings['y_offset'] = y_offset
        if width is not None:
            self._custom_listbox_settings['width'] = width
        if max_height is not None:
            self._custom_listbox_settings['max_height'] = max_height
            self.listbox_max_height = max_height

    def aceitar_item(self, idx=None):
        """Insere o item selecionado no Entry. Retorna True se inseriu algo.

        Args:
            idx: Índice do item a ser aceito (opcional).

        Returns:
            True se um item foi inserido, False caso contrário.
        """
        if self.listbox and self.listbox.size() > 0:
            if idx is None:
                sel = self.listbox.curselection()
                idx = sel[0] if sel else 0
            try:
                value = self.listbox.get(idx)
            except Exception as e:
                Logger.erro("Erro ao obter item da listbox", e, "AutocompleteEntry")
                value = None
            if value:
                self.delete(0, tk.END)
                self.insert(0, value)
                self.config(bg=COR_FEEDBACK_POSITIVO)  # feedback visual
                return True
        return False

    def on_enter(self, event):
        """
        Enter: se a listbox estiver aberta, aceita o item selecionado,
        fecha as sugestões e mantém foco no campo.
        Caso contrário, permite que o Enter faça o comportamento normal.

        Args:
            event: Evento de teclado.

        Returns:
            "break" se o evento foi tratado, None caso contrário.
        """
        if not self.listbox:
            return None  # sem lista, segue com o evento normal

        inserted = False
        try:
            inserted = self.aceitar_item()
        except Exception as e:
            Logger.erro("Erro ao aceitar item", e, "AutocompleteEntry")
            inserted = False

        # fecha popup se estava aberto
        self.hide_suggestions()

        if inserted:
            # mantém foco no campo (útil para edição/confirmar)
            try:
                self.focus_set()
            except Exception as e:
                Logger.erro("Erro ao definir foco", e, "AutocompleteEntry")
            return "break"  # evita que Enter dispare outras ações
        return None

    def on_tab(self, event):
        """
        Tab: se a listbox estiver aberta, aceita o item selecionado, fecha a
        listbox e AVANÇA o foco para o próximo widget.
        Se a listbox tk.NÃO estiver aberta, permite o comportamento normal do Tab.

        Args:
            event: Evento de teclado.

        Returns:
            "break" se o evento foi tratado, None caso contrário.
        """
        if not self.listbox:
            return None  # permite Tab normal

        inserted = False
        try:
            inserted = self.aceitar_item()
        except Exception as e:
            Logger.erro("Erro ao aceitar item", e, "AutocompleteEntry")
            inserted = False

        # fecha popup
        self.hide_suggestions()

        # Sempre avança o foco ao próximo, mesmo se não houver inserção explícita
        try:
            next_widget = getattr(self, 'tab_next', None)
            if not next_widget:
                next_widget = self.tk_focusNext()
            if next_widget:
                next_widget.focus_set()
        except Exception as e:
            Logger.erro("Erro ao avançar foco", e, "AutocompleteEntry")
        return "break"

    def on_listbox_click(self, event):
        """Clique/duplo clique com mouse aceita o item.

        Args:
            event: Evento de mouse.

        Returns:
            "break" para interromper o processamento do evento.
        """
        if self.listbox:
            idx = self.listbox.nearest(event.y)
            if idx >= 0:
                try:
                    value = self.listbox.get(idx)
                except Exception as e:
                    Logger.erro("Erro ao obter item da listbox", e, "AutocompleteEntry")
                    value = None
                if value:
                    self.delete(0, tk.END)
                    self.insert(0, value)
                    self.config(bg=COR_FEEDBACK_POSITIVO)
        self.hide_suggestions()
        # devolve foco ao entry
        try:
            self.focus_set()
        except Exception as e:
            Logger.erro("Erro ao definir foco", e, "AutocompleteEntry")
        return "break"

    def _on_listbox_select_event(self, event):
        """Atualiza o campo quando uma seleção é feita na listbox (visual sync).

        Args:
            event: Evento de seleção.
        """
        if self.listbox:
            sel = self.listbox.curselection()
            if sel:
                idx = sel[0]
                try:
                    value = self.listbox.get(idx)
                    if value:
                        self.delete(0, tk.END)
                        self.insert(0, value)
                except Exception as e:
                    Logger.erro("Erro ao atualizar campo", e, "AutocompleteEntry")

    def show_suggestions(self, matches):
        """Mostra sugestões em uma listbox popup.

        Args:
            matches: Lista de itens para mostrar.
        """
        # Limpa popup anterior
        self.hide_suggestions()

        # Calcula posição relativa à tela
        x = self.winfo_rootx() + self._custom_listbox_settings.get('x_offset', -14)
        y = self.winfo_rooty() + self.winfo_height() + self._custom_listbox_settings.get('y_offset', -2)
        width = self._custom_listbox_settings.get('width', 300)

        # Cria janela popup
        self.listbox_window = tk.Toplevel(self)
        self.listbox_window.withdraw()
        self.listbox_window.overrideredirect(True)
        self.listbox_window.configure(bg='white', bd=1, relief='solid')

        # Cria e configura a listbox
        self.listbox = tk.Listbox(
            self.listbox_window,
            font=self['font'],
            activestyle='none',
            exportselection=False,
            selectmode=tk.SINGLE,
            bg='white',
            fg='black',
            selectbackground=COR_DESTAQUE,
            selectforeground='white',
            relief='flat'
        )
        self.listbox.pack(expand=True, fill=tk.BOTH)

        # Botão Selecionar Todos
        self.botao_selecionar_todos = tk.Button(
            self.listbox_window,
            text="Selecionar Todos",
            bg=COR_DESTAQUE,
            fg="white",
            relief="flat",
            cursor="hand2",
            font=("Segoe UI", 9, "bold"),
            command=lambda: self.listbox.select_set(0, tk.END)
        )
        self.botao_selecionar_todos.pack(fill="x", padx=2, pady=(2, 2))

        # Adiciona os itens (limitando por listbox_max_height)
        for item in matches[:self.listbox_max_height]:
            self.listbox.insert(tk.END, item)

        # Ajusta altura e posição
        self.adjust_listbox_height()
        height = self.listbox.winfo_reqheight()
        self.listbox_window.geometry(f"{width}x{height}+{x}+{y}")

        # Configura topmost e mostra
        try:
            self.listbox_window.attributes("-topmost", True)
        except Exception as e:
            Logger.erro("Erro ao configurar topmost", e, "AutocompleteEntry")
        self.listbox_window.deiconify()

        # Binds na listbox
        self.listbox.bind("<<ListboxSelect>>", self._on_listbox_select_event)
        self.listbox.bind('<ButtonRelease-1>', self.on_listbox_click)  # clique
        self.listbox.bind('<Double-Button-1>', self.on_listbox_click)  # duplo clique
        self.listbox.bind('<Return>', self.on_enter)  # Enter
        self.listbox.bind('<Tab>', self.on_tab)  # Tab
        self.listbox.bind('<Escape>', self.limpar_campo)  # ESC
        self.listbox.bind('<Up>', self.on_arrow_up)  # seta ↑
        self.listbox.bind('<Down>', self.on_arrow_down)  # seta ↓

        # Seleciona o primeiro item por padrão
        if self.listbox.size() > 0:
            self.listbox.selection_set(0)
            self.listbox.activate(0)

    def adjust_listbox_height(self):
        """Ajusta a altura da listbox com base no número de itens."""
        if self.listbox:
            items_count = self.listbox.size()
            height = min(items_count, self._custom_listbox_settings.get('max_height', self.listbox_max_height))
            self.listbox.config(height=height)

    def hide_suggestions(self, event=None):
        """Esconde a listbox de sugestões.

        Args:
            event: Evento (opcional).
        """
        if self.listbox_window:
            try:
                self.listbox_window.destroy()
            except Exception as e:
                Logger.erro("Erro ao destruir listbox_window", e, "AutocompleteEntry")
        self.listbox_window = None
        self.listbox = None

    def on_arrow_up(self, event):
        """Manipula o evento de seta para cima.

        Args:
            event: Evento de teclado.

        Returns:
            "break" se o evento foi tratado, None caso contrário.
        """
        if self.listbox:
            current = self.listbox.curselection()
            if current:
                idx = max(0, current[0] - 1)
            else:
                idx = max(0, self.listbox.size() - 1)
            self.listbox.selection_clear(0, tk.END)
            self.listbox.selection_set(idx)
            self.listbox.activate(idx)
            return "break"
        return None

    def on_arrow_down(self, event):
        """Manipula o evento de seta para baixo.

        Args:
            event: Evento de teclado.

        Returns:
            "break" se o evento foi tratado, None caso contrário.
        """
        if self.listbox:
            current = self.listbox.curselection()
            if current:
                idx = (current[0] + 1) % self.listbox.size()
            else:
                idx = 0
            self.listbox.selection_clear(0, tk.END)
            self.listbox.selection_set(idx)
            self.listbox.activate(idx)
            return "break"
        else:
            # se não há listbox, tentar abrir sugestões
            valor = self.get().strip().upper()
            if not valor:
                matches = self.completion_list
            else:
                valor_sem_acento = remover_acentos(valor)
                matches = [
                    item for item in self.completion_list
                    if valor_sem_acento in remover_acentos(str(item).upper())
                ]
            if matches:
                self.show_suggestions(matches)
                return "break"
        return None

    def on_keyrelease(self, event):
        """Manipula o evento de liberação de tecla.

        Args:
            event: Evento de teclado.
        """
        # evita modificar se estiver readonly
        if event.keysym in ("Up", "Down", "Return", "Escape", "Tab", "Shift_L", "Shift_R", "Control_L", "Control_R"):
            return

        valor = self.get().strip().upper()
        valor_sem_acento = remover_acentos(valor)
        pos = self.index(tk.INSERT)
        # preserva posição do cursor
        self.delete(0, tk.END)
        self.insert(0, valor)
        self.icursor(pos)

        if valor == '':
            self.hide_suggestions()
            return

        matches = [
            item for item in self.completion_list
            if valor_sem_acento in remover_acentos(str(item).upper())
        ]
        if matches:
            self.show_suggestions(matches)
        else:
            self.hide_suggestions()

    def force_hide_suggestions(self):
        """Força o fechamento das sugestões independentemente do estado."""
        if self.listbox_window:
            try:
                self.listbox_window.destroy()
            except Exception as e:
                Logger.erro("Erro ao destruir listbox_window", e, "AutocompleteEntry")
        self.listbox_window = None
        self.listbox = None

    def limpar_campo(self, event=None):
        """Limpa o conteúdo do campo.

        Args:
            event: Evento (opcional).

        Returns:
            "break" para interromper o processamento do evento.
        """
        self.config(state='normal')
        self.delete(0, tk.END)
        self.hide_suggestions()
        return "break"

    @staticmethod
    def ativar_edicao(event=None):
        """Ativa a edição do campo.

        Args:
            event: Evento (opcional).
        """
        widget = event.widget
        widget.config(state='normal')
        if isinstance(widget, tk.Entry):
            widget.icursor(tk.END)
        try:
            item_selecionado = tabela.focus()
            processo = tabela.item(item_selecionado)['values']
            numero_original = str(processo[1]).strip().upper()
            botao_cadastrar.config(
                text="Atualizar",
                state='normal',
                command=lambda: atualizar_processo(numero_original)
            )
        except Exception as e:
            Logger.erro("Erro ao ativar edição", e, "AutocompleteEntry")


# Completar a função
def atualizar_campo_contratado(event):
    cursor_pos = event.widget.index(tk.INSERT)
    texto_atual = event.widget.get()
    texto_maiusculo = texto_atual.upper()

    if texto_atual != texto_maiusculo:
        event.widget.delete(0, tk.END)
        event.widget.insert(0, texto_maiusculo)
        event.widget.icursor(cursor_pos)

    # Mostrar sugestões ao digitar usando apenas nomes salvos em "Contratado"
    if texto_maiusculo:
        valor_sem_acento = remover_acentos(texto_maiusculo)
        matches = [
            item for item in nomes_contratado
            if valor_sem_acento in remover_acentos(str(item).upper())
        ]
        if matches:
            entrada_contratado.show_suggestions(matches)
        else:
            entrada_contratado.hide_suggestions()
    else:
        # Se o campo estiver vazio, esconde as sugestões
        entrada_contratado.hide_suggestions()

    ativar_botao_atualizar()
    if botao_cadastrar.cget('text') == 'Atualizar':
        botao_cadastrar.config(state='normal')


def remover_acentos(txt):
    return ''.join(
        c for c in unicodedata.normalize('NFD', txt)
        if unicodedata.category(c) != 'Mn'
    )


def salvar_larguras_colunas(tabela, colunas, caminho_arquivo='config_colunas.json'):
    larguras = {col: tabela.column(col)['width'] for col in colunas}
    with open(caminho_arquivo, 'w') as f:
        json.dump(larguras, f)


def carregar_larguras_colunas(tabela, colunas, caminho_arquivo='config_colunas.json'):
    if os.path.exists(caminho_arquivo):
        with open(caminho_arquivo, 'r') as f:
            larguras = json.load(f)
            for col in colunas:
                if col in larguras:
                    tabela.column(col, width=larguras[col])


def carregar_tamanho_janela(nome: str, default_w: int = 500, default_h: int = 450,
                            caminho_arquivo: str = 'config_windows.json'):
    try:
        if os.path.exists(caminho_arquivo):
            with open(caminho_arquivo, 'r') as f:
                data = json.load(f) or {}
            conf = data.get(str(nome))
            if isinstance(conf, dict):
                w = int(conf.get('width', default_w))
                h = int(conf.get('height', default_h))
                x = conf.get('x')
                y = conf.get('y')
                xi = int(x) if x is not None else None
                yi = int(y) if y is not None else None
                return w, h, xi, yi
    except Exception:
        pass
    return default_w, default_h, None, None


def salvar_tamanho_janela(nome: str, width: int, height: int, x: int = None, y: int = None,
                          caminho_arquivo: str = 'config_windows.json'):
    try:
        data = {}
        if os.path.exists(caminho_arquivo):
            with open(caminho_arquivo, 'r') as f:
                try:
                    data = json.load(f) or {}
                except Exception:
                    data = {}
        entry = {'width': int(width), 'height': int(height)}
        if x is not None and y is not None:
            entry.update({'x': int(x), 'y': int(y)})
        data[str(nome)] = entry
        with open(caminho_arquivo, 'w') as f:
            json.dump(data, f)
    except Exception:
        pass


def bind_persist_geometry(nome: str, window: tk.Tk):
    state = {'after_id': None}

    def on_configure(event=None):
        try:
            if state['after_id']:
                window.after_cancel(state['after_id'])
        except Exception:
            pass
        try:
            state['after_id'] = window.after(800, lambda: salvar_tamanho_janela(nome, window.winfo_width(),
                                                                                window.winfo_height(), window.winfo_x(),
                                                                                window.winfo_y()))
        except Exception:
            pass

    try:
        window.bind('<Configure>', on_configure)
    except Exception:
        pass


def existe_config_janela(nome: str, caminho_arquivo: str = 'config_windows.json') -> bool:
    try:
        if os.path.exists(caminho_arquivo):
            with open(caminho_arquivo, 'r') as f:
                data = json.load(f) or {}
            return str(nome) in data
    except Exception:
        pass
    return False


def carregar_nomes_autocomplete():
    """Carrega nomes únicos para autocompletar a partir do banco de dados.
    Otimizado para usar uma única consulta SQL.
    """
    cursor.execute(
        """SELECT DISTINCT nome FROM (
            SELECT entregue_por AS nome FROM trabalhos_realizados WHERE entregue_por IS NOT NULL AND entregue_por != ''
            UNION
            SELECT devolvido_a AS nome FROM trabalhos_realizados WHERE devolvido_a IS NOT NULL AND devolvido_a != ''
        ) ORDER BY UPPER(nome)"""
    )
    return [row[0].upper() for row in cursor.fetchall()]


# Lista exclusiva para o campo Contratado

def carregar_nomes_contratado():
    cursor.execute(
        """SELECT DISTINCT UPPER(contratado) AS nome
            FROM trabalhos_realizados
            WHERE contratado IS NOT NULL AND contratado != ''
            ORDER BY UPPER(contratado)"""
    )
    return [row[0] for row in cursor.fetchall()]


def recarregar_listas_autocomplete():
    global nomes_autocomplete, nomes_contratado
    try:
        nomes_autocomplete = carregar_nomes_autocomplete()
        nomes_contratado = carregar_nomes_contratado()
        try:
            entrada_entregue_por.completion_list = nomes_autocomplete
            entrada_devolvido_a.completion_list = nomes_autocomplete
            entrada_contratado.completion_list = nomes_contratado
        except Exception:
            pass
        try:
            cache.invalidate('nomes_autocomplete')
            cache.invalidate('nomes_contratado')
        except Exception:
            pass
    except Exception:
        pass

def exportar_banco():
    global conn, cursor, registros_exportados
    caminho_atual = caminho_banco
    # Adiciona horário e data atual ao nome do arquivo padrão
    data_hora_atual = datetime.now()
    nome_padrao = f"banco_dados_{data_hora_atual.strftime('%H%M%S_%d%m%Y')}.db"
    destino = filedialog.asksaveasfilename(
        defaultextension=".db",
        filetypes=[("Banco de Dados SQLite", "*.db")],
        title="Exportar banco de dados",
        initialfile=nome_padrao
    )
    if destino:
        try:
            conn.commit()
            conn.close()
            shutil.copy2(caminho_atual, destino)
            # Reabre conexão após exportar
            conn = sqlite3.connect(
                caminho_banco,
                detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES
            )
            cursor = conn.cursor()
            try:
                conn.execute("PRAGMA journal_mode=WAL")
                conn.execute("PRAGMA synchronous=NORMAL")
                conn.execute("PRAGMA foreign_keys=ON")
                conn.execute("PRAGMA busy_timeout=5000")
                conn.execute("PRAGMA temp_store=MEMORY")
                conn.execute("PRAGMA cache_size=-32000")
            except Exception:
                pass
            registros_exportados += 1
            atualizar_estatisticas()
            messagebox.showinfo("Sucesso", f"Banco exportado para:\n{destino}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar banco:\n{e}")


def importar_banco():
    global conn, cursor, registros_importados
    origem = filedialog.askopenfilename(
        filetypes=[("Banco de Dados SQLite", "*.db")],
        title="Importar banco de dados"
    )
    if origem:
        if not messagebox.askyesno("Confirmar", "Isto irá substituir TODOS os dados atuais. Continuar?"):
            return
        try:
            conn.close()
            shutil.copy2(origem, caminho_banco)
            conn = sqlite3.connect(
                caminho_banco,
                detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES
            )
            cursor = conn.cursor()
            try:
                conn.execute("PRAGMA journal_mode=WAL")
                conn.execute("PRAGMA synchronous=NORMAL")
                conn.execute("PRAGMA foreign_keys=ON")
                conn.execute("PRAGMA busy_timeout=5000")
                conn.execute("PRAGMA temp_store=MEMORY")
                conn.execute("PRAGMA cache_size=-32000")
            except Exception:
                pass
            listar_processos()
            contar_registros()
            registros_importados += 1
            atualizar_estatisticas()
            messagebox.showinfo("Sucesso", "Banco importado com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao importar banco:\n{e}")


def abrir_janela_banco_dados():
    """Abre uma janela com opções para exportar ou importar banco de dados."""
    janela_banco = Toplevel(janela)
    janela_banco.withdraw()  # Esconde a janela inicialmente para evitar o "piscar"
    janela_banco.title("Gerenciamento do Banco de Dados")
    try:
        w, h, xi, yi = carregar_tamanho_janela('banco_dados', 336, 320)
        if xi is not None and yi is not None:
            janela_banco.geometry(f"{w}x{h}+{xi}+{yi}")
        else:
            center_window(janela_banco, w, h)
    except Exception:
        center_window(janela_banco, 336, 320)
    janela_banco.configure(bg="#F5F7FA")
    janela_banco.resizable(True, True)

    # Centralizar a janela
    janela_banco.transient(janela)
    janela_banco.grab_set()

    # Já centralizado acima com o tamanho padrão salvo

    # Frame principal com borda
    frame_principal = tk.Frame(janela_banco, bg="#F5F7FA", relief="flat", bd=0)
    frame_principal.pack(expand=True, fill="both", padx=20, pady=10)  # Reduzido padding

    # Cabeçalho com ícone e título
    frame_cabecalho = tk.Frame(frame_principal, bg="#F5F7FA")
    frame_cabecalho.pack(fill="x", pady=(0, 12))  # Reduzido espaçamento

    # Título principal
    label_titulo = tk.Label(frame_cabecalho, text="🗄️ Gerenciamento do Banco de Dados",
                            font=("Segoe UI", 12, "bold"), bg="#F5F7FA", fg="#2C3E50")  # Fonte menor
    label_titulo.pack()

    # Subtítulo
    label_subtitulo = tk.Label(frame_cabecalho, text="Faça backup ou restaure seus dados com segurança",
                               font=("Segoe UI", 8), bg="#F5F7FA", fg="#7F8C8D")  # Fonte menor
    label_subtitulo.pack(pady=(3, 0))  # Reduzido espaçamento

    # Separador visual
    separador = tk.Frame(frame_principal, height=2, bg="#E8EBF0")
    separador.pack(fill="x", pady=(0, 10))  # Reduzido espaçamento

    # Frame dos botões com espaçamento melhorado
    frame_botoes = tk.Frame(frame_principal, bg="#F5F7FA")
    frame_botoes.pack(expand=True, fill="both")

    # Botão Exportar Banco - tamanho padrão
    botao_exportar = tk.Button(frame_botoes, text="📤 Exportar Banco",
                               command=lambda: [exportar_banco(), _fechar_banco()],
                               width=14,
                               bg="#4A90E2", fg="white", activebackground="#357ABD", activeforeground="white",
                               relief=tk.FLAT, bd=0, font=("Segoe UI", 10, "bold"))
    botao_exportar.config(padx=15, pady=8)
    botao_exportar.pack(pady=(8, 6))  # Reduzido espaçamento

    # Descrição do botão exportar
    desc_exportar = tk.Label(frame_botoes, text="Salva uma cópia do banco de dados atual",
                             font=("Segoe UI", 7), bg="#F5F7FA", fg="#95A5A6")  # Fonte menor
    desc_exportar.pack(pady=(0, 10))  # Reduzido espaçamento

    # Botão Importar Banco - MESMO tamanho do Exportar Banco
    botao_importar = tk.Button(frame_botoes, text="📥 Importar Banco",
                               command=lambda: [importar_banco(), _fechar_banco()],
                               width=14,
                               bg="#4A90E2", fg="white", activebackground="#357ABD", activeforeground="white",
                               relief=tk.FLAT, bd=0, font=("Segoe UI", 10, "bold"))
    botao_importar.config(padx=15, pady=8)
    botao_importar.pack(pady=(4, 6))  # Reduzido espaçamento

    # Descrição do botão importar
    desc_importar = tk.Label(frame_botoes, text="Restaura dados de um arquivo de backup",
                             font=("Segoe UI", 7), bg="#F5F7FA", fg="#95A5A6")  # Fonte menor
    desc_importar.pack(pady=(0, 8))  # Reduzido espaçamento

    # Removido rodapé com texto de instrução para ESC

    # Efeitos hover usando as cores padrão do programa
    def on_enter_exportar(e):
        botao_exportar.config(bg='#357ABD')

    def on_leave_exportar(e):
        botao_exportar.config(bg='#4A90E2')

    def on_enter_importar(e):
        botao_importar.config(bg='#357ABD')

    def on_leave_importar(e):
        botao_importar.config(bg='#4A90E2')

    botao_exportar.bind("<Enter>", on_enter_exportar)
    botao_exportar.bind("<Leave>", on_leave_exportar)
    botao_importar.bind("<Enter>", on_enter_importar)
    botao_importar.bind("<Leave>", on_leave_importar)

    def _fechar_banco():
        try:
            janela_banco.update_idletasks()
            salvar_tamanho_janela('banco_dados', janela_banco.winfo_width(), janela_banco.winfo_height(),
                                  janela_banco.winfo_x(), janela_banco.winfo_y())
        except Exception:
            pass
        janela_banco.destroy()

    # Bind para fechar com ESC e protocolo de fechar
    janela_banco.bind("<Escape>", lambda e: _fechar_banco())
    try:
        janela_banco.protocol("WM_DELETE_WINDOW", _fechar_banco)
    except Exception:
        pass

    # Focar no primeiro botão
    botao_exportar.focus_set()

    # Persistir tamanho/posição durante redimensionamento/movimento
    try:
        bind_persist_geometry('banco_dados', janela_banco)
    except Exception:
        pass

    # Reaplica geometria salva após montar UI e então mostra
    try:
        if xi is not None and yi is not None:
            janela_banco.geometry(f"{w}x{h}+{xi}+{yi}")
        else:
            center_window(janela_banco, w, h)
    except Exception:
        pass
    janela_banco.deiconify()


def handle_escape(event=None):
    limpar_campos()


#     return "break"  # Isso impede que o evento ESC propague para outros handlers

def visualizar_processo(event=None):
    print("🔍 visualizar_processo chamada!")

    # ✅ MODIFICAÇÃO: Avisar mas continuar permitindo visualização
    if botao_cadastrar.cget('text') == 'Atualizar':
        print("⚠️ Modo atualização ativo - mudando para novo registro")

    item_selecionado = tabela.focus()
    print(f"📋 Item selecionado: {item_selecionado}")

    if not item_selecionado:
        print("❌ Nenhum item selecionado - limpando campos")
        limpar_campos()
        return

    processo = tabela.item(item_selecionado)['values']
    print(f"📊 Dados do processo: {processo}")

    # ✅ CORREÇÃO: Configurar todos os campos como editáveis inicialmente
    campos = [
        entrada_numero, entrada_secretaria, entrada_licitacao, entrada_modalidade,
        entrada_recebimento, entrada_devolucao, entrada_entregue_por,
        entrada_devolvido_a, entrada_contratado
    ]

    for campo in campos:
        campo.config(state='normal')

    # ✅ NÃO DESABILITA mais a caixa de texto de observações aqui

    try:
        # Número do processo
        numero_processo = str(processo[1])
        entrada_numero.delete(0, tk.END)
        entrada_numero.insert(0, numero_processo)
        entrada_numero.config(state='readonly')
        print(f"✅ Número preenchido: {numero_processo}")

        # Secretaria
        sigla_secretaria = processo[2]
        secretaria_formatada = next(
            (s for s in secretarias_formatadas if s.startswith(sigla_secretaria + " - ")),
            ""
        )
        entrada_secretaria.delete(0, tk.END)
        entrada_secretaria.insert(0, secretaria_formatada)
        entrada_secretaria.config(state='readonly')

        # Licitação
        entrada_licitacao.delete(0, tk.END)
        entrada_licitacao.insert(0, processo[3] if processo[3] else "")
        entrada_licitacao.config(state='readonly')

        # Modalidade
        modalidade = processo[5] if len(processo) > 5 else ""
        entrada_modalidade.delete(0, tk.END)
        entrada_modalidade.insert(0, modalidade)
        entrada_modalidade.config(state='readonly')

        # Situação
        situacao_var.set(processo[4])

        # Data de recebimento
        data_inicio = processo[6]
        try:
            data_inicio = DateUtils.para_exibicao(str(data_inicio)) if data_inicio else ""
        except Exception:
            pass
        entrada_recebimento.delete(0, tk.END)
        entrada_recebimento.insert(0, data_inicio if data_inicio else "")
        entrada_recebimento.config(state='readonly')

        # Data de devolução
        data_entrega = processo[7]
        try:
            data_entrega = DateUtils.para_exibicao(str(data_entrega)) if data_entrega else ""
        except Exception:
            pass
        entrada_devolucao.delete(0, tk.END)
        valor_devolucao = (
            "" if (isinstance(data_entrega, str) and data_entrega.strip().lower() == "none")
            else (data_entrega if data_entrega else "")
        )
        entrada_devolucao.insert(0, valor_devolucao)
        entrada_devolucao.config(state='readonly')

        # Entregue por
        entrada_entregue_por.delete(0, tk.END)
        entrada_entregue_por.insert(0, processo[8] if len(processo) > 8 else "")
        entrada_entregue_por.config(state='readonly')

        # Devolvido a
        entrada_devolvido_a.delete(0, tk.END)
        entrada_devolvido_a.insert(0, processo[9] if len(processo) > 9 else "")
        entrada_devolvido_a.config(state='readonly')

        # Contratado
        entrada_contratado.delete(0, tk.END)
        contratado_valor = processo[10] if len(processo) > 10 and processo[10] is not None else ""
        entrada_contratado.insert(0, contratado_valor)
        entrada_contratado.config(state='readonly')

        # ✅ Observações (posição 11) - agora visível e editável
        entrada_descricao.config(state='normal', bg='white')
        entrada_descricao.delete("1.0", tk.END)
        descricao_valor = processo[11] if len(processo) > 11 and processo[11] is not None else ""
        entrada_descricao.insert("1.0", descricao_valor)

        # ✅ Configura o botão como "Atualizar" e habilitado
        botao_cadastrar.config(
            text="Atualizar",
            state='normal',
            command=lambda: atualizar_processo(numero_processo)
        )
        print("✅ Todos os campos preenchidos com sucesso!")

    except Exception as e:
        print(f"❌ Erro ao visualizar processo: {e}")
        messagebox.showerror("Erro", f"Erro ao carregar processo: {e}")

    # Permite cancelar com ESC
    janela.bind("<Escape>", handle_escape)

def ativar_edicao_campo(event):
    widget = event.widget

    # Apenas foca no campo, sem ativar o botão Atualizar automaticamente
    widget.config(state='normal')
    widget.focus_set()
    try:
        if isinstance(widget, tk.Text):
            widget.config(bg="white")
    except Exception:
        pass

    # Só habilita o botão de atualizar se há um registro selecionado na tabela
    try:
        itens_selecionados = tabela.selection()
        if itens_selecionados:
            processo = tabela.item(itens_selecionados[0])['values']
            if processo and len(processo) > 1:
                numero_original = str(processo[1]).strip().upper()
                if numero_original:
                    botao_cadastrar.config(
                        text="Atualizar",
                        state='normal',
                        command=lambda: atualizar_processo(numero_original)
                    )
    except Exception as e:
        print(f"[ERRO] ativar_edicao_campo: {e}")


# Classe ToolTip
class ToolTip:
    def __init__(self, widget):
        self.widget = widget
        self.tip_window = None
        self.label = None
        self.x = self.y = 0
        # Controle de atualização suave
        self.last_update_time = 0
        self.update_interval_ms = 16  # ~60 FPS
        self.min_delta_px = 3

    def show(self, text):
        if not text:
            return

        current_time = time.time() * 1000
        x = self.widget.winfo_pointerx() + 20
        y = self.widget.winfo_pointery() + 20

        if self.tip_window is None:
            self.tip_window = tw = tk.Toplevel(self.widget)
            tw.wm_overrideredirect(True)
            tw.attributes("-topmost", True)
            tw.wm_geometry(f"+{x}+{y}")

            frame = tk.Frame(tw, background="#2C3E50", relief=tk.SOLID, borderwidth=1)
            frame.pack()

            self.label = tk.Label(
                frame,
                text=text,
                justify=tk.LEFT,
                background="#2C3E50",
                foreground="white",
                font=("Segoe UI", "9", "normal"),
                wraplength=400
            )
            self.label.pack(padx=8, pady=6)

            self.x = x
            self.y = y
            self.last_update_time = current_time
        else:
            self.label.config(text=text)

            dx = abs(x - self.x)
            dy = abs(y - self.y)
            if (current_time - self.last_update_time) >= self.update_interval_ms or (dx + dy) >= self.min_delta_px:
                self.x = x
                self.y = y
                self.last_update_time = current_time
                self.tip_window.wm_geometry(f"+{x}+{y}")

    def hide(self):
        """Esconde o tooltip"""
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None
            self.label = None


# --- CONFIGURAÇÕES INICIAIS ---

secretarias_dict = {
    "ASCOM": "Assessoria de Comunicação",
    "AMT": "Autarquia Municipal de Trânsito",
    "CGM": "Controladoria-geral do Município",
    "GABPREF": "Gabinete do Prefeito",
    "GABVICE": "Gabinete do(a) Vice-prefeito(a)",
    "IPM": "Instituto de Previdência do Município de Caucaia",
    "IMAC": "Instituto do Meio Ambiente do Município de Caucaia",
    "OUV": "Ouvidoria-geral do Município",
    "PGM": "Procuradoria-geral do Município",
    "SEAD": "Secretaria de Administração e Recursos Humanos",
    "SEPA": "Secretaria de Proteção Animal",
    "SERJ": "Secretaria Executiva Regional da Grande Jurema",
    "SERL": "Secretaria Executiva Regional do Litoral",
    "SERS": "Secretaria Executiva Regional do Sertão",
    "SECIDT": "Secretaria Municipal de Ciência, Inovação e Desenvolvimento Tecnológico",
    "SECULT": "Secretaria Municipal de Cultura",
    "SEDEC": "Secretaria Municipal de Desenvolvimento Econômico",
    "SEDR": "Secretaria Municipal de Desenvolvimento Rural",
    "SDS": "Secretaria Municipal de Desenvolvimento Social",
    "SME": "Secretaria Municipal de Educação",
    "SEJU": "Secretaria Municipal de Esporte e Juventude",
    "SEFIN": "Secretaria Municipal de Finanças, Planejamento e Orçamento",
    "SEINFRA": "Secretaria Municipal de Infraestrutura",
    "SEPAT": "Secretaria Municipal de Patrimônio e Transporte",
    "SEPLAM": "Secretaria Municipal de Planejamento Urbano e Ambiental",
    "SMS": "Secretaria Municipal de Saúde",
    "SESP": "Secretaria Municipal de Segurança Pública",
    "SETUR": "Secretaria Municipal de Turismo",
    "SETEM": "Secretaria Municipal do Trabalho"
}

secretarias_formatadas = [f"{sigla} - {nome}" for sigla, nome in secretarias_dict.items()]

modalidades_licitacao = [
    "Dispensa", "Convite", "Tomada de Preços", "Concorrência Pública",
    "Concurso", "Leilão", "Pregão", "Inexigibilidade",
    "Ata de Registro de Preço", "Comparação de Preços",
    "Regras Próprias de Organismos Internacionais",
    "Regime Diferenciado de Contratações (RDC)",
    "Chamada Pública", "Não se Aplica", "Diálogo Competitivo"
]

# Dicionário para busca flexível de secretarias

secretarias_busca = {}
for sigla, nome in secretarias_dict.items():
    secretarias_busca[sigla.lower()] = sigla
    secretarias_busca[nome.lower()] = sigla
    for parte in nome.lower().split():
        secretarias_busca[parte] = sigla

# --- BANCO DE DADOS ---

# Caminho absoluto para o arquivo do banco de dados
caminho_banco = os.path.abspath(os.path.join(os.path.dirname(__file__), 'meus_trabalhos.db'))

# Conexão com o banco de dados SQLite
conn = sqlite3.connect(
    caminho_banco,
    detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES
)

cursor = conn.cursor()

# Otimizações e integridade do SQLite
try:
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA busy_timeout=5000")
    conn.execute("PRAGMA temp_store=MEMORY")
    conn.execute("PRAGMA cache_size=-32000")
except Exception:
    pass

# Toast informativo (popup leve)
toast_popup = None


def mostrar_toast(mensagem: str, duracao_ms: int = 2500):
    try:
        global toast_popup
        # Fecha toast anterior, se existir
        if toast_popup and toast_popup.winfo_exists():
            toast_popup.destroy()

        # Posição relativa à janela principal
        janela.update_idletasks()
        root_x = janela.winfo_rootx()
        root_y = janela.winfo_rooty()
        root_w = janela.winfo_width()
        root_h = janela.winfo_height()

        largura = 360
        altura = 36
        x = root_x + (root_w // 2) - (largura // 2)
        y = root_y + root_h - altura - 40

        toast_popup = tk.Toplevel(janela)
        toast_popup.wm_overrideredirect(True)
        toast_popup.configure(bg="#eaf7ea", padx=10, pady=6)
        toast_popup.geometry(f"{largura}x{altura}+{x}+{y}")

        lbl = tk.Label(
            toast_popup,
            text=f"✅ {mensagem}",
            bg="#eaf7ea",
            fg="#1b5e20",
            font=("Segoe UI", 10)
        )
        lbl.pack(expand=True, fill="both")

        # Fecha automaticamente
        toast_popup.after(duracao_ms, toast_popup.destroy)
    except Exception:
        # Evita quebrar fluxo
        pass


# Backup automático do banco (mantém apenas os 10 mais recentes)
def backup_automatico(process_numbers=None):
    try:
        destino = r'C:\\Users\\User\\OneDrive\\MeuGestor\\MiniBanco'
        os.makedirs(destino, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        sufixo = ""
        try:
            if process_numbers:
                if isinstance(process_numbers, (list, tuple)):
                    if len(process_numbers) == 1:
                        import re as _re
                        proc = str(process_numbers[0])
                        proc_sanit = _re.sub(r'[^0-9A-Za-z_-]+', '-', proc)
                        sufixo = f"_proc-{proc_sanit}"
                    else:
                        sufixo = f"_multi-{len(process_numbers)}"
                else:
                    import re as _re
                    proc = str(process_numbers)
                    proc_sanit = _re.sub(r'[^0-9A-Za-z_-]+', '-', proc)
                    sufixo = f"_proc-{proc_sanit}"
        except Exception:
            sufixo = ""

        nome_arquivo = f"meus_trabalhos_{timestamp}{sufixo}.db"
        caminho_backup = os.path.join(destino, nome_arquivo)

        # Usa a API de backup do sqlite para consistência com WAL
        with sqlite3.connect(caminho_backup) as bk_conn:
            conn.backup(bk_conn)

        # Rotação: mantém apenas os 10 backups mais recentes
        arquivos = [
            os.path.join(destino, f) for f in os.listdir(destino)
            if f.startswith('meus_trabalhos_') and f.endswith('.db')
        ]
        arquivos.sort(key=lambda p: os.path.getmtime(p))
        while len(arquivos) > 10:
            mais_antigo = arquivos.pop(0)
            try:
                os.remove(mais_antigo)
            except Exception:
                break

        # Toast de confirmação
        try:
            mostrar_toast(f"Backup criado: {nome_arquivo}")
        except Exception:
            pass
    except Exception as e:
        # Evita quebrar fluxo principal por falha de backup
        try:
            print(f"Aviso: falha no backup automático: {e}")
        except Exception:
            pass


# Garante que a tabela 'promessas' existe
colunas_db = []  # ou outro valor apropriado

# Se a tabela ainda não existe, cria com as colunas necessárias
if not colunas_db:
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS trabalhos_realizados (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data_registro TEXT,
            numero_processo TEXT UNIQUE,
            secretaria TEXT,
            numero_licitacao TEXT,
            situacao TEXT,
            modalidade TEXT,
            data_inicio TEXT,
            data_entrega TEXT,
            entregue_por TEXT,
            devolvido_a TEXT,
            contratado TEXT,
            descricao TEXT
        )
    ''')
    conn.commit()
    print("✅ Tabela 'trabalhos_realizados' criada com sucesso.")
else:
    # Lista de colunas que podem precisar ser adicionadas após uma atualização no sistema
    colunas_necessarias = ['entregue_por', 'devolvido_a', 'data_registro', 'modalidade', 'descricao', 'contratado']

    for coluna in colunas_necessarias:
        if coluna not in colunas_db:
            try:
                cursor.execute(f'ALTER TABLE trabalhos_realizados ADD COLUMN {coluna} TEXT')
                print(f"➕ Coluna '{coluna}' adicionada.")
                # Ajusta valores padrão apenas para data_registro (se necessário)
                if coluna == 'data_registro':
                    cursor.execute('UPDATE trabalhos_realizados SET data_registro = datetime("now", "localtime")')
                conn.commit()
            except sqlite3.OperationalError as e:
                print(f'⚠️ Erro ao adicionar a coluna "{coluna}": {e}')

# Carrega os nomes para autocompletar
nomes_autocomplete = carregar_nomes_autocomplete()
# Lista específica para Contratado
nomes_contratado = carregar_nomes_contratado()

# Após criar trabalhos_realizados
cursor.execute('''
    CREATE TABLE IF NOT EXISTS trabalhos_excluidos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        data_exclusao TEXT,
        data_registro TEXT,
        numero_processo TEXT,
        secretaria TEXT,
        numero_licitacao TEXT,
        situacao TEXT,
        modalidade TEXT,
        data_inicio TEXT,
        data_entrega TEXT,
        entregue_por TEXT,
        devolvido_a TEXT,
        contratado TEXT,
        descricao TEXT
    )
''')
conn.commit()

# Criar tabela promessas para lembretes
cursor.execute('''
    CREATE TABLE IF NOT EXISTS promessas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        data_prometida TEXT NOT NULL,
        descricao TEXT NOT NULL,
        pessoa TEXT,
        notificado INTEGER DEFAULT 0,
        data_criacao TEXT DEFAULT (datetime('now', 'localtime'))
    )
''')
conn.commit()

try:
    cursor.execute("PRAGMA table_info(promessas)")
    colunas_existentes = [coluna[1] for coluna in cursor.fetchall()]

    if 'data_criacao' not in colunas_existentes:
        try:
            cursor.execute('ALTER TABLE promessas ADD COLUMN data_criacao TEXT')
            conn.commit()
            cursor.execute(
                "UPDATE promessas SET data_criacao = datetime('now', 'localtime') WHERE data_criacao IS NULL")
            conn.commit()
        except Exception as e:
            print(f"⚠️ Erro ao adicionar coluna data_criacao: {e}")

except Exception as e:
    print(f"⚠️ Erro ao verificar/atualizar tabela promessas: {e}")


def toggle_lembrete():
    """Ativa/desativa o modo lembrete com feedback visual melhorado"""
    if lembrete_var.get():
        # Modo lembrete ativado
        entrada_descricao.config(
            bg="#FFF9C4",  # Amarelo claro
            fg="#2C3E50",  # Texto escuro para melhor contraste
            relief="solid",
            bd=1
        )
        # Removido: tooltip de modo lembrete ativado
    else:
        # Modo lembrete desativado
        entrada_descricao.config(
            bg="white",  # Fundo branco padrão
            fg="black",  # Texto preto padrão
            relief="flat",
            bd=1
        )

    # Atualiza a cor do botão baseado no estado do checkbox
    atualizar_cor_botao_lembrete_checkbox()

    # Foca no campo de descrição se o lembrete foi ativado
    if lembrete_var.get():
        entrada_descricao.focus_set()


def carregar_lembretes_existente():
    """Atualiza os lembretes na janela já aberta"""
    global lembrete_ids, lembrete_ids_recente, lembrete_ids_antigo
    if janela_lembretes_aberta and janela_lembretes_aberta.winfo_exists():
        # Encontra a lista de lembretes na janela existente
        for widget in janela_lembretes_aberta.winfo_children():
            if isinstance(widget, Frame):
                for subwidget in widget.winfo_children():
                    if isinstance(subwidget, Listbox):
                        # Atualiza a lista
                        subwidget.delete(0, tk.END)
                        cursor.execute("""
                            SELECT 
                                rowid,
                                COALESCE(strftime('%d/%m/%Y', data_criacao), data_prometida) AS data_exibir,
                                descricao 
                            FROM promessas 
                            WHERE pessoa = 'Lembrete'
                            ORDER BY datetime(data_criacao) DESC
                        """)
                        lembretes = cursor.fetchall()
                        # Mapeamentos para abas
                        lembrete_ids_recente = {idx: row[0] for idx, row in enumerate(lembretes)}
                        lembrete_ids_antigo = {idx: row[0] for idx, row in enumerate(reversed(lembretes))}
                        lembrete_ids = lembrete_ids_recente

                        # Preenche conforme a ordem da aba: recente (DESC) ou antigo (ASC)
                        if hasattr(subwidget, 'lembretes_ordem') and subwidget.lembretes_ordem == 'asc':
                            for _, data_exibir, descricao in reversed(lembretes):
                                subwidget.insert(tk.END, f"{data_exibir} - {descricao}")
                        else:
                            for _, data_exibir, descricao in lembretes:
                                subwidget.insert(tk.END, f"{data_exibir} - {descricao}")
                        return


def mostrar_lembretes_iniciais():
    """Exibe em sequência os lembretes cuja descrição começa por 'Lembrar'.

    Ao fechar um lembrete, o próximo é mostrado até acabar.
    """
    try:
        cursor.execute(
            """
            SELECT rowid, strftime('%d/%m/%Y', data_criacao) AS data_exibir, descricao
            FROM promessas
            WHERE pessoa = 'Lembrete'
              AND (
                    descricao LIKE 'Lembrar%'
                 OR descricao LIKE 'LEMBRAR%'
              )
            ORDER BY datetime(data_criacao) ASC, rowid ASC
            """
        )
        lembretes = cursor.fetchall()

        for _, data_exibir, descricao in lembretes:
            try:
                messagebox.showinfo("Lembrete", f"{data_exibir}: {descricao}")
            except Exception:
                # Se messagebox falhar, continuar com os próximos
                continue
    except Exception as e:
        print(f"[ERRO] Ao mostrar lembretes iniciais: {e}")


def abrir_lembretes():
    """Abre janela com os lembretes cadastrados ou atualiza se já estiver aberta"""
    global lembrete_ids, janela_lembretes_aberta

    # Verifica se existem lembretes cadastrados antes de abrir a janela
    cursor.execute("SELECT COUNT(*) FROM promessas WHERE pessoa = 'Lembrete'")
    count_lembretes = cursor.fetchone()[0]

    if count_lembretes == 0:
        messagebox.showinfo("Informação", "Nenhum lembrete cadastrado")
        return

    janela_lembretes = Toplevel(janela)
    janela_lembretes_aberta = janela_lembretes
    janela_lembretes.title("Lembretes")
    # Geometria inicial com persistência: usa último tamanho salvo
    screen_w = janela_lembretes.winfo_screenwidth()
    w, h, xi, yi = carregar_tamanho_janela('lembretes', 500, 450)
    if xi is None or yi is None:
        x = (screen_w - w) // 2
        y = 60
        janela_lembretes.geometry(f"{w}x{h}+{x}+{y}")
    else:
        janela_lembretes.geometry(f"{w}x{h}+{xi}+{yi}")
    janela_lembretes.resizable(True, True)
    janela_lembretes.configure(bg="#F5F7FA")

    def ao_fechar_lembretes():
        global janela_lembretes_aberta
        try:
            janela_lembretes.update_idletasks()
            salvar_tamanho_janela('lembretes', janela_lembretes.winfo_width(), janela_lembretes.winfo_height(),
                                  janela_lembretes.winfo_x(), janela_lembretes.winfo_y())
        except Exception:
            pass
        janela_lembretes_aberta = None
        janela_lembretes.destroy()

    janela_lembretes.protocol("WM_DELETE_WINDOW", ao_fechar_lembretes)

    # Notebook com abas para organização
    notebook = ttk.Notebook(janela_lembretes)
    notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=6)

    # Aba: Mais recentes
    tab_recente = tk.Frame(notebook, bg="#F5F7FA")
    notebook.add(tab_recente, text="Mais recentes")

    scroll_recente = tk.Scrollbar(tab_recente)
    scroll_recente.pack(side=tk.RIGHT, fill=tk.Y)

    lista_recente = tk.Listbox(
        tab_recente,
        yscrollcommand=scroll_recente.set,
        font=("Segoe UI", 10),
        bg="#FFFFFF",
        fg="#000000",
        selectbackground="#0078d4",
        selectforeground="white",
        width=80,
        height=15,
        selectmode=tk.EXTENDED,
        exportselection=False
    )
    lista_recente.pack(fill=tk.BOTH, expand=True)
    scroll_recente.config(command=lista_recente.yview)
    # Flag para ordem
    lista_recente.lembretes_ordem = 'desc'

    # Aba: Mais antigos
    tab_antigo = tk.Frame(notebook, bg="#F5F7FA")
    notebook.add(tab_antigo, text="Mais antigos")

    scroll_antigo = tk.Scrollbar(tab_antigo)
    scroll_antigo.pack(side=tk.RIGHT, fill=tk.Y)

    lista_antigo = tk.Listbox(
        tab_antigo,
        yscrollcommand=scroll_antigo.set,
        font=("Segoe UI", 10),
        bg="#FFFFFF",
        fg="#000000",
        selectbackground="#0078d4",
        selectforeground="white",
        width=80,
        height=15,
        selectmode=tk.EXTENDED,
        exportselection=False
    )
    lista_antigo.pack(fill=tk.BOTH, expand=True)
    scroll_antigo.config(command=lista_antigo.yview)
    # Flag para ordem
    lista_antigo.lembretes_ordem = 'asc'

    tooltip_win = None
    tooltip_index = None
    hover_item_index = None
    last_list = None

    def hide_tooltip():
        nonlocal tooltip_win, tooltip_index, hover_item_index, last_list
        if tooltip_win and tooltip_win.winfo_exists():
            tooltip_win.destroy()
        tooltip_win = None
        tooltip_index = None
        try:
            if last_list is not None and hover_item_index is not None:
                # Restaura cor padrão do item
                default_bg = last_list['bg']
                default_fg = last_list['fg'] if 'fg' in last_list.keys() else None
                try:
                    if default_fg:
                        last_list.itemconfig(hover_item_index, bg=default_bg, fg=default_fg)
                    else:
                        last_list.itemconfig(hover_item_index, bg=default_bg)
                except Exception:
                    pass
        except Exception:
            pass
        hover_item_index = None

    def show_tooltip(text, x, y, selected=False):
        nonlocal tooltip_win
        if not text:
            return
        if tooltip_win and tooltip_win.winfo_exists():
            tooltip_win.geometry(f"+{x}+{y}")
            label = tooltip_win.children.get('label')
            if label:
                label.config(text=text)
                try:
                    bg = TOOLTIP_SELECTED_BG if selected else TOOLTIP_BG
                    fg = "white"
                    label.config(bg=bg, fg=fg)
                    default_bg = bg
                    default_fg = fg
                    hover_bg = TOOLTIP_HOVER_BG
                    hover_fg = TOOLTIP_HOVER_FG
                    try:
                        label.bind("<Enter>", lambda e, _lbl=label: _lbl.config(bg=hover_bg, fg=hover_fg))
                        label.bind("<Leave>",
                                   lambda e, _lbl=label, _bg=default_bg, _fg=default_fg: _lbl.config(bg=_bg, fg=_fg))
                    except Exception:
                        pass
                except Exception:
                    pass
            return
        tooltip_win = tk.Toplevel(janela_lembretes)
        tooltip_win.wm_overrideredirect(True)
        tooltip_win.geometry(f"+{x}+{y}")

        bg = TOOLTIP_SELECTED_BG if selected else TOOLTIP_BG
        fg = "white"
        lbl = tk.Label(tooltip_win, name='label', text=text, bg=bg, fg=fg, padx=8, pady=6,
                       font=("Segoe UI", 9), justify=tk.LEFT, wraplength=380)
        lbl.pack()
        try:
            default_bg = bg
            default_fg = fg
            hover_bg = TOOLTIP_HOVER_BG
            hover_fg = TOOLTIP_HOVER_FG
            lbl.bind("<Enter>", lambda e, _lbl=lbl: _lbl.config(bg=hover_bg, fg=hover_fg))
            lbl.bind("<Leave>", lambda e, _lbl=lbl, _bg=default_bg, _fg=default_fg: _lbl.config(bg=_bg, fg=_fg))
        except Exception:
            pass

    def on_list_motion(event, lst):
        nonlocal tooltip_index
        try:
            idx = lst.nearest(event.y)
            box = lst.bbox(idx)
            if not box:
                hide_tooltip()
                return
            y0 = box[1]
            h = box[3]
            if not (y0 <= event.y <= y0 + h):
                hide_tooltip()
                return
            item = lst.get(idx)
            parts = item.split(' - ', 1)
            text = parts[1] if len(parts) > 1 else item
            if tooltip_index == idx and tooltip_win:
                return
            tooltip_index = idx
            xr = lst.winfo_rootx() + event.x + 16
            yr = lst.winfo_rooty() + event.y + 16
            try:
                sel = lst.curselection() or ()
                # Tkinter retorna índices como strings; converte para int
                selected_indices = set()
                for i in sel:
                    try:
                        selected_indices.add(int(i))
                    except Exception:
                        selected_indices.add(i)
                selected = idx in selected_indices
            except Exception:
                selected = False
            show_tooltip(text, xr, yr, selected)
            # Destaca item sob o cursor com azul claro quando não está selecionado
            try:
                nonlocal hover_item_index, last_list
                last_list = lst
                if not selected:
                    if hover_item_index is not None and hover_item_index != idx:
                        # Restaura cor do item anterior
                        default_bg = lst['bg']
                        try:
                            lst.itemconfig(hover_item_index, bg=default_bg)
                        except Exception:
                            pass
                    # Aplica cor de hover
                    lst.itemconfig(idx, bg=TOOLTIP_HOVER_BG)
                    hover_item_index = idx
            except Exception:
                pass
        except Exception:
            hide_tooltip()

    lista_recente.bind('<Motion>', lambda e: on_list_motion(e, lista_recente))
    lista_recente.bind('<Leave>', lambda e: hide_tooltip())
    lista_antigo.bind('<Motion>', lambda e: on_list_motion(e, lista_antigo))
    lista_antigo.bind('<Leave>', lambda e: hide_tooltip())

    def carregar_lembretes():
        """Carrega os lembretes do banco de dados e preenche as abas"""
        global lembrete_ids, lembrete_ids_recente, lembrete_ids_antigo
        cursor.execute("""
            SELECT 
                rowid,
                COALESCE(strftime('%d/%m/%Y', data_criacao), data_prometida) AS data_exibir,
                descricao 
            FROM promessas 
            WHERE pessoa = 'Lembrete'
            ORDER BY datetime(data_criacao) DESC
        """)
        lembretes = cursor.fetchall()
        # Mapeamentos por aba
        lembrete_ids_recente = {idx: row[0] for idx, row in enumerate(lembretes)}
        lembrete_ids_antigo = {idx: row[0] for idx, row in enumerate(reversed(lembretes))}
        # Por compatibilidade legado
        lembrete_ids = lembrete_ids_recente

        # Aba: Mais recentes
        lista_recente.delete(0, tk.END)
        for _, data_exibir, descricao in lembretes:
            lista_recente.insert(tk.END, f"{data_exibir} - {descricao}")

        # Aba: Mais antigos
        lista_antigo.delete(0, tk.END)
        for _, data_exibir, descricao in reversed(lembretes):
            lista_antigo.insert(tk.END, f"{data_exibir} - {descricao}")

    # --- Habilitar seleção estendida e atalhos de teclado nas listas ---
    def habilitar_selecao_estendida(lst: tk.Listbox):
        try:
            lst.configure(selectmode=tk.EXTENDED, exportselection=False)

            def _set_anchor_on_click(event):
                try:
                    idx = lst.nearest(event.y)
                    lst.selection_anchor(idx)
                except Exception:
                    pass

            def _shift_select(direction):
                try:
                    current = lst.index(tk.ACTIVE)
                    size = lst.size()
                    if size <= 0:
                        return None
                    new_index = max(0, current - 1) if direction == -1 else min(size - 1, current + 1)
                    anchor = lst.index(tk.ANCHOR)
                    start = min(anchor, new_index)
                    end = max(anchor, new_index)
                    lst.selection_clear(0, tk.END)
                    lst.selection_set(start, end)
                    lst.activate(new_index)
                    lst.see(new_index)
                    return 'break'
                except Exception:
                    return None

            def _on_shift_up(event):
                return _shift_select(-1)

            def _on_shift_down(event):
                return _shift_select(1)

            def _ctrl_toggle_active(event):
                try:
                    idx = lst.index(tk.ACTIVE)
                    if idx in lst.curselection():
                        lst.selection_clear(idx)
                    else:
                        lst.selection_set(idx)
                    return 'break'
                except Exception:
                    return None

            lst.bind("<Button-1>", _set_anchor_on_click)
            lst.bind("<Shift-Up>", _on_shift_up)
            lst.bind("<Shift-Down>", _on_shift_down)
            lst.bind("<Control-space>", _ctrl_toggle_active)
        except Exception:
            pass

    habilitar_selecao_estendida(lista_recente)
    habilitar_selecao_estendida(lista_antigo)

    def listbox_ativa():
        """Retorna a Listbox da aba ativa"""
        tab_id = notebook.select()
        tab_widget = notebook.nametowidget(tab_id)
        for child in tab_widget.winfo_children():
            if isinstance(child, tk.Listbox):
                return child
        return None

    def selecionar_todos():
        """Seleciona todos os lembretes da listbox ativa"""
        lb = listbox_ativa()
        if lb:
            lb.select_set(0, tk.END)
            print(f"[DEBUG] Ctrl+A pressionado - {lb.size()} lembretes selecionados")

    def excluir_selecionados():
        """Exclui todos os lembretes selecionados na listbox ativa"""
        lb = listbox_ativa()
        if not lb:
            return
        selecoes = lb.curselection()
        if not selecoes:
            messagebox.showwarning("Aviso", "Selecione pelo menos um lembrete para excluir!", parent=janela_lembretes)
            return

        if not messagebox.askyesno(
                "Confirmar Exclusão",
                f"Deseja excluir {len(selecoes)} lembrete(s) selecionado(s)?",
                parent=janela_lembretes
        ):
            return

        try:
            # Escolhe mapeamento correto por aba
            mapping = lembrete_ids_recente if lb is lista_recente else lembrete_ids_antigo
            for idx in sorted(selecoes, reverse=True):
                id_lembrete = mapping.get(idx)
                if id_lembrete is None:
                    continue
                cursor.execute("DELETE FROM promessas WHERE rowid = ?", (id_lembrete,))
                lb.delete(idx)

            conn.commit()
            cache.invalidate('count_promessas')
            atualizar_cor_botao_lembrete()
            carregar_lembretes()

            messagebox.showinfo("Sucesso", f"{len(selecoes)} lembrete(s) excluído(s) com sucesso!",
                                parent=janela_lembretes)

            if lb.size() == 0:
                ao_fechar_lembretes()

        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao excluir lembretes:\n{str(e)}", parent=janela_lembretes)
            conn.rollback()

    # Função intermediária para evitar múltiplas chamadas rápidas
    def visualizar_com_debounce():
        """Evita múltiplas chamadas rápidas de visualização"""
        janela_lembretes.after(10, visualizar_lembrete_selecionado)

    def visualizar_lembrete_selecionado():
        """Visualiza o lembrete selecionado da listbox ativa em uma janela modal,
        com navegação entre lembretes (Anterior/Próximo)."""
        global janela_visualizar_ativa, ultimo_clique_visualizar

        # Verifica se já foi clicado recentemente (dentro de 500ms)
        tempo_atual = time.time() * 1000  # Tempo em milissegundos
        if tempo_atual - ultimo_clique_visualizar < 500:  # 500ms de intervalo
            return  # Ignora o clique duplicado
        ultimo_clique_visualizar = tempo_atual

        # Se já existir uma janela de visualização aberta, apenas foca nela e sai
        if janela_visualizar_ativa and janela_visualizar_ativa.winfo_exists():
            janela_visualizar_ativa.focus_set()
            return

        lb = listbox_ativa()
        if not lb:
            return
        selecoes = lb.curselection()
        # Sem seleção: se houver itens na lista, abre no primeiro
        if not selecoes:
            if lb.size() > 0:
                selecoes = (0,)
            else:
                messagebox.showwarning("Aviso", "Nenhum lembrete disponível para visualizar.", parent=janela_lembretes)
                return

        if len(selecoes) > 1:
            messagebox.showwarning("Aviso", "Selecione apenas um lembrete para visualizar!", parent=janela_lembretes)
            return

        # Seleciona o mapeamento correto pela aba ativa
        mapping = lembrete_ids_recente if lb.lembretes_ordem == 'desc' else lembrete_ids_antigo
        total_lembretes = len(mapping)
        indice_atual = tk.IntVar(value=selecoes[0])
        current_rowid = None

        win = LembreteDetailsWindow(janela_lembretes, mapping, cursor, conn, carregar_lembretes,
                                    start_index=indice_atual.get())
        globals()['janela_visualizar_ativa'] = getattr(win, 'win', None)
        try:
            janela_visualizar_ativa.focus_set()
        except Exception:
            pass

    def duplo_clique_visualizar(event):
        """Visualiza lembrete com duplo clique usando debounce"""
        visualizar_com_debounce()

    # Menu de contexto
    menu_contexto = tk.Menu(janela_lembretes, tearoff=0)
    menu_contexto.add_command(label="Visualizar", command=visualizar_com_debounce)
    menu_contexto.add_command(label="Selecionar Todos", command=lambda: selecionar_ou_desmarcar())
    menu_contexto.add_separator()
    menu_contexto.add_command(label="Excluir", command=excluir_selecionados)

    def mostrar_menu_contexto(event):
        """Mostra o menu de contexto na listbox alvo"""
        lb = event.widget
        try:
            index = lb.nearest(event.y)
            if index >= 0 and index < lb.size():
                lb.selection_clear(0, tk.END)
                lb.selection_set(index)
                lb.activate(index)
            menu_contexto.tk_popup(event.x_root, event.y_root)
        finally:
            menu_contexto.grab_release()

    # Eventos nas duas listas
    for lb in (lista_recente, lista_antigo):
        lb.bind("<Button-3>", mostrar_menu_contexto)
        lb.bind("<Double-Button-1>", duplo_clique_visualizar)
        lb.bind("<<ListboxSelect>>", lambda e: update_select_button_label())

    # Bind global para Ctrl-A selecionar todos
    janela_lembretes.bind_all("<Control-a>", lambda e: selecionar_todos())

    # Atualiza rótulo quando trocar de aba
    notebook.bind("<<NotebookTabChanged>>", lambda e: update_select_button_label())

    # Variável para controlar rótulo do botão Selecionar
    btn_selecionar = None

    # Funções do botão Selecionar Todos / Desmarcar
    def todos_selecionados(lb):
        try:
            size = lb.size()
            return size > 0 and len(lb.curselection()) == size
        except Exception:
            return False

    def update_select_button_label():
        lb = listbox_ativa()
        if not lb:
            return
        try:
            if todos_selecionados(lb):
                btn_selecionar.config(text="Desmarcar")
            else:
                btn_selecionar.config(text="Selecionar Todos")
        except Exception:
            # Caso o botão ainda não tenha sido criado no ciclo de inicialização
            pass

    def selecionar_ou_desmarcar():
        lb = listbox_ativa()
        if not lb:
            return
        if todos_selecionados(lb):
            lb.selection_clear(0, tk.END)
        else:
            lb.select_set(0, tk.END)
        update_select_button_label()

    # Alterna rótulo/estilo do botão Editar -> Atualizar
    def alternar_botao_editar(editando=False):
        try:
            if editando:
                btn_editar.config(text="Atualizar", bg=BUTTON_PRIMARY_BG, fg=BUTTON_PRIMARY_FG,
                                  activebackground=BUTTON_PRIMARY_ACTIVE_BG, activeforeground=BUTTON_PRIMARY_ACTIVE_FG,
                                  highlightbackground=BUTTON_PRIMARY_HIGHLIGHT)
            else:
                btn_editar.config(text="Editar", bg=BUTTON_PRIMARY_BG, fg=BUTTON_PRIMARY_FG,
                                  activebackground=BUTTON_PRIMARY_ACTIVE_BG, activeforeground=BUTTON_PRIMARY_ACTIVE_FG,
                                  highlightbackground=BUTTON_PRIMARY_HIGHLIGHT)
        except Exception:
            pass

    def editar_lembrete(rowid: int):
        """Abre janela para editar o lembrete e salva alterações."""
        # Buscar dados atuais
        dados = get_promessa_by_id(cursor, rowid)
        if not dados:
            messagebox.showerror("Erro", "Lembrete não encontrado.", parent=janela_lembretes)
            return

        desc_atual, _, data_criacao = dados
        data_atual = data_criacao if data_criacao else ""

        # Janela de edição
        win = tk.Toplevel(janela_lembretes)
        win.title("Editar Lembrete")
        try:
            w, h, xi, yi = carregar_tamanho_janela('editar_lembrete', 420, 240)
            if xi is not None and yi is not None:
                win.geometry(f"{w}x{h}+{xi}+{yi}")
            else:
                center_window(win, w, h)
        except Exception:
            center_window(win, 420, 240)
        win.transient(janela_lembretes)
        win.grab_set()
        try:
            bind_persist_geometry('editar_lembrete', win)
        except Exception:
            pass

        tk.Label(win, text="Data prometida (dd/mm/aaaa)").pack(anchor=tk.W, padx=10, pady=(10, 0))
        entry_data = tk.Entry(win)
        entry_data.pack(fill=tk.X, padx=10)
        entry_data.insert(0, data_atual or "")

        tk.Label(win, text="Descrição").pack(anchor=tk.W, padx=10, pady=(10, 0))
        text_desc = tk.Text(win, height=5)
        try:
            enhance_text_controls(text_desc)
        except Exception:
            pass
        text_desc.pack(fill=tk.BOTH, expand=True, padx=10)
        text_desc.insert("1.0", desc_atual or "")

        def salvar():
            nova_data = entry_data.get().strip()
            nova_desc = text_desc.get("1.0", "end-1c").strip()

            if not nova_data or not nova_desc:
                messagebox.showwarning("Aviso", "Preencha data e descrição.", parent=win)
                return

            # Validação simples de data dd/mm/aaaa
            try:
                datetime.strptime(nova_data, "%d/%m/%Y")
            except Exception:
                messagebox.showwarning("Aviso", "Data deve estar no formato dd/mm/aaaa.", parent=win)
                return

            try:
                cursor.execute(
                    "UPDATE promessas SET data_prometida = ?, descricao = ? WHERE rowid = ?",
                    (nova_data, nova_desc, rowid)
                )
                conn.commit()
                cache.invalidate('count_promessas')
                atualizar_cor_botao_lembrete()
                carregar_lembretes()
                messagebox.showinfo("Sucesso", "Lembrete atualizado com sucesso!", parent=janela_lembretes)
                win.destroy()
                alternar_botao_editar(False)
            except Exception as e:
                conn.rollback()
                messagebox.showerror("Erro", f"Falha ao atualizar lembrete:\n{str(e)}", parent=win)

        def cancelar():
            try:
                win.update_idletasks()
                salvar_tamanho_janela('editar_lembrete', win.winfo_width(), win.winfo_height(), win.winfo_x(),
                                      win.winfo_y())
            except Exception:
                pass
            win.destroy()
            alternar_botao_editar(False)

        botoes = tk.Frame(win)
        botoes.pack(pady=10)
        btn_salvar = tk.Button(botoes, text="Salvar", command=salvar)
        btn_cancelar = tk.Button(botoes, text="Cancelar", command=cancelar)
        btn_salvar.pack(side=tk.LEFT, padx=5)
        btn_cancelar.pack(side=tk.LEFT, padx=5)

    def editar_selecionado():
        """Edita o lembrete selecionado na aba ativa"""
        lb = listbox_ativa()
        if not lb:
            return
        selecoes = lb.curselection()
        if not selecoes:
            messagebox.showwarning("Aviso", "Selecione um lembrete para editar!", parent=janela_lembretes)
            return
        if len(selecoes) > 1:
            messagebox.showwarning("Aviso", "Selecione apenas um lembrete para editar!", parent=janela_lembretes)
            return

        # Mapeamento conforme aba
        mapping = lembrete_ids_recente if lb is lista_recente else lembrete_ids_antigo
        idx = selecoes[0]
        rowid = mapping.get(idx)
        if rowid is None:
            messagebox.showwarning("Aviso", "Não foi possível identificar o lembrete.", parent=janela_lembretes)
            return

        alternar_botao_editar(True)
        editar_lembrete(rowid)

    # Frame para botões inferiores
    frame_botoes = tk.Frame(janela_lembretes, bg="#F5F7FA")
    frame_botoes.pack(pady=10)

    estilo_botao_padrao = {
        "bg": BUTTON_PRIMARY_BG,
        "fg": BUTTON_PRIMARY_FG,
        "activebackground": BUTTON_PRIMARY_ACTIVE_BG,
        "activeforeground": BUTTON_PRIMARY_ACTIVE_FG,
        "font": ("Segoe UI", 10, "bold"),
        "relief": tk.FLAT,
        "bd": 0,
        "highlightthickness": 1,
        "highlightbackground": BUTTON_PRIMARY_HIGHLIGHT,
    }
    btn_visualizar = tk.Button(
        frame_botoes,
        text="Visualizar",
        command=visualizar_com_debounce,
        width=10,
        **estilo_botao_padrao
    )
    btn_visualizar.pack(side=tk.LEFT, padx=5)
    janela_lembretes.bind("<F10>", lambda e: (piscar_botao(btn_visualizar), visualizar_com_debounce()))
    btn_selecionar = tk.Button(
        frame_botoes,
        text="Selecionar Todos",
        command=selecionar_ou_desmarcar,
        width=16,
        **estilo_botao_padrao
    )
    btn_selecionar.pack(side=tk.LEFT, padx=5)

    btn_editar = tk.Button(
        frame_botoes,
        text="Editar",
        command=editar_selecionado,
        **estilo_botao_padrao
    )
    btn_editar.pack(side=tk.LEFT, padx=5)

    estilo_excluir = dict(estilo_botao_padrao)
    estilo_excluir.update({
        "bg": BUTTON_DANGER_BG,
        "fg": BUTTON_DANGER_FG,
        "activebackground": BUTTON_DANGER_ACTIVE_BG,
        "activeforeground": BUTTON_DANGER_ACTIVE_FG,
        "highlightbackground": BUTTON_DANGER_HIGHLIGHT,
        "highlightthickness": 2
    })
    btn_excluir = tk.Button(
        frame_botoes,
        text="Excluir",
        command=excluir_selecionados,
        width=10,
        **estilo_excluir
    )
    btn_excluir.pack(side=tk.LEFT, padx=5)

    # Inicializa rótulo do botão Selecionar/Desmarcar conforme estado atual
    update_select_button_label()

    # Carrega os lembretes
    carregar_lembretes()

    # Foca na janela
    janela_lembretes.focus_set()
    janela_lembretes.bind("<Escape>", lambda e: ao_fechar_lembretes())


def deletar_lembrete(lista, janela_lembretes):
    """Deleta o lembrete selecionado e fecha a janela"""
    global lembrete_ids

    selecionado = lista.curselection()
    if not selecionado:
        messagebox.showwarning("Aviso", "Selecione um lembrete para deletar!")
        return

    try:
        # Obtém o ID do lembrete selecionado
        id_lembrete = lembrete_ids[selecionado[0]]

        # Executa a exclusão no banco de dados
        try:
            delete_promessa(cursor, id_lembrete)
        except Exception:
            cursor.execute("DELETE FROM promessas WHERE rowid = ?", (id_lembrete,))
        conn.commit()

        # Remove da lista visual
        lista.delete(selecionado[0])

        # Atualiza o dicionário de IDs
        lembrete_ids = {i: id for i, (idx, id) in enumerate(lembrete_ids.items()) if idx != selecionado[0]}

        # Invalida cache de promessas
        cache.invalidate('count_promessas')

        # Atualiza a cor do botão lembrete
        atualizar_cor_botao_lembrete()

        messagebox.showinfo("Sucesso", "Lembrete excluído com sucesso!")

        # Se a lista ficou vazia, fecha a janela
        if lista.size() == 0:
            janela_lembretes.destroy()

    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao excluir lembrete:\n{str(e)}")
        conn.rollback()

    def ao_fechar_lembretes():
        global janela_lembretes_aberta
        janela_lembretes_aberta = None
        janela_lembretes.destroy()

    janela_lembretes.bind("<Escape>", lambda e: ao_fechar_lembretes())
    janela_lembretes.protocol("WM_DELETE_WINDOW", ao_fechar_lembretes)

    # Frame principal
    frame_principal = tk.Frame(janela_lembretes)
    frame_principal.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    # Scrollbar
    scrollbar = tk.Scrollbar(frame_principal)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # Lista de lembretes
    lista_lembretes = tk.Listbox(
        frame_principal,
        yscrollcommand=scrollbar.set,
        font=("Segoe UI", 10),
        bg="#FFFFFF",
        fg="#000000",
        selectbackground="#0078d4",
        selectforeground="white",
        width=80,
        height=15
    )
    lista_lembretes.pack(fill=tk.BOTH, expand=True)
    scrollbar.config(command=lista_lembretes.yview)

    # ESC limpa qualquer seleção na lista de lembretes
    janela_lembretes.bind('<Escape>', lambda e: lista_lembretes.selection_clear(0, tk.END))

    # Frame para botões
    frame_botoes = tk.Frame(janela_lembretes)
    frame_botoes.pack(pady=10)

    def carregar_lembretes():
        """Carrega os lembretes do banco de dados"""
        global lembrete_ids
        lista_lembretes.delete(0, tk.END)
        cursor.execute("""
            SELECT rowid, data_prometida, descricao 
            FROM promessas 
            WHERE pessoa = 'Lembrete'
            ORDER BY date(substr(data_prometida, 7, 4) || '-' || 
                     substr(data_prometida, 4, 2) || '-' || 
                     substr(data_prometida, 1, 2)) DESC
        """)
        lembretes = cursor.fetchall()
        lembrete_ids = {idx: row[0] for idx, row in enumerate(lembretes)}

        for idx, (rowid, data, descricao) in enumerate(lembretes):
            lista_lembretes.insert(tk.END, f"{data} - {descricao}")

    # Função para selecionar todos os lembretes
    def selecionar_todos():
        """Seleciona todos os lembretes da lista"""
        lista_lembretes.select_set(0, tk.END)

    # Função para excluir lembretes selecionados
    def excluir_selecionados():
        """Exclui todos os lembretes selecionados"""
        selecoes = lista_lembretes.curselection()
        if not selecoes:
            messagebox.showwarning("Aviso", "Selecione pelo menos um lembrete para excluir!", parent=janela_lembretes)
            return

        if not messagebox.askyesno(
                "Confirmar Exclusão",
                f"Deseja excluir {len(selecoes)} lembrete(s) selecionado(s)?",
                parent=janela_lembretes
        ):
            return

        try:
            for idx in sorted(selecoes, reverse=True):
                id_lembrete = lembrete_ids[idx]
                cursor.execute("DELETE FROM promessas WHERE rowid = ?", (id_lembrete,))
                lista_lembretes.delete(idx)

            conn.commit()
            cache.invalidate('count_promessas')
            atualizar_cor_botao_lembrete()
            carregar_lembretes()

            messagebox.showinfo("Sucesso", f"{len(selecoes)} lembrete(s) excluído(s) com sucesso!",
                                parent=janela_lembretes)

            if lista_lembretes.size() == 0:
                ao_fechar_lembretes()

        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao excluir lembretes:\n{str(e)}", parent=janela_lembretes)
            conn.rollback()

    # Função para editar lembrete selecionado
    def editar_selecionado():
        """Edita o lembrete selecionado"""
        selecoes = lista_lembretes.curselection()
        if not selecoes:
            messagebox.showwarning("Aviso", "Selecione um lembrete para editar!", parent=janela_lembretes)
            return

        if len(selecoes) > 1:
            messagebox.showwarning("Aviso", "Selecione apenas um lembrete para editar!", parent=janela_lembretes)
            return

        id_lembrete = lembrete_ids[selecoes[0]]
        editar_lembrete(id_lembrete, lista_lembretes, janela_lembretes)
        # Transforma o botão Editar em Atualizar quando estiver editando
        alternar_botao_editar(True)

        # Função para alternar entre Editar e Atualizar
        def alternar_botao_editar(editando=False):
            """Alterna o botão entre Editar e Atualizar"""
            if editando:
                btn_editar.config(text="Atualizar")
            else:
                btn_editar.config(text="Editar")

    def editar_lembrete(rowid: int, lista: tk.Listbox, parent: tk.Toplevel):
        """Abre janela para editar o lembrete e salva alterações."""
        # Buscar dados atuais
        dados = get_promessa_by_id(cursor, rowid)
        if not dados:
            messagebox.showerror("Erro", "Lembrete não encontrado.", parent=parent)
            return

        desc_atual, _, data_criacao = dados
        data_atual = data_criacao if data_criacao else ""

        # Janela de edição
        win = tk.Toplevel(parent)
        win.title("Editar Lembrete")
        try:
            w, h, xi, yi = carregar_tamanho_janela('editar_lembrete', 420, 240)
            if xi is not None and yi is not None:
                win.geometry(f"{w}x{h}+{xi}+{yi}")
            else:
                center_window(win, w, h)
        except Exception:
            center_window(win, 420, 240)
        win.transient(parent)
        win.grab_set()
        try:
            bind_persist_geometry('editar_lembrete', win)
        except Exception:
            pass

        tk.Label(win, text="Data prometida (dd/mm/aaaa)").pack(anchor=tk.W, padx=10, pady=(10, 0))
        entry_data = tk.Entry(win)
        entry_data.pack(fill=tk.X, padx=10)
        entry_data.insert(0, data_atual or "")

        tk.Label(win, text="Descrição").pack(anchor=tk.W, padx=10, pady=(10, 0))
        text_desc = tk.Text(win, height=5)
        try:
            enhance_text_controls(text_desc)
        except Exception:
            pass
        text_desc.pack(fill=tk.BOTH, expand=True, padx=10)
        text_desc.insert("1.0", desc_atual or "")

        def salvar():
            nova_data = entry_data.get().strip()
            nova_desc = text_desc.get("1.0", "end-1c").strip()

            if not nova_data or not nova_desc:
                messagebox.showwarning("Aviso", "Preencha data e descrição.", parent=win)
                return

            # Validação simples de data dd/mm/aaaa
            try:
                datetime.strptime(nova_data, "%d/%m/%Y")
            except Exception:
                messagebox.showwarning("Aviso", "Data deve estar no formato dd/mm/aaaa.", parent=win)
                return

            try:
                cursor.execute(
                    "UPDATE promessas SET data_prometida = ?, descricao = ? WHERE rowid = ?",
                    (nova_data, nova_desc, rowid)
                )
                conn.commit()
                cache.invalidate('count_promessas')
                atualizar_cor_botao_lembrete()
                carregar_lembretes()
                messagebox.showinfo("Sucesso", "Lembrete atualizado com sucesso!", parent=parent)
                win.destroy()
                # Volta o botão para "Editar" após atualizar
                alternar_botao_editar(False)
            except Exception as e:
                conn.rollback()
                messagebox.showerror("Erro", f"Falha ao atualizar lembrete:\n{str(e)}", parent=win)

        def cancelar():
            try:
                win.update_idletasks()
                salvar_tamanho_janela('editar_lembrete', win.winfo_width(), win.winfo_height(), win.winfo_x(),
                                      win.winfo_y())
            except Exception:
                pass
            win.destroy()
            # Volta o botão para "Editar" ao cancelar
            alternar_botao_editar(False)

        botoes = tk.Frame(win)
        botoes.pack(pady=10)
        tk.Button(botoes, text="Salvar", command=salvar).pack(side=tk.LEFT, padx=5)
        tk.Button(botoes, text="Cancelar", command=cancelar).pack(side=tk.LEFT, padx=5)

    # MENU DE CONTEXTO
    menu_contexto = tk.Menu(janela_lembretes, tearoff=0)
    menu_contexto.add_command(label="Visualizar", command=visualizar_lembrete_selecionado)
    menu_contexto.add_command(label="Editar", command=editar_selecionado)
    menu_contexto.add_separator()
    menu_contexto.add_command(label="Selecionar Todos", command=selecionar_todos)
    menu_contexto.add_command(label="Atualizar", command=carregar_lembretes)
    menu_contexto.add_command(label="Excluir", command=excluir_selecionados)

    def mostrar_menu_contexto(event):
        """Mostra o menu de contexto"""
        try:
            # Seleciona o item clicado
            index = lista_lembretes.nearest(event.y)
            if index >= 0 and index < lista_lembretes.size():
                lista_lembretes.selection_clear(0, tk.END)
                lista_lembretes.selection_set(index)
                lista_lembretes.activate(index)

            menu_contexto.tk_popup(event.x_root, event.y_root)
        finally:
            menu_contexto.grab_release()

    def duplo_clique_visualizar(event):
        """Visualiza lembrete com duplo clique"""
        index = lista_lembretes.nearest(event.y)
        if index >= 0 and index < lista_lembretes.size():
            lista_lembretes.selection_clear(0, tk.END)
            lista_lembretes.selection_set(index)
            lista_lembretes.activate(index)
            visualizar_lembrete_selecionado()

    # EVENTOS DE CLIQUE
    def duplo_clique_editar(event):
        """Abre a edição do lembrete com duplo clique no item."""
        index = lista_lembretes.nearest(event.y)
        if index >= 0 and index < lista_lembretes.size():
            lista_lembretes.selection_clear(0, tk.END)
            lista_lembretes.selection_set(index)
            lista_lembretes.activate(index)
            id_lembrete = lembrete_ids.get(index)
            if id_lembrete:
                editar_lembrete(id_lembrete, lista_lembretes, janela_lembretes)

    lista_lembretes.bind("<Button-3>", mostrar_menu_contexto)  # Clique direito
    lista_lembretes.bind("<Double-Button-1>", duplo_clique_editar)  # Duplo clique esquerdo

    estilo_botao_padrao = {
        "bg": BUTTON_PRIMARY_BG,
        "fg": BUTTON_PRIMARY_FG,
        "activebackground": BUTTON_PRIMARY_ACTIVE_BG,
        "activeforeground": BUTTON_PRIMARY_ACTIVE_FG,
        "font": ("Segoe UI", 10, "bold"),
        "relief": tk.FLAT,
        "bd": 0,
        "highlightthickness": 1,
        "highlightbackground": BUTTON_PRIMARY_HIGHLIGHT,
    }
    btn_visualizar = tk.Button(
        frame_botoes,
        text="Visualizar",
        command=visualizar_com_debounce,
        width=10,
        **estilo_botao_padrao
    )
    btn_visualizar.pack(side=tk.LEFT, padx=5)
    btn_selecionar = tk.Button(
        frame_botoes,
        text="Selecionar Todos",
        command=selecionar_ou_desmarcar,
        width=16,
        **estilo_botao_padrao
    )
    btn_selecionar.pack(side=tk.LEFT, padx=5)

    btn_editar = tk.Button(
        frame_botoes,
        text="Editar",
        command=editar_selecionado,
        **estilo_botao_padrao
    )
    btn_editar.pack(side=tk.LEFT, padx=5)

    estilo_excluir = dict(estilo_botao_padrao)
    estilo_excluir.update({
        "bg": BUTTON_DANGER_BG,
        "fg": BUTTON_DANGER_FG,
        "activebackground": BUTTON_DANGER_ACTIVE_BG,
        "activeforeground": BUTTON_DANGER_ACTIVE_FG,
        "highlightbackground": BUTTON_DANGER_HIGHLIGHT,
        "highlightthickness": 2
    })
    btn_excluir = tk.Button(
        frame_botoes,
        text="Excluir",
        command=excluir_selecionados,
        width=10,
        **estilo_excluir
    )
    btn_excluir.pack(side=tk.LEFT, padx=5)

    # Inicializa o rótulo do botão Selecionar/Desmarcar conforme estado atual
    update_select_button_label()

    # Carrega os lembretes inicialmente
    carregar_lembretes()
    janela_lembretes.focus_set()


def adicionar_lembrete_atual():
    """Adiciona o texto atual como lembrete."""
    texto = entrada_descricao.get("1.0", "end-1c").strip()
    if texto:
        # Verifica se há data no texto antes de tentar registrar
        padrao_data = re.search(r'(\d{2}/\d{2}/\d{4})|(\d{2}/\d{2})', texto)
        if padrao_data:
            registrar_promessas(texto, "Lembrete")
            carregar_lista()
            messagebox.showinfo("Sucesso", "Lembrete adicionado com sucesso!")
        else:
            messagebox.showwarning("Aviso",
                                   "Para criar um lembrete, você deve especificar uma data no formato dd/mm ou dd/mm/aaaa no texto.")
    atualizar_cor_botao_lembrete()


def atualizar_cor_botao_lembrete():
    """Atualiza a cor do botão Lembretes baseado na existência de lembretes com cache"""
    try:
        # Sempre consulta o banco diretamente para garantir dados atualizados
        cursor.execute("SELECT COUNT(*) FROM promessas WHERE pessoa = 'Lembrete'")
        count = cursor.fetchone()[0]

        # Atualiza o cache com o novo valor
        cache.set('count_promessas', count, 300)

        # Atualiza a cor do botão baseado na contagem
        if count > 0:
            botao_lembrete.config(bg="#ECEFF1", fg="#e74c3c")  # Texto vermelho se houver lembretes
        else:
            botao_lembrete.config(bg="#ECEFF1", fg="#37474F")  # Texto cinza se não houver

        print(f"[DEBUG] Lembretes encontrados: {count}, Cor atualizada")
    except Exception as e:
        # Em caso de erro, usa cor padrão
        botao_lembrete.config(bg="#ECEFF1", fg="#37474F")
        print(f"Erro ao atualizar cor do botão lembretes: {e}")


def atualizar_cor_botao_lembrete_checkbox():
    """Atualiza a cor do botão Lembretes baseado no estado do checkbox e conteúdo das observações"""
    if lembrete_var.get() and entrada_descricao.get("1.0", "end-1c").strip():
        botao_lembrete.config(fg="#FF0000")  # Amarelo quando checkbox ativo e com texto
    else:
        # Volta à cor determinada pela função original (vermelho ou cinza)
        atualizar_cor_botao_lembrete()


# --- FUNÇÕES AUXILIARES ---
def validar_data(data: str) -> bool:
    """
    Valida datas no formato DD/MM/AAAA.
    Rejeita datas fora do calendário (ex: 32/13/2022).
    """
    if not data or not isinstance(data, str):
        return False

    if not re.match(r"^(0[1-9]|[12][0-9]|3[01])/(0[1-9]|1[0-2])/\d{4}$", data):
        return False

    try:
        datetime.strptime(data, "%d/%m/%Y")
        return True
    except ValueError:
        return False


def checar_data_entry(entry_widget):
    """
    Verifica se a data no campo de entrada está no formato correto.
    Se não estiver, limpa o campo ou tenta corrigir o formato.
    Também valida as regras de negócio para datas de recebimento e devolução.
    """
    data = entry_widget.get().strip()
    if not data:
        return  # Campo vazio, nada a fazer

    # Tenta formatar números sem separadores (DDMMAAAA)
    apenas_numeros = ''.join(c for c in data if c.isdigit())
    if len(apenas_numeros) == 8 and '/' not in data:
        dia = apenas_numeros[0:2]
        mes = apenas_numeros[2:4]
        ano = apenas_numeros[4:8]
        data_corrigida = f"{dia}/{mes}/{ano}"
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, data_corrigida)
        data = data_corrigida

    # Tenta corrigir formatos comuns de data
    # Formato: DD-MM-AAAA para DD/MM/AAAA
    elif re.match(r"^(0[1-9]|[12][0-9]|3[01])-(0[1-9]|1[0-2])-\d{4}$", data):
        partes = data.split('-')
        data_corrigida = f"{partes[0]}/{partes[1]}/{partes[2]}"
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, data_corrigida)
        data = data_corrigida

    # Formato: AAAA-MM-DD para DD/MM/AAAA
    elif re.match(r"^\d{4}-(0[1-9]|1[0-2])-(0[1-9]|[12][0-9]|3[01])$", data):
        partes = data.split('-')
        data_corrigida = f"{partes[2]}/{partes[1]}/{partes[0]}"
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, data_corrigida)
        data = data_corrigida

    # Verifica se está no formato correto DD/MM/AAAA
    if not validar_data(data):
        messagebox.showwarning("Formato de data inválido",
                               "Por favor, use o formato DD/MM/AAAA.\nExemplo: 01/01/2023")
        entry_widget.delete(0, tk.END)
        entry_widget.focus_set()
        return

    # Validações de regras de negócio
    validar_regras_datas(entry_widget)


def validar_regras_datas(entry_widget_atual):
    """
    Valida as regras de negócio para datas:
    1. A data de recebimento não pode ser anterior à devolução
    2. A devolução não pode ser posterior ao dia atual
    3. A data de recebimento não pode ser posterior ao dia atual
    """
    try:
        data_recebimento = entrada_recebimento.get().strip()
        data_devolucao = entrada_devolucao.get().strip()
        hoje = DateUtils.obter_data_atual()

        # Converte strings para objetos datetime para comparação
        if data_recebimento and validar_data(data_recebimento):
            recebimento_obj = datetime.strptime(data_recebimento, "%d/%m/%Y")
        else:
            recebimento_obj = None

        if data_devolucao and validar_data(data_devolucao):
            devolucao_obj = datetime.strptime(data_devolucao, "%d/%m/%Y")
        else:
            devolucao_obj = None

        hoje_obj = datetime.strptime(hoje, "%d/%m/%Y")

        # Regra 1: Data de recebimento não pode ser anterior à devolução
        if recebimento_obj and devolucao_obj:
            if recebimento_obj > devolucao_obj:
                messagebox.showerror("Erro de Validação",
                                     "A data de recebimento não pode ser posterior à data de devolução!")
                entry_widget_atual.delete(0, tk.END)
                entry_widget_atual.focus_set()
                return False

        # Regra 2: Devolução não pode ser posterior ao dia atual
        if devolucao_obj and devolucao_obj > hoje_obj:
            messagebox.showerror("Erro de Validação",
                                 "A data de devolução não pode ser posterior ao dia atual!")
            if entry_widget_atual == entrada_devolucao:
                entry_widget_atual.delete(0, tk.END)
                entry_widget_atual.focus_set()
            return False

        # ✅ NOVA REGRA 3: Data de recebimento não pode ser posterior ao dia atual
        if recebimento_obj and recebimento_obj > hoje_obj:
            messagebox.showerror("Erro de Validação",
                                 "A data de recebimento não pode ser posterior ao dia atual!")
            if entry_widget_atual == entrada_recebimento:
                entry_widget_atual.delete(0, tk.END)
                entry_widget_atual.focus_set()
            return False

        return True

    except Exception as e:
        print(f"[ERRO] Validação de datas: {e}")
        return False


def validar_campos_obrigatorios(numero_processo, secretaria, data_inicio, data_entrega=None) -> bool:
    """
    Verifica se os campos obrigatórios estão preenchidos e se as datas são válidas e lógicas.
    - Campos obrigatórios: número_processo, secretaria, data_inicio
    - Formato das datas: DD/MM/AAAA
    - Lógica das datas: devolução não pode ser antes do recebimento nem no futuro
    """
    # Verifica campos obrigatórios
    if not numero_processo.strip() or not secretaria.strip() or not data_inicio.strip():
        messagebox.showwarning("Campos obrigatórios", "Preencha todos os campos obrigatórios!")
        return False

    # Valida data de início
    if not validar_data(data_inicio):
        messagebox.showerror("Data inválida", "Data de recebimento inválida. Use o formato DD/MM/AAAA.")
        entrada_recebimento.focus_set()
        return False

    # Verifica campo "Entregue por"
    if not entrada_entregue_por.get().strip():
        messagebox.showwarning("Campo obrigatório", "Preencha o campo 'Entregue por'.")
        entrada_entregue_por.focus_set()
        return False

    # Verifica a lógica das datas se data_entrega estiver preenchida
    if data_entrega and data_entrega.strip():
        if not validar_data(data_entrega):
            messagebox.showerror("Data inválida", "Data de devolução inválida. Use o formato DD/MM/AAAA.")
            entrada_devolucao.focus_set()
            return False

    try:
        # Sempre valida e converte a data de início (já garantida acima por validar_data)
        data_inicio_dt = datetime.strptime(data_inicio, "%d/%m/%Y")

        # Só tenta converter e comparar a data de entrega se o campo estiver preenchido
        hoje = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        if data_entrega and data_entrega.strip():
            data_entrega_dt = datetime.strptime(data_entrega, "%d/%m/%Y")

            if data_entrega_dt < data_inicio_dt:
                messagebox.showerror("Erro de data", "A devolução não pode ser anterior ao recebimento.")
                entrada_devolucao.focus_set()
                return False

            if data_entrega_dt > hoje:
                messagebox.showerror("Erro de data", "A devolução não pode ser no futuro.")
                entrada_devolucao.focus_set()
                return False

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao validar datas: {e}")
        return False

    return True


popup_erro = None  # Variável global


def validar_entregue_por(event=None):
    """Valida se o campo 'Entregue por' está preenchido e mostra um aviso visual se estiver vazio."""
    if not entrada_entregue_por.get().strip():
        entrada_entregue_por.config(bg="#ffd6d6")
        mostrar_popup_erro(entrada_entregue_por, "Campo obrigatório!")
    else:
        entrada_entregue_por.config(bg="white")

    # tk.Não é necessário vincular o evento novamente aqui, pois isso causaria vinculações múltiplas
    # e potencialmente chamadas recursivas


def mostrar_popup_erro(entry, mensagem):
    """Exibe um popup de erro próximo ao widget de entrada especificado.

    Args:
        entry: O widget de entrada ao qual o popup será associado
        mensagem: A mensagem de erro a ser exibida
    """
    global popup_erro

    # Fecha qualquer popup existente antes de criar um novo
    if popup_erro:
        popup_erro.destroy()

    # Calcula a posição do popup em relação ao widget de entrada
    entry_x = entry.winfo_rootx()
    entry_y = entry.winfo_rooty()
    entry_w = entry.winfo_width()
    entry_h = entry.winfo_height()

    # Configuração do tamanho e posição do popup
    largura_popup = 200
    x = entry_x + (entry_w // 2) - (largura_popup // 2)
    y = entry_y + entry_h + 10  # 10px abaixo do widget

    # Cria o popup
    popup_erro = tk.Toplevel(entry)
    popup_erro.wm_overrideredirect(True)  # Remove a barra de título
    popup_erro.configure(bg="#ffdddd", padx=6, pady=3, borderwidth=1, relief="solid")
    popup_erro.geometry(f"+{x}+{y}")

    # Adiciona a mensagem de erro ao popup
    label = tk.Label(
        popup_erro,
        text=mensagem,
        bg="#ffdddd",
        fg="black",
        font=("Segoe UI", 9),
        justify="center",
        wraplength=largura_popup
    )
    label.pack()

    # Configura o popup para fechar quando o mouse sair dele
    popup_erro.bind("<Leave>", lambda e: popup_erro.destroy())


def extrair_promessas(texto):
    """Extrai promessas do formato 'prometeu que dia DD/MM' ou similar.

    Args:
        texto: O texto a ser analisado para extração de promessas

    Returns:
        Uma lista de tuplas (data_str, promessa) onde data_str está no formato DD/MM
        e promessa é a descrição da promessa extraída
    """
    if not texto or not isinstance(texto, str):
        return []

    # Padrões para identificar promessas no texto
    padroes = [
        r'prometeu que dia (\d{2}/\d{2}) (.*)',  # "prometeu que dia 15/07 entregaria"
        r'afirmou que dia (\d{2}/\d{2}) (.*)',  # "afirmou que dia 20/07 enviaria"
        r'dia (\d{2}/\d{2}) (.*)',  # "dia 25/07 mandará os documentos"
        r'ficou de (.*) dia (\d{2}/\d{2})',  # "ficou de entregar dia 30/07"
        r'(\d{2}/\d{2})\s*[:-]\s*(.*)'  # "30/07: entregará os documentos"
    ]

    promessas = []
    ano_atual = datetime.now().year

    for padrao in padroes:
        matches = re.finditer(padrao, texto, re.IGNORECASE)
        for match in matches:
            # Os dois últimos padrões têm grupos diferentes
            if 'ficou de' in padrao:
                promessa = match.group(1).strip()
                data_str = match.group(2)
            elif '[:-]' in padrao:
                data_str = match.group(1)
                promessa = match.group(2).strip()
            else:
                data_str = match.group(1)
                promessa = match.group(2).strip()

            # Adiciona o ano atual para facilitar comparações futuras
            promessas.append((data_str, promessa))

    return promessas


def verificar_promessas_do_dia():
    """Verifica e notifica sobre promessas agendadas para o dia atual."""
    try:
        hoje = DateUtils.obter_data_atual()
    except Exception:
        hoje = datetime.now().strftime("%d/%m/%Y")

        # Busca APENAS lembretes não notificados para hoje, excluindo os com data padrão (01/01/2000)
        cursor.execute('''
            SELECT descricao, pessoa FROM promessas 
            WHERE data_prometida = ? AND notificado = 0 AND pessoa = 'Lembrete' AND data_prometida != '01/01/2000'
        ''', (hoje,))

        promessas = cursor.fetchall()

        if promessas:
            mensagens = []
            for descricao, pessoa in promessas:
                mensagem = f"Lembrete: {descricao}"
                mensagens.append(mensagem)

                # Marca cada promessa como notificada
                cursor.execute('''
                    UPDATE promessas SET notificado = 1 
                    WHERE descricao = ? AND pessoa = ? AND data_prometida = ?
                ''', (descricao, pessoa, hoje))

            conn.commit()
            messagebox.showinfo("Lembretes do Dia", "\n\n".join(mensagens))

        # Verifica também lembretes atrasados
        verificar_promessas_atrasadas()

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao verificar promessas: {e}")
        print(f"[ERRO] Verificação de promessas: {e}")


def verificar_promessas_atrasadas():
    """Verifica se existem lembretes não cumpridos de dias anteriores."""
    try:
        try:
            hoje = DateUtils.obter_data_atual()
        except Exception:
            hoje = datetime.now().strftime("%d/%m/%Y")

        # Busca APENAS lembretes de dias anteriores que não foram notificados e não são da data padrão
        cursor.execute('''
            SELECT data_prometida, descricao, pessoa FROM promessas 
            WHERE data_prometida < ? AND notificado = 0 AND pessoa = 'Lembrete' AND data_prometida != '01/01/2000'
            ORDER BY data_prometida
        ''', (hoje,))

        promessas_atrasadas = cursor.fetchall()

        if promessas_atrasadas:
            mensagens = []
            for data, descricao, pessoa in promessas_atrasadas:
                mensagem = f"{data}: {descricao}"
                mensagens.append(mensagem)

            messagebox.showwarning("Lembretes Atrasados",
                                   "Os seguintes lembretes estão atrasados:\n\n" +
                                   "\n".join(mensagens))
    except Exception as e:
        print(f"[ERRO] Verificação de lembretes atrasados: {e}")


def cadastrar_processo():
    global nomes_autocomplete
    try:
        # Verifica se é um lembrete
        if lembrete_var.get():
            descricao = entrada_descricao.get("1.0", "end-1c").strip()
            if not descricao:
                messagebox.showwarning("Aviso", "Para criar um lembrete, é necessário preencher o campo Observações!")
                return

            # Registra o lembrete
            inseridos_msg = "Lembrete cadastrado com sucesso!"
            # Se não houver data explícita no texto, avisar uso da data padrão
            if not re.search(r"\b(\d{1,2})/(\d{1,2})(/(\d{2,4}))?\b", descricao):
                inseridos_msg = "Lembrete cadastrado"

            registrar_promessas(descricao, "Lembrete")
            messagebox.showinfo("Sucesso", inseridos_msg)
            # Limpa apenas a descrição e desmarca o checkbox
            entrada_descricao.delete("1.0", tk.END)
            lembrete_var.set(False)
            entrada_descricao.config(bg="white")
            botao_cadastrar.config(text="Cadastrar")  # Restaura o texto do botão
            # Atualiza a cor do botão de lembretes
            atualizar_cor_botao_lembrete()
            # Atualiza a janela de lembretes se estiver aberta
            try:
                carregar_lembretes_existente()
            except Exception:
                pass
            return

        # Se não for lembrete, prossegue com o cadastro normal
        # Coleta os dados [mantido igual]
        numero_processo = entrada_numero.get().strip()
        secretaria = entrada_secretaria.get().split(' - ')[0] if entrada_secretaria.get() else ''
        numero_licitacao = entrada_licitacao.get().strip()
        modalidade = entrada_modalidade.get()
        situacao = situacao_var.get()
        data_inicio = entrada_recebimento.get().strip()
        data_entrega = entrada_devolucao.get().strip()
        descricao = entrada_descricao.get("1.0", "end-1c").strip()
        entregue_por = entrada_entregue_por.get().strip().upper()
        devolvido_a = entrada_devolvido_a.get().strip().upper()
        contratado = entrada_contratado.get().strip().upper()  # Novo campo

        # Validação e lógica principal [mantido igual até a inserção]
        if not validar_campos_obrigatorios(numero_processo, secretaria, data_inicio, data_entrega):
            return

        data_inicio_db = converter_data_para_banco(data_inicio, "Data de recebimento")
        if not data_inicio_db:
            return

        data_entrega_db = None
        if data_entrega:
            data_entrega_db = converter_data_para_banco(data_entrega, "Data de devolução")
            if not data_entrega_db:
                return

        # Verificação de processo excluído [mantido igual]
        cursor.execute('SELECT COUNT(*) FROM trabalhos_excluidos WHERE numero_processo = ?', (numero_processo,))
        if cursor.fetchone()[0] > 0:
            resposta = messagebox.askyesno("Processo Excluído",
                                           f"O processo '{numero_processo}' foi excluído anteriormente.\n\nDeseja restaurá-lo ou usar um número diferente?\n\n• SIM = Restaurar processo excluído\n• tk.NÃO = Cancelar e usar número diferente")
            if resposta:
                restaurar_registro_excluido(numero_processo)
                return
            else:
                messagebox.showwarning("Aviso", "Operação cancelada. Use um número de processo diferente.")
                return

        # Verificação de duplicidade [mantido igual]
        cursor.execute("SELECT 1 FROM trabalhos_realizados WHERE numero_processo = ?", (numero_processo,))
        if cursor.fetchone():
            messagebox.showwarning("Aviso", f"O processo '{numero_processo}' já está cadastrado.")
            return

        # Insere no banco (agora com 12 campos, incluindo contratado)
        cursor.execute('''
            INSERT INTO trabalhos_realizados (
                data_registro, numero_processo, secretaria, numero_licitacao,
                modalidade, data_inicio, data_entrega,
                entregue_por, devolvido_a, contratado, situacao, descricao
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'), numero_processo, secretaria, numero_licitacao,
            modalidade, data_inicio_db, data_entrega_db,
            entregue_por, devolvido_a, contratado, situacao, descricao
        ))

        conn.commit()  # APENAS UM COMMIT AQUI
        # Backup após inclusão
        try:
            backup_automatico([numero_processo])
        except Exception:
            pass

        # Atualiza cache e interface
        cache.invalidate('count_concluidos')
        cache.invalidate('count_andamento')
        cache.invalidate('nomes_autocomplete')

        recarregar_listas_autocomplete()

        messagebox.showinfo("Sucesso", "Processo cadastrado com sucesso!")
        limpar_campos()
        listar_processos()
        contar_registros()

    except sqlite3.IntegrityError as e:
        if "UNIQUE constraint failed" in str(e):
            messagebox.showerror("Erro de Duplicação",
                                 f"Já existe um processo com o número '{numero_processo}'.\n\nVerifique se não foi excluído anteriormente.")
        else:
            messagebox.showerror("Erro de Integridade", f"Erro inesperado: {str(e)}")
        print(f"[ERRO] IntegrityError: {e}")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao cadastrar: {str(e)}")
        print(f"[ERRO] Cadastro: {e}")


def registrar_promessas(texto, pessoa, data_especifica=None):
    """Registra promessas ou lembretes no banco de dados."""
    if not texto or not isinstance(texto, str):
        return

    # Para lembretes, salva mesmo sem data
    if pessoa == "Lembrete":
        data = None
        tem_data_explicita = False

        if data_especifica:
            data = data_especifica
            tem_data_explicita = True
        else:
            # Tenta extrair data do texto do lembrete
            padrao_data = re.search(r'(\d{2}/\d{2}/\d{4})', texto)
            if padrao_data:
                data = padrao_data.group(1)
                tem_data_explicita = True
            else:
                # Se não encontrar data completa, tenta dd/mm
                padrao_data_curta = re.search(r'(\d{2}/\d{2})', texto)
                if padrao_data_curta:
                    data_str = padrao_data_curta.group(1)
                    try:
                        data_obj = datetime.strptime(f"{data_str}/{datetime.now().year}", "%d/%m/%Y")
                        data = data_obj.strftime("%d/%m/%Y")
                        tem_data_explicita = True
                    except ValueError:
                        # Data inválida - usará data padrão
                        pass

        # Se não tem data explícita, usa a data vigente (data do cadastro)
        if not tem_data_explicita:
            try:
                data = DateUtils.obter_data_atual()
            except Exception:
                data = datetime.now().strftime("%d/%m/%Y")

        try:
            cursor.execute('''
                INSERT INTO promessas (data_prometida, descricao, pessoa, data_criacao)
                VALUES (?, ?, ?, datetime('now','localtime'))
            ''', (data, texto, pessoa))
            conn.commit()

            # Invalida cache de promessas
            cache.invalidate('count_promessas')
            atualizar_cor_botao_lembrete()

            if 'janela' in globals():
                janela.update_idletasks()

        except Exception as e:
            print(f"[ERRO] Falha ao registrar promessa: {e}")
    else:

        # Verificação de Lembrete
        if lembrete_var.get() and descricao:
            registrar_promessas(descricao, "Lembrete")
            entrada_descricao.delete("1.0", tk.END)
            entrada_descricao.config(bg="white")
            lembrete_var.set(False)
            botao_cadastrar.config(text="Cadastrar")  # Restaura o texto do botão
            messagebox.showinfo("Sucesso", "Lembrete cadastrado com sucesso!")
            return


def converter_data_para_banco(data_str, campo_nome):
    """Converte uma data no formato DD/MM/AAAA para YYYY-MM-DD."""
    try:
        data_obj = datetime.strptime(data_str, '%d/%m/%Y')
        return data_obj.strftime('%Y-%m-%d')
    except ValueError:
        messagebox.showerror("Erro", f"{campo_nome} inválida!")
        return None


def atualizar_lista_autocomplete(entregue_por, devolvido_a, contratado=None):
    """Atualiza listas de autocompletar e propaga por widget (separado para Contratado)."""
    global nomes_autocomplete, nomes_contratado

    atualizado = False
    # normalize upper-case
    if entregue_por:
        e = entregue_por.strip().upper()
        if e and e not in nomes_autocomplete:
            nomes_autocomplete.append(e)
            atualizado = True

    if devolvido_a:
        d = devolvido_a.strip().upper()
        if d and d not in nomes_autocomplete:
            nomes_autocomplete.append(d)
            atualizado = True

    if contratado:
        c = contratado.strip().upper()
        if c and c not in nomes_contratado:
            nomes_contratado.append(c)
            atualizado = True

    if not atualizado:
        return  # nada novo

    # Ordena e remove duplicatas
    nomes_autocomplete = sorted(set([n for n in nomes_autocomplete if n]), key=lambda x: x.lower())
    nomes_contratado = sorted(set([n for n in nomes_contratado if n]), key=lambda x: x.lower())

    # Propaga listas específicas aos widgets
    try:
        entrada_entregue_por.completion_list = nomes_autocomplete
        entrada_devolvido_a.completion_list = nomes_autocomplete
        entrada_contratado.completion_list = nomes_contratado
    except NameError:
        pass
    except Exception:
        pass

    # Invalida cache relevante
    try:
        cache.invalidate('nomes_autocomplete')
        cache.invalidate('nomes_contratado')
    except Exception:
        pass


def formatar_data_hora(event):
    """Formata automaticamente a data enquanto o usuário digita.
    Converte formatos como 07072025 para 07/07/2025.
    """
    entry = event.widget
    texto = entry.get().strip()

    # Remove caracteres não numéricos
    apenas_numeros = ''.join(c for c in texto if c.isdigit())

    # Se temos 8 dígitos (DDMMAAAA), formata como DD/MM/AAAA
    if len(apenas_numeros) == 8:
        dia = apenas_numeros[0:2]
        mes = apenas_numeros[2:4]
        ano = apenas_numeros[4:8]
        entry.delete(0, tk.END)
        entry.insert(0, f"{dia}/{mes}/{ano}")


def formatar_data_hora_str(data):
    if not data:
        return ""
    try:
        return DateUtils.formatar_data_hora(str(data))
    except Exception:
        return str(data)


def abrir_janela_restaurar():
    global janela_restaurar_instancia
    janela_restaurar = Toplevel(janela)
    janela_restaurar.title("Restaurar Registro Excluído")
    try:
        w, h, xi, yi = carregar_tamanho_janela('restaurar_registro', 480, 380)
        if xi is not None and yi is not None:
            janela_restaurar.geometry(f"{w}x{h}+{xi}+{yi}")
        else:
            center_window(janela_restaurar, w, h)
    except Exception:
        center_window(janela_restaurar, 480, 380)
    try:
        janela_restaurar.resizable(True, True)
    except Exception:
        pass

    # Registrar instância global para permitir fechamento posterior
    try:
        janela_restaurar_instancia = janela_restaurar
    except Exception:
        pass

    # Função para fechar a janela com ESC
    def fechar_janela(event=None):
        try:
            try:
                janela_restaurar.update_idletasks()
                salvar_tamanho_janela('restaurar_registro', janela_restaurar.winfo_width(),
                                      janela_restaurar.winfo_height(), janela_restaurar.winfo_x(),
                                      janela_restaurar.winfo_y())
            except Exception:
                pass
            janela_restaurar.destroy()
        finally:
            try:
                janela_restaurar_instancia = None
            except Exception:
                pass

    # Bind da tecla ESC para fechar a janela
    janela_restaurar.bind('<Escape>', fechar_janela)
    janela_restaurar.bind('<KeyPress-Escape>', fechar_janela)
    try:
        janela_restaurar.protocol("WM_DELETE_WINDOW", fechar_janela)
    except Exception:
        pass

    # Garante que a janela capture eventos de teclado
    janela_restaurar.focus_force()
    janela_restaurar.grab_set()  # Torna a janela modal

    # Lista com seleção múltipla (removendo a duplicação)
    lista = tk.Listbox(
        janela_restaurar,
        width=80,
        height=12,
        selectmode=tk.EXTENDED,
        bg="#F7F9FC",
        fg="#1F2937",
        selectbackground="#0078d4",
        selectforeground="white"
    )
    lista.pack(padx=10, pady=10, fill="both", expand=True)

    # Também adiciona o bind na lista para garantir
    lista.bind('<Escape>', fechar_janela)

    # Carrega os dados iniciais
    cursor.execute('''
        SELECT numero_processo, secretaria, data_exclusao
        FROM trabalhos_excluidos
        ORDER BY data_exclusao DESC
    ''')
    backups = cursor.fetchall()
    for proc, sec, data_exc in backups:
        nome_secretaria = secretarias_dict.get(sec, sec)
        try:
            from datetime import datetime
            s = str(data_exc)
            try:
                dt = datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
                data_fmt = dt.strftime("%d/%m/%Y %H:%M")
            except Exception:
                try:
                    dt = datetime.strptime(s, "%Y-%m-%d %H:%M")
                    data_fmt = dt.strftime("%d/%m/%Y %H:%M")
                except Exception:
                    try:
                        dt = datetime.strptime(s[:10], "%Y-%m-%d")
                        data_fmt = dt.strftime("%d/%m/%Y")
                    except Exception:
                        data_fmt = s
        except Exception:
            data_fmt = str(data_exc)
        lista.insert(tk.END, f"{proc} | {nome_secretaria} | Excluído em: {data_fmt}")

    def ajustar_tamanho_janela():
        try:
            from tkinter import font as tkfont
            lista.update_idletasks()
            itens = lista.get(0, tk.END)
            longest = ""
            for it in itens:
                if len(it) > len(longest):
                    longest = it
            f = tkfont.nametofont(lista.cget("font"))
            text_w = f.measure(longest) if longest else 0
            pad_w = 60
            w = max(480, min(text_w + pad_w, janela_restaurar.winfo_screenwidth() - 80))
            line_h = f.metrics("linespace")
            vis_lines = min(max(lista.size(), 8), 20)
            list_h = line_h * vis_lines + 80
            buttons_h = 70
            h = max(380, min(list_h + buttons_h, janela_restaurar.winfo_screenheight() - 120))
            lista.config(width=max(len(longest) + 2, 80), height=vis_lines)
            center_window(janela_restaurar, int(w), int(h))
        except Exception:
            pass

    ajustar_tamanho_janela()

    def restaurar_selecionado():
        indices = lista.curselection()
        if not indices:
            messagebox.showwarning("Aviso", "Selecione um ou mais registros para restaurar.")
            return

        processos_restaurados = []
        erros = []

        for idx in indices:
            try:
                numero_processo = backups[idx][0]
                restaurar_registro_excluido(numero_processo)
                processos_restaurados.append(numero_processo)
            except Exception as e:
                erros.append(f"Erro ao restaurar {backups[idx][0]}: {str(e)}")

        # Mostra resultado
        if processos_restaurados:
            messagebox.showinfo("Sucesso", f"Restaurados {len(processos_restaurados)} registro(s).")

            # ✅ CORREÇÃO: Fecha a janela após restauração bem-sucedida
            try:
                janela_restaurar.destroy()
            finally:
                try:
                    janela_restaurar_instancia = None
                except Exception:
                    pass
            return  # Sai da função após fechar a janela

        if erros:
            messagebox.showerror("Erros", "\n".join(erros))

        # Atualiza a lista apenas se não fechou a janela (em caso de erro)
        atualizar_lista_restaurar()

    def visualizar_selecionado():
        indices = lista.curselection()
        # Sem seleção: se existir um único item, auto-seleciona; se houver vários, abre todos
        if not indices:
            total = lista.size()
            if total == 0:
                messagebox.showwarning("Aviso", "Nenhum registro disponível para visualizar.")
                return
            if total == 1:
                indices = (0,)
            else:
                # Abre a janela de visualização com todos os registros listados
                processos_selecionados = [item[0] for item in backups]
                visualizar_registros_excluidos(processos_selecionados)
                return

        # Coleta os números dos processos selecionados
        processos_selecionados = [backups[idx][0] for idx in indices]
        visualizar_registros_excluidos(processos_selecionados)

    def excluir_permanentemente():
        indices = lista.curselection()
        if not indices:
            messagebox.showwarning("Aviso", "Selecione um ou mais registros para excluir permanentemente.")
            return

        quantidade = len(indices)
        if quantidade == 1:
            mensagem = f"Tem certeza que deseja excluir PERMANENTEMENTE o processo {backups[indices[0]][0]}?"
        else:
            mensagem = f"Tem certeza que deseja excluir PERMANENTEMENTE {quantidade} processos selecionados?"

        mensagem += "\n\nEsta ação não pode ser desfeita!"

        if messagebox.askyesno("Confirmar Exclusão Permanente", mensagem):
            processos_excluidos = []
            erros = []

            for idx in indices:
                try:
                    numero_processo = backups[idx][0]
                    cursor.execute('DELETE FROM trabalhos_excluidos WHERE numero_processo = ?', (numero_processo,))
                    processos_excluidos.append(numero_processo)
                except Exception as e:
                    erros.append(f"Erro ao excluir {backups[idx][0]}: {str(e)}")

            if processos_excluidos:
                conn.commit()
                messagebox.showinfo("Sucesso", f"Excluídos permanentemente {len(processos_excluidos)} registro(s).")
                # Realiza backup após alteração
                try:
                    backup_automatico(processos_excluidos)
                except Exception:
                    pass
                # Fecha a janela após exclusão permanente
                try:
                    janela_restaurar.destroy()
                finally:
                    try:
                        janela_restaurar_instancia = None
                    except Exception:
                        pass
                return
            if erros:
                messagebox.showerror("Erros", "\n".join(erros))

            # Atualiza a lista (apenas se não fechou a janela)
            atualizar_lista_restaurar()

    def atualizar_lista_restaurar():
        lista.delete(0, tk.END)
        cursor.execute('''
            SELECT numero_processo, secretaria, data_exclusao
            FROM trabalhos_excluidos
            ORDER BY data_exclusao DESC
        ''')
        nonlocal backups
        backups = cursor.fetchall()
        for proc, sec, data_exc in backups:
            nome_secretaria = secretarias_dict.get(sec, sec)
            try:
                from datetime import datetime
                s = str(data_exc)
                try:
                    dt = datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
                    data_fmt = dt.strftime("%d/%m/%Y %H:%M")
                except Exception:
                    try:
                        dt = datetime.strptime(s, "%Y-%m-%d %H:%M")
                        data_fmt = dt.strftime("%d/%m/%Y %H:%M")
                    except Exception:
                        try:
                            dt = datetime.strptime(s[:10], "%Y-%m-%d")
                            data_fmt = dt.strftime("%d/%m/%Y")
                        except Exception:
                            data_fmt = s
            except Exception:
                data_fmt = str(data_exc)
            lista.insert(tk.END, f"{proc} | {nome_secretaria} | Excluído em: {data_fmt}")

        try:
            ajustar_tamanho_janela()
        except Exception:
            pass

    def alternar_selecao():
        """Alterna entre selecionar todos e limpar seleção."""
        if lista.curselection():
            # Se há itens selecionados, limpa a seleção
            lista.selection_clear(0, tk.END)
            btn_selecionar_todos.config(text="Selecionar Todos")
        else:
            # Se não há itens selecionados, seleciona todos
            lista.selection_set(0, tk.END)
            btn_selecionar_todos.config(text="Limpar Seleção")

    # Frame para os botões
    frame_botoes = tk.Frame(janela_restaurar, bg="#ECEFF1")
    frame_botoes.pack(pady=5)

    # Botões na ordem solicitada: Visualizar, Selecionar Todos, Restaurar Selecionado, Excluir
    btn_visualizar = tk.Button(frame_botoes, text="Visualizar", command=visualizar_selecionado)
    btn_visualizar.pack(side=tk.LEFT, padx=5)

    btn_selecionar_todos = tk.Button(frame_botoes, text="Selecionar Todos", command=alternar_selecao)
    btn_selecionar_todos.pack(side=tk.LEFT, padx=5)

    btn_restaurar = tk.Button(frame_botoes, text="Restaurar Selecionado", command=restaurar_selecionado)
    btn_restaurar.pack(side=tk.LEFT, padx=5)

    btn_excluir = tk.Button(frame_botoes, text="Excluir", command=excluir_permanentemente,
                            bg=BUTTON_DANGER_BG, fg=BUTTON_DANGER_FG, activebackground=BUTTON_DANGER_ACTIVE_BG,
                            activeforeground=BUTTON_DANGER_ACTIVE_FG,
                            relief=tk.FLAT, bd=0, font=("Segoe UI", 10, "bold"), highlightthickness=2,
                            highlightbackground=BUTTON_DANGER_HIGHLIGHT)
    btn_excluir.pack(side=tk.LEFT, padx=5)

    # Criação do menu de contexto
    menu_contexto = tk.Menu(janela_restaurar, tearoff=0)
    menu_contexto.add_command(label="Visualizar", command=visualizar_selecionado)
    menu_contexto.add_command(label="Selecionar Todos", command=alternar_selecao)
    menu_contexto.add_command(label="Restaurar Selecionado", command=restaurar_selecionado)
    menu_contexto.add_separator()
    menu_contexto.add_command(label="Excluir", command=excluir_permanentemente)

    def mostrar_menu_contexto(event):
        try:
            # tk.Não limpa a seleção se já há itens selecionados
            if not lista.curselection():
                lista.selection_clear(0, tk.END)
                lista.selection_set(lista.nearest(event.y))
            menu_contexto.post(event.x_root, event.y_root)
        except:
            pass

    def duplo_clique(event):
        """Mostra o menu de contexto no duplo clique."""
        try:
            if not lista.curselection():
                lista.selection_clear(0, tk.END)
                lista.selection_set(lista.nearest(event.y))
            menu_contexto.post(event.x_root, event.y_root)
        except:
            pass

    lista.bind("<Button-3>", mostrar_menu_contexto)  # Botão direito
    lista.bind("<Double-Button-1>", duplo_clique)  # Duplo clique esquerdo


def visualizar_registros_excluidos(numeros_processos):
    """Visualiza múltiplos registros excluídos em uma única janela com navegação."""
    if not numeros_processos:
        return

    # Fecha a janela "Restaurar Registro Excluído" se estiver aberta
    try:
        global janela_restaurar_instancia
        if 'janela_restaurar_instancia' in globals() and janela_restaurar_instancia and janela_restaurar_instancia.winfo_exists():
            try:
                janela_restaurar_instancia.destroy()
            finally:
                janela_restaurar_instancia = None
    except Exception:
        pass

    # Busca todos os registros
    registros = []
    for numero in numeros_processos:
        cursor.execute('''
            SELECT data_registro, numero_processo, secretaria, numero_licitacao,
                   situacao, modalidade, data_inicio, data_entrega,
                   entregue_por, devolvido_a, contratado, descricao, data_exclusao
            FROM trabalhos_excluidos
            WHERE numero_processo = ?
        ''', (numero,))
        registro = cursor.fetchone()
        if registro:
            registros.append(registro)

    if not registros:
        messagebox.showerror("Erro", "Nenhum registro encontrado.")
        return

    # Se já existe uma janela aberta, apenas traz para frente
    try:
        if janela_visualizar_excluida_ativa and janela_visualizar_excluida_ativa.winfo_exists():
            janela_visualizar_excluida_ativa.lift()
            janela_visualizar_excluida_ativa.focus_set()
            return
    except Exception:
        pass

    # Cria janela de visualização
    janela_vis = Toplevel(janela)
    try:
        janela_vis.withdraw()
    except Exception:
        pass
    janela_vis.title(f"Visualizar Registros Excluídos ({len(registros)} registro(s))")
    base_width, base_height, xi, yi = carregar_tamanho_janela('visualizar_excluidos', 520, 400)
    if xi is not None and yi is not None:
        janela_vis.geometry(f"{base_width}x{base_height}+{xi}+{yi}")
    else:
        janela_vis.geometry(f"{base_width}x{base_height}")
    try:
        janela_vis.resizable(True, True)
    except Exception:
        pass
    bg_color = "#F5F7FA"
    janela_vis.configure(bg=bg_color)
    try:
        janela.update_idletasks()
        px = janela.winfo_rootx()
        py = janela.winfo_rooty()
        pw = janela.winfo_width()
        ph = janela.winfo_height()
        cx = px + (pw // 2) - (base_width // 2)
        cy = py + (ph // 2) - (base_height // 2)
        if xi is None or yi is None:
            janela_vis.geometry(f"{base_width}x{base_height}+{cx}+{cy}")
    except Exception:
        pass
    try:
        def _fechar_vis():
            try:
                janela_vis.update_idletasks()
                salvar_tamanho_janela('visualizar_excluidos', janela_vis.winfo_width(), janela_vis.winfo_height(),
                                      janela_vis.winfo_x(), janela_vis.winfo_y())
            except Exception:
                pass
            janela_vis.destroy()

        janela_vis.protocol("WM_DELETE_WINDOW", _fechar_vis)
    except Exception:
        pass

    # Variável para controlar o registro atual
    indice_atual = tk.IntVar(value=0)

    # Frame de navegação
    frame_nav = tk.Frame(janela_vis, bg=bg_color)
    frame_nav.pack(fill="x", padx=10, pady=(5, 0))

    def atualizar_navegacao():
        idx = indice_atual.get()
        btn_anterior.config(state="normal" if idx > 0 else "disabled")
        btn_proximo.config(state="normal" if idx < len(registros) - 1 else "disabled")
        lbl_contador.config(text=f"Registro {idx + 1} de {len(registros)}")

    def ir_anterior():
        if indice_atual.get() > 0:
            indice_atual.set(indice_atual.get() - 1)
            atualizar_campos()
            atualizar_navegacao()

    def ir_proximo():
        if indice_atual.get() < len(registros) - 1:
            indice_atual.set(indice_atual.get() + 1)
            atualizar_campos()
            atualizar_navegacao()

    btn_anterior = tk.Button(frame_nav, text="◀ Anterior", command=ir_anterior)
    btn_anterior.pack(side=tk.LEFT, padx=5)

    lbl_contador = tk.Label(frame_nav, text="", font=("Arial", 10, "bold"), bg=bg_color)
    lbl_contador.pack(side=tk.LEFT, expand=True)

    btn_proximo = tk.Button(frame_nav, text="Próximo ▶", command=ir_proximo)
    btn_proximo.pack(side=tk.RIGHT, padx=5)

    # Frame principal com scroll
    main_frame = tk.Frame(janela_vis, bg=bg_color)
    main_frame.pack(expand=True, fill="both", padx=10, pady=10)
    try:
        main_frame.grid_columnconfigure(0, weight=0)
        main_frame.grid_columnconfigure(1, weight=1, minsize=50)
    except Exception:
        pass

    # Labels para os campos (serão atualizados)
    labels_valores = {}
    labels_chaves = {}

    campos_info = [
        ("Data de Registro:", 0),
        ("Número do Processo:", 1),
        ("Secretaria:", 2),
        ("Número da Licitação:", 3),
        ("Situação:", 4),
        ("Modalidade:", 5),
        ("Data de Recebimento:", 6),
        ("Data de Devolução:", 7),
        ("Entregue por:", 8),
        ("Devolvido a:", 9),
        ("Contratado:", 10),
        ("Data de Exclusão:", 12)
    ]

    row = 0
    for label_text, campo_idx in campos_info:
        lbl_key = tk.Label(main_frame, text=label_text, font=("Arial", 10, "bold"), anchor="w", bg=bg_color)
        lbl_key.grid(row=row, column=0, sticky="w", padx=5, pady=2)
        labels_chaves[campo_idx] = lbl_key
        lbl_valor = tk.Label(main_frame, text="", font=("Arial", 10), anchor="w", bg=bg_color)
        lbl_valor.grid(row=row, column=1, sticky="ew", padx=5, pady=2)
        labels_valores[campo_idx] = lbl_valor
        row += 1

    # Descrição (campo maior)
    tk.Label(main_frame, text="Descrição:", font=("Arial", 10, "bold"), anchor="w", bg=bg_color).grid(
        row=row, column=0, sticky="nw", padx=5, pady=2
    )

    # Observações/Descrição: aumentar uma linha na altura
    # Ajuste proporcional da altura da descrição
    descricao_text = tk.Text(main_frame, height=2, width=0, wrap="word", state="normal", bd=0, highlightthickness=0,
                             bg="#f7f7f7")
    try:
        enhance_text_controls(descricao_text)
    except Exception:
        pass
    # Tag para texto de alerta (SEM REGISTRO)
    descricao_text.tag_configure("alert", foreground="red")
    descricao_text.grid(row=row, column=1, sticky="ew", padx=5, pady=2)

    def ajustar_largura_secretaria():
        try:
            janela_vis.update_idletasks()
            val_label = labels_valores.get(2)
            left_px = val_label.winfo_rootx() - janela_vis.winfo_rootx()
            needed_w = left_px + val_label.winfo_reqwidth() + 30
            screen_w = janela_vis.winfo_screenwidth()
            min_w = base_width
            final_w = min(max(min_w, needed_w), screen_w - 20)
            current_h = janela_vis.winfo_height() or base_height
            janela_vis.geometry(f"{int(final_w)}x{int(current_h)}")
            return
        except Exception:
            pass
        try:
            import tkinter.font as tkfont
            secretaria_text = labels_valores.get(2).cget("text")
            f_val = tkfont.Font(font=("Arial", 10))
            f_lbl = tkfont.Font(font=("Arial", 10, "bold"))
            val_px = f_val.measure(str(secretaria_text))
            lbl_px = f_lbl.measure("Secretaria:")
            padding_total = 120
            needed_w = lbl_px + val_px + padding_total
            screen_w = janela_vis.winfo_screenwidth()
            min_w = base_width
            final_w = min(max(min_w, needed_w), screen_w - 20)
            current_h = janela_vis.winfo_height() or base_height
            janela_vis.geometry(f"{int(final_w)}x{int(current_h)}")
        except Exception:
            pass

    def atualizar_campos():
        """Atualiza os campos com os dados do registro atual."""
        idx = indice_atual.get()
        registro = registros[idx]

        # Atualiza título da janela
        janela_vis.title(f"Visualizar Registro Excluído - {registro[1]} ({idx + 1}/{len(registros)})")

        # Atualiza labels
        for campo_idx, label in labels_valores.items():
            valor = registro[campo_idx] if registro[campo_idx] is not None else "N/A"
            # Exibir secretaria como SIGLA - Nome completo
            if campo_idx == 2 and valor and valor != "N/A":
                try:
                    from config.settings import SECRETARIAS
                    sigla = str(valor).strip().upper()
                    nome = SECRETARIAS.get(sigla)
                    if nome:
                        valor = f"{sigla} - {nome}"
                except Exception:
                    pass
            # Formatar Data de Exclusão no padrão brasileiro
            if campo_idx == 12 and valor and valor != "N/A":
                try:
                    valor = DateUtils.formatar_data_hora(str(valor))
                except Exception:
                    pass
            label.config(text=str(valor))

        ajustar_largura_secretaria()

    # Reajusta largura após exibir, garantindo cálculo com UI pronta
    try:
        janela_vis.after(0, ajustar_largura_secretaria)
    except Exception:
        pass
        # Atualiza descrição e tamanho da janela dinamicamente
        descricao_text.config(state="normal")
        descricao_text.delete("1.0", tk.END)

        desc = registro[11] if registro[11] else ""
        tem_texto = bool(str(desc).strip())

        if tem_texto:
            # Aumenta altura em 1/3, focando a área inferior (descrição)
            nova_altura = int(base_height * 4 / 3)
            current_w = janela_vis.winfo_width() or base_width
            janela_vis.geometry(f"{current_w}x{nova_altura}")
            descricao_text.configure(height=10)
            descricao_text.insert("1.0", str(desc))
        else:
            # Mantém tamanho atual e mostra aviso em vermelho
            current_w = janela_vis.winfo_width() or base_width
            janela_vis.geometry(f"{current_w}x{base_height}")
            descricao_text.configure(height=6)
            descricao_text.insert("1.0", "SEM REGISTRO")
            descricao_text.tag_add("alert", "1.0", "end-1c")

        descricao_text.config(state="disabled")

    # Ações: Restaurar, Excluir e Reabrir Restaurar
    def restaurar_registro_atual():
        try:
            idx = indice_atual.get()
            numero = registros[idx][1]
            restaurar_registro_excluido(numero)
            try:
                listar_processos()
                contar_registros()
            except Exception:
                pass
            janela_vis.destroy()
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao restaurar: {str(e)}")

    def excluir_registro_atual():
        try:
            idx = indice_atual.get()
            numero = registros[idx][1]
            if messagebox.askyesno("Confirmar Exclusão",
                                   f"Excluir definitivamente o processo {numero}?\n\nEsta ação não pode ser desfeita!"):
                cursor.execute('DELETE FROM trabalhos_excluidos WHERE numero_processo = ?', (numero,))
                try:
                    conn.commit()
                except Exception:
                    pass
                try:
                    backup_automatico([numero])
                except Exception:
                    pass
                try:
                    listar_processos()
                    contar_registros()
                except Exception:
                    pass
                messagebox.showinfo("Excluído", f"Processo {numero} excluído permanentemente.")
                janela_vis.destroy()
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao excluir: {str(e)}")

    def reabrir_restaurar():
        try:
            abrir_janela_restaurar()
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao reabrir janela Restaurar: {str(e)}")
        finally:
            janela_vis.destroy()

    frame_botoes = tk.Frame(janela_vis, bg=bg_color)
    frame_botoes.pack(fill="x", padx=10, pady=6)
    # Botões solicitados na área marcada de verde
    tk.Button(frame_botoes, text="Restaurar Registro Excluído", command=restaurar_registro_atual).pack(side=tk.LEFT,
                                                                                                       padx=5)
    tk.Button(frame_botoes, text="Janela Registro Excluído", command=reabrir_restaurar).pack(side=tk.LEFT, padx=5)

    # Botão fechar
    def _on_close():
        try:
            globals()['janela_visualizar_excluida_ativa'] = None
        except Exception:
            pass
        try:
            janela_vis.destroy()
        except Exception:
            pass

    btn_fechar = tk.Button(janela_vis, text="Fechar", command=_on_close)
    btn_fechar.pack(side="bottom", pady=10)

    # Inicializa a visualização
    atualizar_campos()
    atualizar_navegacao()
    try:
        janela_vis.deiconify()
    except Exception:
        pass

    # Atalhos de teclado
    janela_vis.bind("<Left>", lambda e: ir_anterior())
    janela_vis.bind("<Right>", lambda e: ir_proximo())
    # ESC agora reabre a janela "Restaurar Registro Excluído" e fecha a visualização
    janela_vis.bind("<Escape>", lambda e: reabrir_restaurar())
    try:
        janela_vis.protocol("WM_DELETE_WINDOW", _on_close)
    except Exception:
        pass

    # Centraliza
    try:
        globals()['janela_visualizar_excluida_ativa'] = janela_vis
    except Exception:
        pass
    janela_vis.transient(janela)
    janela_vis.grab_set()
    janela_vis.focus_set()


def restaurar_registro_excluido(numero_processo):
    global registros_restaurados
    cursor.execute('''
        SELECT data_registro, numero_processo, secretaria, numero_licitacao,
               situacao, modalidade, data_inicio, data_entrega,
               entregue_por, devolvido_a, contratado, descricao
        FROM trabalhos_excluidos
        WHERE numero_processo = ?
    ''', (numero_processo,))
    registro = cursor.fetchone()
    if not registro:
        messagebox.showerror("Erro", "Registro não encontrado no backup.")
        return

    try:
        cursor.execute('''
            INSERT INTO trabalhos_realizados (
                data_registro, numero_processo, secretaria, numero_licitacao,
                situacao, modalidade, data_inicio, data_entrega,
                entregue_por, devolvido_a, contratado, descricao
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', registro)
        cursor.execute('DELETE FROM trabalhos_excluidos WHERE numero_processo = ?', (numero_processo,))
        conn.commit()
        # Backup após restauração
        try:
            backup_automatico([numero_processo])
        except Exception:
            pass

        # Atualiza a lista
        listar_processos()
        contar_registros()

        try:
            recarregar_listas_autocomplete()
        except Exception:
            pass

        # Encontra e destaca o registro restaurado
        for item in tabela.get_children():
            if tabela.item(item)['values'][1] == numero_processo:
                # Move para o topo
                tabela.move(item, '', 0)
                # Destaca visualmente
                tabela.selection_set(item)
                tabela.focus(item)
                tabela.see(item)
                # Configura tags para destaque
                tabela.item(item, tags=('destaque',))
                # Agendamento para remover o destaque após 3 segundos
                janela.after(3000, lambda i=item: remover_destaque(i))
                break

        registros_restaurados += 1
        atualizar_estatisticas()
        messagebox.showinfo("Sucesso", "Registro restaurado com sucesso!")

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao restaurar: {str(e)}")


def remover_destaque(item):
    """Remove o destaque do item após o tempo definido"""
    situacao = tabela.item(item)['values'][4]
    tag = 'concluido' if situacao == 'Concluído' else 'andamento' if situacao == 'Em Andamento' else ''
    tabela.item(item, tags=(tag,))
    tabela.selection_remove(item)


def piscar_registro(item):
    """Apenas destaca visualmente o item restaurado na tabela"""
    tabela.selection_set(item)
    tabela.focus(item)
    tabela.see(item)


def editar_processo():
    item_selecionado = tabela.focus()
    if not item_selecionado:
        messagebox.showwarning("Aviso", "Selecione um processo para editar!")
        return

    processo = tabela.item(item_selecionado)['values']

    # Preenche os campos com os valores do processo selecionado
    entrada_numero.delete(0, tk.END)
    entrada_numero.insert(0, processo[1])
    entrada_numero.config(state='normal')

    sigla_secretaria = processo[2]
    secretaria_formatada = next((s for s in secretarias_formatadas if s.startswith(sigla_secretaria + " - ")), "")
    entrada_secretaria.delete(0, tk.END)
    entrada_secretaria.insert(0, secretaria_formatada)
    entrada_secretaria.config(state='normal')

    entrada_licitacao.delete(0, tk.END)
    entrada_licitacao.insert(0, processo[3] if processo[3] else "")
    entrada_licitacao.config(state='normal')

    modalidade = processo[5] if len(processo) > 5 else ""
    entrada_modalidade.delete(0, tk.END)
    entrada_modalidade.insert(0, modalidade)
    entrada_modalidade.config(state='normal')

    situacao_var.set(processo[4])

    # Datas
    data_inicio = processo[6]
    try:
        data_inicio = DateUtils.para_exibicao(str(data_inicio)) if data_inicio else ""
    except Exception:
        pass
    entrada_recebimento.delete(0, tk.END)
    entrada_recebimento.insert(0, data_inicio if data_inicio else "")
    entrada_recebimento.config(state='normal')

    data_entrega = processo[7]
    if data_entrega and isinstance(data_entrega, str) and "/" not in data_entrega:
        try:
            data_entrega = DateUtils.para_exibicao(str(data_entrega))
        except Exception:
            pass
    entrada_devolucao.delete(0, tk.END)
    # Exibir vazio quando o valor for textual 'None'
    valor_devolucao = (
        "" if (isinstance(data_entrega, str) and data_entrega.strip().lower() == "none")
        else (data_entrega if data_entrega else "")
    )
    entrada_devolucao.insert(0, valor_devolucao)
    entrada_devolucao.config(state='normal')

    entrada_entregue_por.delete(0, tk.END)
    entrada_entregue_por.insert(0, processo[8] if len(processo) > 8 else "")
    entrada_entregue_por.config(state='normal')

    entrada_devolvido_a.delete(0, tk.END)
    entrada_devolvido_a.insert(0, processo[9] if len(processo) > 9 else "")
    entrada_devolvido_a.config(state='normal')

    # ✅ ADICIONADO: Preenchimento do campo Contratado
    entrada_contratado.delete(0, tk.END)
    entrada_contratado.insert(0, processo[10] if len(processo) > 10 and processo[10] else "")
    entrada_contratado.config(state='normal')

    entrada_descricao.delete("1.0", tk.END)
    entrada_descricao.insert("1.0", processo[11] if len(processo) > 11 else "")
    entrada_descricao.config(state='normal', bg="white")

    # ✅ Aqui é onde você coloca a conversão para string
    numero_processo_original = str(processo[1])

    # ✅ Atualiza o botão para "Atualizar" com o número correto
    botao_cadastrar.config(
        text="Atualizar",
        state='normal',
        command=lambda: atualizar_processo(numero_processo_original)
    )


def habilitar_campos_para_edicao():
    campos = [
        entrada_numero, entrada_secretaria, entrada_licitacao, entrada_modalidade,
        entrada_recebimento, entrada_devolucao, entrada_entregue_por, entrada_devolvido_a, entrada_contratado
    ]
    for campo in campos:
        campo.config(state='normal')

    entrada_descricao.config(state='normal', bg="white")


def atualizar_processo(numero_processo_original):
    """Atualiza um processo existente no banco de dados."""
    global registros_editados, nomes_autocomplete

    try:
        # Obter valores dos campos
        numero_processo = entrada_numero.get().strip()
        secretaria = entrada_secretaria.get().split(' - ')[0] if entrada_secretaria.get() else ''
        numero_licitacao = entrada_licitacao.get().strip()
        modalidade = entrada_modalidade.get()
        situacao = situacao_var.get()
        data_inicio = entrada_recebimento.get().strip()
        data_entrega = entrada_devolucao.get().strip()
        descricao = entrada_descricao.get("1.0", "end-1c").strip()
        entregue_por = entrada_entregue_por.get().strip().upper()
        devolvido_a = entrada_devolvido_a.get().strip().upper()
        contratado = entrada_contratado.get().strip().upper()

        # Extrai e registra novas promessas da descrição
        if descricao and entregue_por:
            registrar_promessas(descricao, entregue_por)

        # Validações
        if not validar_campos_obrigatorios(numero_processo, secretaria, data_inicio, data_entrega):
            return

        # Converte datas para o formato do banco (YYYY-MM-DD)
        data_inicio_db = converter_data_para_banco(data_inicio, "Data de recebimento")
        if not data_inicio_db:
            return

        data_entrega_db = None
        if data_entrega:
            data_entrega_db = converter_data_para_banco(data_entrega, "Data de devolução")
            if not data_entrega_db:
                return

        # Verifica se o número do processo foi alterado
        if numero_processo != numero_processo_original:
            # Se foi alterado, verifica se o novo número já existe
            cursor.execute("SELECT 1 FROM trabalhos_realizados WHERE numero_processo = ? AND numero_processo != ?",
                           (numero_processo, numero_processo_original))
            if cursor.fetchone():
                messagebox.showerror("Erro", "Número de processo já existe!")
                return

        # Atualiza o processo no banco
        cursor.execute('''
            UPDATE trabalhos_realizados
            SET numero_processo = ?,
                secretaria = ?,
                numero_licitacao = ?,
                situacao = ?,
                data_inicio = ?,
                data_entrega = ?,
                entregue_por = ?,
                devolvido_a = ?,
                contratado = ?,
                modalidade = ?,
                descricao = ?
            WHERE numero_processo = ?
        ''', (
            numero_processo, secretaria, numero_licitacao, situacao,
            data_inicio_db, data_entrega_db, entregue_por, devolvido_a,
            contratado, modalidade, descricao, numero_processo_original
        ))
        conn.commit()
        # Backup após atualização
        try:
            backup_automatico([numero_processo])
        except Exception:
            pass

        recarregar_listas_autocomplete()

        # Atualiza contadores e interface
        registros_editados += 1
        contar_registros()
        listar_processos()

        # Encontra e destaca o registro atualizado
        for item in tabela.get_children():
            if tabela.item(item)['values'][1] == numero_processo:
                # Move para o topo
                tabela.move(item, '', 0)
                # Destaca visualmente
                tabela.selection_set(item)
                tabela.focus(item)
                tabela.see(item)
                # Configura tags para destaque
                tabela.item(item, tags=('destaque',))
                # Agendamento para remover o destaque após 3 segundos
                janela.after(3000, lambda i=item: remover_destaque(i))
                break

        messagebox.showinfo("Sucesso", "Processo atualizado com sucesso!")
        manter_campos_pos_atualizacao()

    except sqlite3.IntegrityError:
        messagebox.showerror("Erro", "Conflito no banco de dados. Verifique se o número de processo já existe.")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao atualizar processo: {str(e)}")
        print(f"[ERRO] Atualização de processo: {e}")


def manter_campos_pos_atualizacao():
    """Mantém os campos preenchidos após atualização em estado readonly (não editáveis)"""
    # Preserva o conteúdo atual da descrição
    descricao_atual = entrada_descricao.get("1.0", "end-1c")

    # ✅ CORREÇÃO: Configura todos os campos como 'readonly' (não editáveis)
    for widget in [entrada_numero, entrada_secretaria, entrada_licitacao, entrada_modalidade,
                   entrada_recebimento, entrada_devolucao, entrada_entregue_por, entrada_devolvido_a,
                   entrada_contratado]:
        widget.config(state='readonly')

    # ✅ CORREÇÃO: Campo descrição fica 'disabled' (não editável)
    entrada_descricao.config(state='disabled', bg="#ECEFF1")
    entrada_descricao.delete("1.0", tk.END)
    entrada_descricao.insert("1.0", descricao_atual)

    # Desabilita o botão de cadastrar
    botao_cadastrar.config(text="Atualizar", state='disabled')

    # ✅ ADICIONA: Vincula a tecla Escape para sair do modo visualização
    janela.bind('<Escape>', handle_escape)


def limpar_campos(event=None):
    """Limpa todos os campos do formulário e de busca, reseta para valores padrão."""
    try:
        # Guarda o widget que tem o foco atualmente
        widget_com_foco = janela.focus_get()

        # Limpa campos do formulário de cadastro
        campos_entry = [
            entrada_numero, entrada_secretaria, entrada_licitacao, entrada_modalidade,
            entrada_recebimento, entrada_devolucao, entrada_entregue_por,
            entrada_devolvido_a, entrada_contratado  # Adicionar este campo
        ]

        for widget in campos_entry:
            widget.config(state='normal')
            widget.delete(0, tk.END)

        # Limpa campo de descrição
        entrada_descricao.config(state='normal')
        entrada_descricao.delete("1.0", tk.END)

        # Limpa campos de busca
        entrada_busca.delete(0, tk.END)
        entrada_filtro_secretaria.delete(0, tk.END)
        entrada_filtro_situacao.delete(0, tk.END)
        entrada_filtro_modalidade.delete(0, tk.END)

        # Reseta a situação para o valor padrão
        situacao_var.set("Em Andamento")

        # Desmarca o checkbox Lembretes e restaura a cor
        lembrete_var.set(False)
        entrada_descricao.config(bg="white")
        atualizar_cor_botao_lembrete()

        # Restaura o botão para modo de cadastro
        botao_cadastrar.config(text="Cadastrar", command=cadastrar_processo, state='normal')

        # Remove seleção e foco da tabela para evitar ativação indevida
        sel = tabela.selection()
        if sel:
            tabela.selection_remove(sel)
        tabela.focus("")

        # Desmarca o botão "Selecionar Todos"
        toggle_var.set(False)
        selecionar_todos_var.set(False)
        checkbox_selecionar_todos.deselect()

        # Restaura cores padrão
        entrada_recebimento.config(bg="white")
        entrada_devolucao.config(bg="white")

        # Restaura o foco para o widget original, se ainda existir
        if widget_com_foco and widget_com_foco.winfo_exists():
            widget_com_foco.focus_set()

        # Move a rolagem da tabela para o topo após limpar
        try:
            tabela.yview_moveto(0)
        except Exception:
            pass

        return "break"

    except Exception as e:
        print(f"Erro ao limpar campos: {e}")
        return "break"


def excluir_processo():
    """Exclui os processos selecionados na tabela, movendo-os para a tabela de registros excluídos.

    Verifica se há processos selecionados, confirma a exclusão com o usuário,
    e então move os registros para a tabela de backup antes de excluí-los da tabela principal.
    """
    global registros_apagados

    # Verifica se há itens selecionados
    itens_selecionados = tabela.selection()
    if not itens_selecionados:
        messagebox.showwarning("Aviso", "Selecione um ou mais processos para excluir!")
        return

    # Impede a exclusão de todos os registros
    total_registros = len(tabela.get_children())
    if len(itens_selecionados) >= total_registros:
        messagebox.showwarning("Aviso", "Não é permitido apagar todos os registros!")
        return

    # Confirma a exclusão com o usuário
    quantidade = len(itens_selecionados)
    if quantidade == 1:
        mensagem = "Tem certeza que deseja excluir este processo?"
    else:
        mensagem = f"Tem certeza que deseja excluir {quantidade} processos selecionados?"

    if not messagebox.askyesno("Confirmar Exclusão", mensagem):
        return

    processos_excluidos = []
    erros = []
    try:
        data_exclusao = DateUtils.obter_data_hora_atual_banco()
    except Exception:
        data_exclusao = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    try:
        # Processa cada item selecionado
        for item_selecionado in itens_selecionados:
            try:
                # Obtém os dados do processo selecionado
                processo = tabela.item(item_selecionado)['values']
                if not processo:
                    erros.append(f"Não foi possível identificar um dos registros selecionados.")
                    continue

                numero_processo = str(processo[1]).strip()

                # Salva no backup antes de excluir
                cursor.execute('''
                    INSERT INTO trabalhos_excluidos (
                        data_exclusao, data_registro, numero_processo, secretaria, numero_licitacao,
                        situacao, modalidade, data_inicio, data_entrega, entregue_por, devolvido_a, contratado, descricao
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    data_exclusao, *processo
                ))

                # Exclui do banco de dados principal
                cursor.execute("DELETE FROM trabalhos_realizados WHERE numero_processo = ?", (numero_processo,))

                processos_excluidos.append(numero_processo)

            except Exception as e:
                erros.append(
                    f"Erro ao excluir processo {numero_processo if 'numero_processo' in locals() else 'desconhecido'}: {str(e)}")
                continue

        # Confirma as alterações no banco
        conn.commit()
        # Backup após exclusão
        try:
            backup_automatico(processos_excluidos)
        except Exception:
            pass

        # Remove os itens da interface (apenas os que foram excluídos com sucesso)
        for item_selecionado in itens_selecionados:
            try:
                processo = tabela.item(item_selecionado)['values']
                numero_processo = str(processo[1]).strip()
                if numero_processo in processos_excluidos:
                    tabela.delete(item_selecionado)
            except:
                pass

        # Atualiza estatísticas
        registros_apagados += len(processos_excluidos)
        contar_registros()

        # Limpa backups antigos automaticamente
        limpar_backups_antigos()

        # Exibe resultado
        if processos_excluidos and not erros:
            if len(processos_excluidos) == 1:
                messagebox.showinfo("Sucesso", f"Processo {processos_excluidos[0]} excluído com sucesso!")
            else:
                messagebox.showinfo("Sucesso", f"{len(processos_excluidos)} processos excluídos com sucesso!")
        elif processos_excluidos and erros:
            mensagem = f"{len(processos_excluidos)} processos excluídos com sucesso.\n\nErros encontrados:\n" + "\n".join(
                erros[:5])
            if len(erros) > 5:
                mensagem += f"\n... e mais {len(erros) - 5} erros."
            messagebox.showwarning("Parcialmente Concluído", mensagem)
        else:
            messagebox.showerror("Erro", "Nenhum processo foi excluído.\n\nErros:\n" + "\n".join(erros[:5]))

    except Exception as e:
        messagebox.showerror("Erro", f"Erro geral ao excluir processos: {str(e)}")
        print(f"[ERRO] Exclusão de processos: {e}")
        # Tenta reverter a transação em caso de erro
        try:
            conn.rollback()
        except:
            pass


def limpar_backups_antigos(dias=30):
    """Remove registros antigos da tabela de backups.

    Args:
        dias: tk.Número de dias para manter os backups (padrão: 30)
    """
    try:
        limite = (datetime.now() - timedelta(days=dias)).strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute('DELETE FROM trabalhos_excluidos WHERE data_exclusao < ?', (limite,))
        registros_removidos = cursor.rowcount
        conn.commit()

        if registros_removidos > 0:
            print(f"[INFO] {registros_removidos} backups antigos foram removidos.")

    except Exception as e:
        print(f"[ERRO] Falha ao limpar backups antigos: {e}")
        try:
            conn.rollback()
        except:
            pass


def interpretar_periodo_tempo(termo_busca):
    """Interpreta termos de tempo digitados pelo usuário e retorna intervalo de datas.

    Suporta:
    - Meses: janeiro, fevereiro, etc. (maiúscula ou minúscula)
    - Múltiplos meses: "janeiro fevereiro" ou "janeiro, fevereiro"
    - Semanas: 1semana, 2semanas, etc.
    - Meses: 1mes, 1mês, 2meses, etc.
    - Períodos: mês/mes (vigente), bimestre, trimestre, semestre, ano (vigentes)

    Returns:
        tuple: (data_inicio, data_fim) no formato 'YYYY-MM-DD' ou None se não reconhecido
    """
    import re
    from datetime import datetime, timedelta

    from dateutil.relativedelta import relativedelta

    termo = termo_busca.strip().lower()
    hoje = datetime.now()

    # Dicionário de meses
    meses = {
        'janeiro': 1, 'fevereiro': 2, 'março': 3, 'marco': 3, 'abril': 4,
        'maio': 5, 'junho': 6, 'julho': 7, 'agosto': 8,
        'setembro': 9, 'outubro': 10, 'novembro': 11, 'dezembro': 12
    }

    # Busca por meses específicos por nome (p.ex.: "junho", "julho", etc.)
    meses_encontrados = []
    for nome_mes, numero_mes in meses.items():
        if nome_mes in termo:
            meses_encontrados.append(numero_mes)

    if meses_encontrados:
        # Remove duplicatas e ordena
        meses_encontrados = sorted(list(set(meses_encontrados)))

        # Calcula o intervalo de datas
        ano_atual = hoje.year
        primeiro_mes = min(meses_encontrados)
        ultimo_mes = max(meses_encontrados)

        data_inicio = datetime(ano_atual, primeiro_mes, 1)

        # Último dia do último mês
        if ultimo_mes == 12:
            data_fim = datetime(ano_atual + 1, 1, 1) - timedelta(days=1)
        else:
            data_fim = datetime(ano_atual, ultimo_mes + 1, 1) - timedelta(days=1)

        return (data_inicio.strftime('%Y-%m-%d'), data_fim.strftime('%Y-%m-%d'))

    # Busca por semanas (1semana, 2semanas, etc.)
    padrao_semanas = re.search(r'(\d+)\s*semanas?', termo)
    if padrao_semanas:
        num_semanas = int(padrao_semanas.group(1))
        # Início da semana atual (segunda-feira)
        inicio_semana = hoje - timedelta(days=hoje.weekday())
        data_inicio = inicio_semana - timedelta(weeks=num_semanas - 1)
        data_fim = inicio_semana + timedelta(days=6)  # Domingo da semana atual

        return (data_inicio.strftime('%Y-%m-%d'), data_fim.strftime('%Y-%m-%d'))

    # Busca por meses (1mes, 1mês, 2meses, etc.)
    padrao_meses = re.search(r'(\d+)\s*m[eê]s(?:es)?', termo)
    if padrao_meses:
        num_meses = int(padrao_meses.group(1))
        data_fim = hoje
        data_inicio = hoje - relativedelta(months=num_meses - 1)
        # Primeiro dia do mês de início
        data_inicio = data_inicio.replace(day=1)

        return (data_inicio.strftime('%Y-%m-%d'), data_fim.strftime('%Y-%m-%d'))

    # Períodos específicos
    # Período: mês vigente (garante que 'mes/mês' seja palavra isolada, não parte de 'bimestre', 'trimestre' ou 'semestre')
    if re.search(r"\b(m[eê]s)\b", termo):
        ano_atual = hoje.year
        mes_atual = hoje.month
        data_inicio = datetime(ano_atual, mes_atual, 1)
        if mes_atual == 12:
            data_fim = datetime(ano_atual + 1, 1, 1) - timedelta(days=1)
        else:
            data_fim = datetime(ano_atual, mes_atual + 1, 1) - timedelta(days=1)
        return (data_inicio.strftime('%Y-%m-%d'), data_fim.strftime('%Y-%m-%d'))

    if 'bimestre' in termo:
        # Bimestre vigente (pares: Jan-Fev, Mar-Abr, ...)
        ano_atual = hoje.year
        mes_atual = hoje.month
        inicio_mes = mes_atual - ((mes_atual - 1) % 2)
        data_inicio = datetime(ano_atual, inicio_mes, 1)
        fim_mes = inicio_mes + 1
        if fim_mes == 12:
            data_fim = datetime(ano_atual + 1, 1, 1) - timedelta(days=1)
        else:
            data_fim = datetime(ano_atual, fim_mes + 1, 1) - timedelta(days=1)
        return (data_inicio.strftime('%Y-%m-%d'), data_fim.strftime('%Y-%m-%d'))

    elif 'trimestre' in termo:
        # Trimestre vigente (Q1: Jan-Mar, Q2: Abr-Jun, Q3: Jul-Set, Q4: Out-Dez)
        ano_atual = hoje.year
        mes_atual = hoje.month
        inicio_mes = 3 * ((mes_atual - 1) // 3) + 1
        data_inicio = datetime(ano_atual, inicio_mes, 1)
        fim_mes = inicio_mes + 2
        if fim_mes == 12:
            data_fim = datetime(ano_atual + 1, 1, 1) - timedelta(days=1)
        else:
            data_fim = datetime(ano_atual, fim_mes + 1, 1) - timedelta(days=1)
        return (data_inicio.strftime('%Y-%m-%d'), data_fim.strftime('%Y-%m-%d'))

    elif 'semestre' in termo:
        # Semestre vigente (S1: Jan-Jun, S2: Jul-Dez)
        ano_atual = hoje.year
        mes_atual = hoje.month
        inicio_mes = 6 * ((mes_atual - 1) // 6) + 1
        data_inicio = datetime(ano_atual, inicio_mes, 1)
        fim_mes = inicio_mes + 5
        if fim_mes == 12:
            data_fim = datetime(ano_atual + 1, 1, 1) - timedelta(days=1)
        else:
            data_fim = datetime(ano_atual, fim_mes + 1, 1) - timedelta(days=1)
        return (data_inicio.strftime('%Y-%m-%d'), data_fim.strftime('%Y-%m-%d'))

    elif 'ano' in termo:
        # Ano vigente
        data_inicio = datetime(hoje.year, 1, 1)
        data_fim = datetime(hoje.year, 12, 31)
        return (data_inicio.strftime('%Y-%m-%d'), data_fim.strftime('%Y-%m-%d'))

    # Padrões numéricos:
    # 0006  => mês 06 do ano vigente
    # 000610 => mês 06 do ano 2010 (aceita 2 ou 4 dígitos no ano)
    # 02024 => ano 2024
    padrao_mes_atual = re.match(r"^00(\d{2})$", termo)
    if padrao_mes_atual:
        mm = int(padrao_mes_atual.group(1))
        ano = hoje.year
        if 1 <= mm <= 12:
            data_inicio = datetime(ano, mm, 1)
            if mm == 12:
                data_fim = datetime(ano + 1, 1, 1) - timedelta(days=1)
            else:
                data_fim = datetime(ano, mm + 1, 1) - timedelta(days=1)
            return (data_inicio.strftime('%Y-%m-%d'), data_fim.strftime('%Y-%m-%d'))

    padrao_mes_ano = re.match(r"^00(\d{2})(\d{2}|\d{4})$", termo)
    if padrao_mes_ano:
        mm = int(padrao_mes_ano.group(1))
        ano_str = padrao_mes_ano.group(2)
        ano = int(ano_str)
        if len(ano_str) == 2:
            ano += 2000
        if 1 <= mm <= 12:
            data_inicio = datetime(ano, mm, 1)
            if mm == 12:
                data_fim = datetime(ano + 1, 1, 1) - timedelta(days=1)
            else:
                data_fim = datetime(ano, mm + 1, 1) - timedelta(days=1)
            return (data_inicio.strftime('%Y-%m-%d'), data_fim.strftime('%Y-%m-%d'))

    padrao_ano = re.match(r"^0(\d{4})$", termo)
    if padrao_ano:
        ano = int(padrao_ano.group(1))
        data_inicio = datetime(ano, 1, 1)
        data_fim = datetime(ano, 12, 31)
        return (data_inicio.strftime('%Y-%m-%d'), data_fim.strftime('%Y-%m-%d'))

    return None


def buscar_processos():
    """Busca processos no banco de dados com base nos filtros aplicados.

    Realiza a busca com base nos seguintes critérios:
    - Termo de busca (texto livre, períodos de tempo ou intervalo de datas)
    - Filtro por secretaria
    - Filtro por situação
    - Filtro por modalidade

    Suporta busca por:
    - Períodos de tempo (janeiro, fevereiro, 1semana, 2meses, bimestre, etc.)
    - Intervalo de datas no formato ddmmaDDMM (ex: 0101a3101)
    """
    try:
        # Obtém os valores dos filtros
        termo_busca = entrada_busca.get().strip().lower()
        filtro_secretaria = entrada_filtro_secretaria.get().strip()
        filtro_situacao = entrada_filtro_situacao.get().strip()
        filtro_modalidade = entrada_filtro_modalidade.get().strip()

        # Reseta seleções
        selecionar_todos_var.set(False)
        toggle_var.set(False)
        checkbox_selecionar_todos.deselect()
        checkbox_selecionar_todos.update_idletasks()

        # Limpa a tabela antes de preencher com novos resultados
        for row in tabela.get_children():
            tabela.delete(row)

        # Constrói a consulta SQL base
        query = '''SELECT data_registro, numero_processo, secretaria, numero_licitacao,
                        situacao, modalidade, data_inicio, data_entrega,
                        entregue_por, devolvido_a, contratado, descricao
                FROM trabalhos_realizados WHERE 1=1'''
        params = []

        # Aplica filtro por secretaria
        if filtro_secretaria:
            if ' - ' in filtro_secretaria:
                sigla_secretaria = filtro_secretaria.split(' - ')[0]
            else:
                # Busca a sigla correspondente ao nome da secretaria
                sigla_secretaria = next((sigla for sigla, nome in secretarias_dict.items()
                                         if filtro_secretaria.lower() in nome.lower()), filtro_secretaria)
            query += " AND secretaria = ?"
            params.append(sigla_secretaria)

        # Aplica filtro por situação
        if filtro_situacao:
            query += " AND situacao = ?"
            params.append(filtro_situacao)

        # Aplica filtro por modalidade
        if filtro_modalidade:
            query += " AND modalidade = ?"
            params.append(filtro_modalidade)

        # Verifica se o termo de busca é um período de tempo
        periodo_tempo = interpretar_periodo_tempo(termo_busca)

        if periodo_tempo:
            data_inicio, data_fim = periodo_tempo
            # data_registro está em formato 'DD/MM/YYYY HH:MM'; converte para 'YYYY-MM-DD' para comparação
            query += " AND (substr(data_registro,7,4)||'-'||substr(data_registro,4,2)||'-'||substr(data_registro,1,2)) BETWEEN ? AND ?"
            params.append(data_inicio)
            params.append(data_fim)
            termo_busca = ""  # evita filtro textual desnecessário
        else:
            # Verifica se o termo de busca é um intervalo de datas
            # Suporta:
            # - ddmmaDDMM (aplica ano atual e também ano anterior como fallback)
            # - ddmmYYaDDMMYY / ddmmYYYYaDDMMYYYY (anos explícitos para início e fim)
            import re
            from datetime import datetime

            ano_atual = datetime.now().year

            # Intervalo com anos opcionais (cada data pode ter 2 ou 4 dígitos de ano)
            padrao_intervalo_com_ano = re.match(r"^(\d{2})(\d{2})(\d{2}|\d{4})?a(\d{2})(\d{2})(\d{2}|\d{4})?$",
                                                termo_busca)
            if padrao_intervalo_com_ano:
                d1, m1, y1, d2, m2, y2 = padrao_intervalo_com_ano.groups()
                try:
                    # Determina anos (se não fornecidos, usa ano atual)
                    if y1 is None:
                        ano1 = ano_atual
                    else:
                        ano1 = int(y1)
                        if len(y1) == 2:
                            ano1 += 2000

                    if y2 is None:
                        ano2 = ano_atual
                    else:
                        ano2 = int(y2)
                        if len(y2) == 2:
                            ano2 += 2000

                    # Cria objetos de data
                    data_inicio_obj = datetime(ano1, int(m1), int(d1))
                    data_fim_obj = datetime(ano2, int(m2), int(d2))

                    # Formata para SQL
                    data_inicio = data_inicio_obj.strftime("%Y-%m-%d")
                    data_fim = data_fim_obj.strftime("%Y-%m-%d")

                    # Compara usando data_registro convertido para 'YYYY-MM-DD'
                    query += " AND (substr(data_registro,7,4)||'-'||substr(data_registro,4,2)||'-'||substr(data_registro,1,2)) BETWEEN ? AND ?"
                    params.append(data_inicio)
                    params.append(data_fim)
                    termo_busca = ""  # evita filtro textual desnecessário
                except ValueError as e:
                    print(f"[AVISO] Formato de data inválido: {e}")
                    # Continua com a busca normal
            else:
                # Intervalo sem ano explícito: ddmmaDDMM
                padrao_intervalo = re.match(r"^(\d{2})(\d{2})a(\d{2})(\d{2})$", termo_busca)
                if padrao_intervalo:
                    dia1, mes1, dia2, mes2 = padrao_intervalo.groups()
                    try:
                        # Ano atual
                        data_inicio_atual = datetime(ano_atual, int(mes1), int(dia1))
                        data_fim_atual = datetime(ano_atual, int(mes2), int(dia2))

                        # Ano anterior (fallback)
                        ano_anterior = ano_atual - 1
                        data_inicio_prev = datetime(ano_anterior, int(mes1), int(dia1))
                        data_fim_prev = datetime(ano_anterior, int(mes2), int(dia2))

                        # Formata para SQL
                        data_inicio_atual_sql = data_inicio_atual.strftime("%Y-%m-%d")
                        data_fim_atual_sql = data_fim_atual.strftime("%Y-%m-%d")
                        data_inicio_prev_sql = data_inicio_prev.strftime("%Y-%m-%d")
                        data_fim_prev_sql = data_fim_prev.strftime("%Y-%m-%d")

                        # Faixa mmdd para qualquer ano
                        mmdd_inicio = f"{int(mes1):02d}{int(dia1):02d}"
                        mmdd_fim = f"{int(mes2):02d}{int(dia2):02d}"

                        # Aplica filtro para o ano atual OU ano anterior OU qualquer ano pelo mmdd
                        query += (
                            " AND ((substr(data_registro,7,4)||'-'||substr(data_registro,4,2)||'-'||substr(data_registro,1,2)) BETWEEN ? AND ? "
                            "OR (substr(data_registro,7,4)||'-'||substr(data_registro,4,2)||'-'||substr(data_registro,1,2)) BETWEEN ? AND ? "
                            "OR (substr(data_registro,4,2)||substr(data_registro,1,2)) BETWEEN ? AND ?)"
                        )
                        params.extend([
                            data_inicio_atual_sql, data_fim_atual_sql,
                            data_inicio_prev_sql, data_fim_prev_sql,
                            mmdd_inicio, mmdd_fim
                        ])
                        termo_busca = ""  # evita filtro textual desnecessário
                    except ValueError as e:
                        print(f"[AVISO] Formato de data inválido: {e}")
                        # Continua com a busca normal

        # Aplica filtro por termos textuais se ainda houver termo de busca
        if termo_busca:
            query += ''' AND (numero_processo LIKE ? OR 
                            numero_licitacao LIKE ? OR
                            descricao LIKE ? OR
                            entregue_por LIKE ? OR
                            devolvido_a LIKE ? OR
                            contratado LIKE ?)'''
            params.extend([f"%{termo_busca}%"] * 6)

        # Ordena do mais recente para o mais antigo
        # Ordena por data convertida (YYYY-MM-DD) e depois horário (HH:MM) desc
        query += " ORDER BY (substr(data_registro,7,4)||'-'||substr(data_registro,4,2)||'-'||substr(data_registro,1,2)) DESC, substr(data_registro,12,5) DESC"

        # Executa a consulta
        cursor.execute(query, params)
        resultados = cursor.fetchall()

        # Verifica se há resultados
        if not resultados:
            # Exibe mensagem quando não há resultados
            resp = messagebox.showinfo("Busca", "NENHUMA OCORRÊNCIA ENCONTRADA")
            # Ao confirmar (OK), limpar filtros e reexibir todos os registros
            try:
                # Preserva ordenação atual e apenas recarrega dados completos
                limpar_filtros(preservar_ordem=True)
            except Exception:
                try:
                    # Versões sem parâmetro: apenas limpa filtros e recarrega
                    limpar_filtros()
                except Exception:
                    # Fallback: repopular a tabela completa
                    listar_processos()
        else:
            # Preenche a tabela com os resultados
            for row in resultados:
                valores = list(row)

                # Formata as datas para exibição
                valores[0] = formatar_data_hora_str(valores[0])  # data_registro

                # Formata data_inicio
                if valores[6]:
                    valores[6] = converter_data_para_exibicao(valores[6])

                # Formata data_entrega
                if valores[7]:
                    valores[7] = converter_data_para_exibicao(valores[7])

                # Define tag com base na situação para estilização
                if valores[4] == "Em Andamento":
                    tag = "andamento"
                elif valores[4] == "Concluído":
                    tag = "concluido"
                else:
                    tag = ""

                # Insere na tabela
                tabela.insert("", "end", values=valores, tags=(tag,))

        # Atualiza o contador de registros
        contar_registros()

        # Garante que a rolagem volte ao topo após a busca
        try:
            tabela.yview_moveto(0)
        except Exception:
            pass

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao buscar processos: {str(e)}")
        print(f"[ERRO] Busca de processos: {e}")


def converter_data_para_exibicao(data_str):
    """Converte uma data do formato do banco para o formato de exibição.

    Args:
        data_str: String de data no formato '%Y-%m-%d'

    Returns:
        String de data no formato '%d/%m/%Y'
    """
    try:
        return DateUtils.para_exibicao(str(data_str))
    except Exception:
        return str(data_str)


def limpar_filtros(preservar_ordem: bool = False):
    """Limpa todos os filtros de busca e recarrega a lista completa de processos.

    Se `preservar_ordem` for True, mantém a ordenação atual da tabela
    (não reseta `ordem_colunas_reversa` nem chama `ordenar_coluna`).
    """
    try:
        global selecionar_todos_var, checkbox_selecionar_todos

        # Reseta as variáveis de seleção
        toggle_var.set(False)
        selecionar_todos_var.set(False)
        checkbox_selecionar_todos.deselect()
        checkbox_selecionar_todos.update()

        # Limpa o campo de busca (corrigido a duplicação do delete)
        entrada_busca.delete(0, tk.END)

        # Limpa os campos de autocomplete
        for campo in [entrada_filtro_secretaria, entrada_filtro_situacao, entrada_filtro_modalidade]:
            estado_anterior = campo.cget('state')
            campo.config(state='normal')  # Permite edição temporariamente
            campo.delete(0, tk.END)
            campo.config(state=estado_anterior)  # Restaura o estado original

        # Recarrega a lista completa
        listar_processos()

        # Reseta a ordenação para padrão (situação), a menos que seja para preservar
        if not preservar_ordem:
            for col in ordem_colunas_reversa:
                ordem_colunas_reversa[col] = False
            ordenar_coluna('situacao')

        # Garante que a rolagem volte ao topo após limpar filtros
        try:
            tabela.yview_moveto(0)
        except Exception:
            pass

        # Foca no campo de busca para facilitar nova pesquisa
        entrada_busca.focus_set()

    except Exception as e:
        print(f"[ERRO] Falha ao limpar filtros: {e}")
        messagebox.showerror("Erro", f"Erro ao limpar filtros: {str(e)}")


def mostrar_opcoes_situacao(event):
    entrada_filtro_situacao.show_suggestions(["Em Andamento", "Concluído"])


def exportar_pdf():
    """Exporta os processos selecionados para um arquivo PDF formatado.

    Cria um relatório PDF com os dados dos processos selecionados na tabela,
    incluindo uma seção de observações completas para processos em andamento.
    """
    try:
        # Importa as bibliotecas necessárias
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                        Table, TableStyle)

        # Verifica se há itens selecionados, se não houver, seleciona todos
        itens_selecionados = tabela.selection()
        if not itens_selecionados:
            # Se nenhum item estiver selecionado, exporta todos os registros
            itens_selecionados = tabela.get_children()

        # Cabeçalhos em maiúsculas e negrito
        cabecalhos_pdf = [
            "#",
            "RECEBIMENTO",
            "DEVOLUÇÃO",
            "CONTRATO",
            "LICITAÇÃO",
            "MODALIDADE",
            "SITUAÇÃO",
            "OBSERVAÇÕES"
        ]

        dados = [cabecalhos_pdf]
        descricoes_longas = []
        estilo = getSampleStyleSheet()

        # Paleta de cores claras e distintas
        CORES_PENDENCIAS = [
            colors.HexColor("#FFF2CC"),  # Amarelo claro
            colors.HexColor("#D5E8D4"),  # Verde claro
            colors.HexColor("#DAE8FC"),  # Azul claro
            colors.HexColor("#E1D5E7"),  # Lilás claro
            colors.HexColor("#F5F5F5"),  # Cinza muito claro
            colors.HexColor("#FFE6CC"),  # Laranja claro
            colors.HexColor("#E2F0D9"),  # Verde menta
            colors.HexColor("#FDEADA"),  # Bege
            colors.HexColor("#E8EAF6"),  # Azul lavanda
            colors.HexColor("#FCE4EC")  # Rosa claro
        ]

        # Define estilos de texto
        estilo_centralizado = ParagraphStyle(
            'Centralizado',
            parent=estilo['Normal'],
            alignment=1,
            leading=10,
            fontSize=8,
            textColor=colors.black,
            fontName='Helvetica'
        )

        estilo_normal = ParagraphStyle(
            'Normal',
            parent=estilo['Normal'],
            fontSize=8,
            textColor=colors.black,
            fontName='Helvetica'
        )

        # Processa cada item selecionado
        for idx, item in enumerate(itens_selecionados, start=1):
            valores = tabela.item(item)["values"]
            num_registro = str(idx)

            # Converter valores para string para evitar erros
            valores = [str(v) if v is not None else "" for v in valores]

            # Processa a descrição
            descricao = valores[10] if valores[10] else "Sem pendências"
            if valores[4] == "Em Andamento" and descricao:
                # Pega apenas a primeira linha para a tabela principal
                descricao = str(descricao).split('\n')[0]

            # Verifica se tem pendência
            tem_pendencia = descricao.strip().lower() not in ["sem alterações", "sem alteracoes", "sem pendências"]

            # Define o texto a ser exibido na tabela principal
            if tem_pendencia:
                descricao_texto = f"Ver registro {idx}"
                descricoes_longas.append((idx, descricao))
            else:
                descricao_texto = descricao

            # Cria parágrafos formatados
            descricao_paragraph = Paragraph(f"<para align='center'>{descricao_texto}</para>", estilo_centralizado)
            modalidade_paragraph = Paragraph(f"<para align='center'>{valores[5]}</para>", estilo_centralizado)
            situacao_paragraph = Paragraph(f"<para align='center'>{valores[4]}</para>", estilo_centralizado)

            # Monta a linha de dados
            linha = [
                num_registro,
                valores[6], valores[7], valores[1], valores[3],
                modalidade_paragraph,
                situacao_paragraph,
                descricao_paragraph
            ]
            dados.append(linha)

        # Define o nome do arquivo no padrão solicitado: Relatório_PDF_DDMMYYYY_HHMMSS
        data_hora_atual = datetime.now()
        nome_arquivo = f"Relatório_PDF_{data_hora_atual.strftime('%d%m%Y_%H%M%S')}.pdf"

        # Adiciona a secretaria ao nome do arquivo se disponível
        secretaria_filtro = entrada_filtro_secretaria.get().strip()
        if secretaria_filtro:
            nome_arquivo = f"{secretaria_filtro}_{nome_arquivo}"

        arquivo = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            title="Salvar relatório como",
            initialfile=nome_arquivo
        )
        if not arquivo:
            return

        # Cria o documento PDF
        doc = SimpleDocTemplate(arquivo, pagesize=A4, rightMargin=18, leftMargin=18, topMargin=18, bottomMargin=18)
        story = []

        # Configuração das colunas
        largura_total = 540
        larguras = [25, 65, 60, 80, 80, 90, 60, 85]
        soma = sum(larguras)
        larguras = [largura_total * (w / soma) for w in larguras]

        # Cria a tabela principal
        tabela_pdf = Table(dados, repeatRows=1, colWidths=larguras)

        # Define o estilo da tabela
        estilo_tabela = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#E0E0E0")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), "CENTER"),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 0.7, colors.HexColor("#777777")),  # Linhas mais escuras e grossas
        ]

        # Aplica cores claras para linhas com pendências
        cor_idx = 0
        for i in range(1, len(dados)):
            if "Ver registro" in dados[i][7].text:
                estilo_tabela.append(('BACKGROUND', (0, i), (-1, i), CORES_PENDENCIAS[cor_idx % len(CORES_PENDENCIAS)]))
                cor_idx += 1

        tabela_pdf.setStyle(TableStyle(estilo_tabela))

        # Adiciona o título do relatório
        titulo_style = ParagraphStyle(
            'Titulo',
            parent=estilo['Title'],
            fontName='Helvetica-Bold',
            textColor=colors.black,
            fontSize=12,
            spaceAfter=12
        )
        story.append(Paragraph("RELATÓRIO DE PROCESSOS", titulo_style))
        try:
            gerado_em = DateUtils.obter_data_hora_atual()
        except Exception:
            gerado_em = datetime.now().strftime('%d/%m/%Y %H:%M')
        story.append(Paragraph(f"Gerado em: {gerado_em}", estilo_normal))
        story.append(Spacer(1, 10))
        story.append(tabela_pdf)

        # Adiciona a seção de observações completas
        if descricoes_longas:
            story.append(Spacer(1, 20))
            subtitulo_style = ParagraphStyle(
                'Subtitulo',
                parent=estilo['Heading2'],
                fontName='Helvetica-Bold',
                textColor=colors.black,
                fontSize=10,
                spaceAfter=6
            )
            story.append(Paragraph("OBSERVAÇÕES COMPLETAS", subtitulo_style))

            # Adiciona cada observação com formatação
            cor_idx = 0
            for idx, texto in descricoes_longas:
                story.append(Spacer(1, 6))
                tabela_obs = Table([
                    [Paragraph(f"Registro {idx}", estilo_centralizado),
                     Paragraph(texto, estilo_normal)]
                ], colWidths=[60, 460])

                cor = CORES_PENDENCIAS[cor_idx % len(CORES_PENDENCIAS)]
                tabela_obs.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), cor),
                    ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor("#555555")),
                    ('VALIGN', (0, 0), (-1, -1), "TOP"),
                    ('ALIGN', (0, 0), (0, -1), "CENTER"),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ]))
                story.append(tabela_obs)
                cor_idx += 1

        # Gera o PDF
        doc.build(story)

        # Pergunta se o usuário deseja abrir o arquivo
        abrir = messagebox.askyesno("Exportação Concluída",
                                    f"Relatório PDF exportado com sucesso!\n\n{arquivo}\n\nDeseja abrir o arquivo agora?")
        if abrir:
            try:
                os.startfile(arquivo)  # Windows
            except AttributeError:
                try:
                    subprocess.run(["open", arquivo])  # macOS
                except FileNotFoundError:
                    subprocess.run(["xdg-open", arquivo])  # Linux
            except Exception as e:
                messagebox.showerror("Erro", f"Não foi possível abrir o arquivo:\n{str(e)}")

    except ImportError as e:
        messagebox.showerror("Erro",
                             f"Biblioteca necessária não encontrada: {e}\n\nInstale as bibliotecas necessárias com: pip install reportlab")
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao exportar PDF:\n{e}")
        print(f"[ERRO] Exportação PDF: {e}")


def exportar_excel():
    """Exporta os processos selecionados para um arquivo Excel formatado.

    Cria uma planilha Excel com os dados dos processos selecionados na tabela,
    incluindo uma seção de descrições para processos em andamento.
    Determina automaticamente o tipo de relatório com base nas datas dos processos.
    """
    try:
        # Importa as bibliotecas necessárias
        from math import ceil

        from openpyxl.utils import get_column_letter

        # Verifica se há itens selecionados, se não houver, seleciona todos
        itens_selecionados = tabela.selection()
        if not itens_selecionados:
            # Se nenhum item estiver selecionado, exporta todos os registros
            itens_selecionados = tabela.get_children()

        # Coleta os dados dos itens selecionados
        dados = []
        datas_recebimento = []

        for idx, item in enumerate(itens_selecionados, start=1):
            valores = tabela.item(item)["values"]
            data_recebimento = valores[6] or ""

            # Tenta converter a data de recebimento para objeto datetime
            if data_recebimento:
                try:
                    try:
                        data_obj = datetime.strptime(data_recebimento, "%d/%m/%Y")
                    except ValueError:
                        data_obj = datetime.strptime(data_recebimento, "%Y-%m-%d")
                    datas_recebimento.append(data_obj)
                except Exception as e:
                    print(f"[AVISO] Erro ao converter data '{data_recebimento}': {e}")

            # Adiciona os dados do item à lista
            devolucao_valor = valores[7] or ""
            if isinstance(devolucao_valor, str) and devolucao_valor.strip().lower() == "none":
                devolucao_valor = ""
            dados.append({
                "n": idx,
                "recebimento": valores[6] or "",
                "devolucao": devolucao_valor,
                "contrato": str(valores[1]) if valores[1] else "",
                "licitacao": str(valores[3]) if valores[3] else "",
                "modalidade": valores[5] or "",
                "situacao": valores[4] or "",
                "descricao": valores[10] or "",
                "id_original": item
            })

        # Determina o tipo de relatório com base nas datas
        tipo_relatorio = "Relatório"
        if datas_recebimento:
            data_inicio = min(datas_recebimento)
            data_fim = max(datas_recebimento)
            diferenca = (data_fim - data_inicio).days

            if diferenca == 0:
                tipo_relatorio = "Relatório Diário"
            elif diferenca <= 7:
                tipo_relatorio = "Relatório Semanal"
            elif diferenca <= 31:
                tipo_relatorio = "Relatório Mensal"
            elif diferenca <= 365:
                tipo_relatorio = "Relatório Anual"

            periodo = f"de {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
            data_hora_atual = datetime.now()
            # Padrão: Relatório_EXCEL_DDMMYYYY_HHMMSS
            nome_arquivo = f"Relatório_EXCEL_{data_hora_atual.strftime('%d%m%Y_%H%M%S')}.xlsx"
        else:
            periodo = "sem data definida"
            data_hora_atual = datetime.now()
            nome_arquivo = f"Relatório_EXCEL_{data_hora_atual.strftime('%d%m%Y_%H%M%S')}.xlsx"

        # Adiciona a secretaria ao nome do arquivo se disponível
        secretaria_filtro = entrada_filtro_secretaria.get().strip()
        if secretaria_filtro:
            nome_arquivo = f"{secretaria_filtro}_{nome_arquivo}"

        # Solicita o local para salvar o arquivo
        arquivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")],
            title="Salvar relatório como",
            initialfile=nome_arquivo
        )

        if not arquivo:
            return

        # Cria a planilha Excel
        wb = Workbook()
        ws = wb.active
        ws.title = tipo_relatorio

        # Define os estilos
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        font_cabecalho = Font(bold=True)
        fonte_padrao = Font(name="Calibri", size=9)
        alinhamento_central = Alignment(horizontal='center', vertical='center', wrap_text=True)
        alinhamento_esquerda = Alignment(horizontal='left', vertical='top', wrap_text=True)

        fundo_cinza = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        fundo_azul_claro = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        fundo_amarelo = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        fundo_azul_vazio = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

        # Adiciona o título do relatório
        ws.merge_cells('A1:G1')
        titulo_celula = ws.cell(row=1, column=1, value=f"{tipo_relatorio.upper()} {periodo}")
        titulo_celula.font = Font(bold=True, size=14)
        titulo_celula.alignment = Alignment(horizontal='center', vertical='center')
        titulo_celula.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
        ws.row_dimensions[1].height = 30

        # Adiciona a data de geração
        ws.merge_cells('A2:G2')
        try:
            gerado_em = DateUtils.obter_data_hora_atual()
        except Exception:
            gerado_em = datetime.now().strftime('%d/%m/%Y %H:%M')
        data_geracao = ws.cell(row=2, column=1, value=f"Gerado em: {gerado_em}")
        data_geracao.font = fonte_padrao
        data_geracao.alignment = Alignment(horizontal='right', vertical='center')
        ws.row_dimensions[2].height = 20

        # Adiciona os cabeçalhos da tabela principal
        cabecalhos = ["#", "Recebimento", "Devolução", "Contrato", "Licitação", "Modalidade", "Situação"]
        linha_cabecalho = 3
        for col_num, cab in enumerate(cabecalhos, 1):
            cell = ws.cell(row=linha_cabecalho, column=col_num, value=cab)
            cell.font = font_cabecalho
            cell.alignment = alinhamento_central
            cell.fill = fundo_cinza
            cell.border = thin_border

        # Função para limpar e preparar o texto
        def preparar_texto(texto):
            """Limpa e formata o texto para exibição na planilha."""
            if not texto or not isinstance(texto, str):
                return ""
            texto = texto.strip()
            texto = re.sub(r'\s+', ' ', texto)
            return texto.replace('\r', '').replace('\n', ' ')

        # Função para calcular altura da linha
        def calcular_altura_linha(texto, largura_coluna):
            """Calcula a altura ideal da linha com base no conteúdo e largura."""
            if not texto or not isinstance(texto, str):
                return 15  # Altura mínima

            # Remove espaços extras e quebras de linha
            texto = ' '.join(str(texto).split())

            # Configurações precisas
            altura_base = 15  # Altura mínima
            pixels_por_linha = 12  # Altura por linha adicional
            margem = 2  # Margem mínima
            max_altura = 80  # Altura máxima permitida

            # Calcula caracteres por linha baseado na largura
            caracteres_por_linha = max(1, (largura_coluna * 1.8))  # Fator de ajuste empírico

            # Calcula linhas necessárias
            num_linhas = max(1, ceil(len(texto) / caracteres_por_linha))

            # Calcula altura final
            altura = altura_base + ((num_linhas - 1) * pixels_por_linha) + margem
            return min(altura, max_altura)

        # Preenche a tabela principal
        for row_idx, item in enumerate(dados, start=linha_cabecalho + 1):
            for col_idx, col in enumerate(
                    ['n', 'recebimento', 'devolucao', 'contrato', 'licitacao', 'modalidade', 'situacao'], 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                valor = preparar_texto(str(item[col])) if item[col] else "não preenchido"
                cell.value = valor
                cell.font = fonte_padrao
                cell.alignment = alinhamento_central
                cell.border = thin_border

                # Aplica formatação condicional
                if not item[col]:
                    cell.fill = fundo_azul_vazio
                elif item["situacao"].strip().lower() == "em andamento":
                    cell.fill = fundo_amarelo

            ws.row_dimensions[row_idx].height = 25

        # Ajusta as larguras das colunas
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 13
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['G'].width = 11

        # Função para calcular larguras máximas de forma segura
        def safe_max_width(header, values):
            """Calcula a largura máxima de uma coluna de forma segura."""
            try:
                return max(len(str(header)), *[len(str(v)) for v in values if v is not None]) + 2
            except Exception:
                return len(str(header)) + 2

        # Ajusta as larguras das colunas com base no conteúdo
        if dados:
            ws.column_dimensions['D'].width = min(safe_max_width("Contrato", [d['contrato'] for d in dados]), 30)
            ws.column_dimensions["E"].width = min(safe_max_width("Licitação", [d['licitacao'] for d in dados]), 30)
            ws.column_dimensions['F'].width = min(safe_max_width("Modalidade", [d['modalidade'] for d in dados]), 30)

        # Adiciona uma linha de separação
        linha_separacao = len(dados) + linha_cabecalho + 1
        ws.merge_cells(start_row=linha_separacao, start_column=2, end_row=linha_separacao, end_column=7)

        for col in range(2, 8):
            cell = ws.cell(row=linha_separacao, column=col)
            cell.fill = fundo_azul_claro
            cell.border = thin_border

        ws.row_dimensions[linha_separacao].height = 25

        # Adiciona os cabeçalhos da seção de descrição
        linha_cabecalho_desc = linha_separacao + 1

        # Coluna #
        cell_num = ws.cell(row=linha_cabecalho_desc, column=1, value="#")
        cell_num.font = font_cabecalho
        cell_num.alignment = alinhamento_central
        cell_num.fill = fundo_cinza
        cell_num.border = thin_border

        # Coluna Descrição
        ws.merge_cells(start_row=linha_cabecalho_desc, start_column=2, end_row=linha_cabecalho_desc, end_column=7)
        cell_desc = ws.cell(row=linha_cabecalho_desc, column=2, value="Descrição")
        cell_desc.font = font_cabecalho
        cell_desc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell_desc.fill = fundo_cinza
        for col in range(1, 8):
            ws.cell(row=linha_cabecalho_desc, column=col).border = thin_border
        ws.row_dimensions[linha_cabecalho_desc].height = 25

        # Preenche a seção de descrição (apenas para processos em andamento)
        linha_atual = linha_cabecalho_desc + 1
        processos_em_andamento = 0

        for item in dados:
            if item["situacao"].strip().lower() == "em andamento":
                processos_em_andamento += 1
                descricao = preparar_texto(item['descricao'])

                # Coluna #
                cell_num = ws.cell(row=linha_atual, column=1, value=item['n'])
                cell_num.font = fonte_padrao
                cell_num.alignment = alinhamento_central
                cell_num.border = thin_border
                cell_num.fill = fundo_amarelo

                # Coluna Descrição
                ws.merge_cells(start_row=linha_atual, start_column=2, end_row=linha_atual, end_column=7)
                cell_desc = ws.cell(row=linha_atual, column=2)
                cell_desc.value = descricao if descricao else "não preenchido"
                cell_desc.font = fonte_padrao
                cell_desc.alignment = Alignment(
                    horizontal='left',
                    vertical='top',
                    wrap_text=True,
                    shrink_to_fit=False
                )

                # Aplica bordas e fundo
                for col in range(1, 8):
                    ws.cell(row=linha_atual, column=col).border = thin_border
                    ws.cell(row=linha_atual, column=col).fill = fundo_amarelo if descricao else fundo_azul_vazio

                # Calcula altura baseada no conteúdo
                total_width = sum(ws.column_dimensions[get_column_letter(c)].width for c in range(2, 8))
                ws.row_dimensions[linha_atual].height = calcular_altura_linha(descricao, total_width)

                linha_atual += 1

        # Se não houver processos em andamento, adiciona uma mensagem
        if processos_em_andamento == 0:
            ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=7)
            cell = ws.cell(row=linha_atual, column=1, value="Não há processos em andamento")
            cell.font = fonte_padrao
            cell.alignment = alinhamento_central
            cell.border = thin_border
            cell.fill = fundo_azul_claro
            ws.row_dimensions[linha_atual].height = 25

        # Configura as opções de impressão
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = True
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        # Salva o arquivo
        wb.save(arquivo)

        # Pergunta se o usuário deseja abrir o arquivo
        abrir = messagebox.askyesno("Exportação Concluída",
                                    f"{tipo_relatorio} exportado com sucesso!\n\n{arquivo}\n\nDeseja abrir o arquivo agora?")
        if abrir:
            try:
                os.startfile(arquivo)  # Windows
            except AttributeError:
                try:
                    subprocess.run(["open", arquivo])  # macOS
                except FileNotFoundError:
                    subprocess.run(["xdg-open", arquivo])  # Linux
            except Exception as e:
                messagebox.showerror("Erro", f"Não foi possível abrir o arquivo:\n{str(e)}")

    except ImportError as e:
        messagebox.showerror("Erro",
                             f"Biblioteca necessária não encontrada: {e}\n\nInstale as bibliotecas necessárias com: pip install openpyxl")
    except Exception as e:
        messagebox.showerror("Erro",
                             f"Erro ao exportar {tipo_relatorio.lower() if 'tipo_relatorio' in locals() else 'relatório'}:\n{str(e)}")
        print(f"[ERRO] Exportação Excel: {e}")


def exportar_txt():
    """Exporta os processos selecionados para um arquivo de texto formatado.

    Cria um arquivo de texto com os dados dos processos selecionados na tabela,
    organizados de forma legível e estruturada.
    """
    try:
        # Verifica se há itens selecionados, se não houver, seleciona todos
        itens_selecionados = tabela.selection()
        if not itens_selecionados:
            # Se nenhum item estiver selecionado, exporta todos os registros
            itens_selecionados = tabela.get_children()

        # Solicita o local para salvar o arquivo
        arquivo_destino = filedialog.asksaveasfilename(
            title="Salvar arquivo de texto",
            defaultextension=".txt",
            filetypes=[("Arquivos de Texto", "*.txt")],
            initialfile=f"Relatório_TXT_{datetime.now().strftime('%d%m%Y_%H%M%S')}.txt"
        )

        if not arquivo_destino:
            return

        # Coleta os dados dos itens selecionados
        dados = []
        for item in itens_selecionados:
            valores = tabela.item(item)["values"]
            dados.append({
                'data_registro': valores[0] or "",
                'numero_processo': valores[1] or "",
                'secretaria': valores[2] or "",
                'numero_licitacao': valores[3] or "",
                'situacao': valores[4] or "",
                'modalidade': valores[5] or "",
                'descricao': valores[10] or ""
            })

        # Cria o conteúdo do arquivo de texto
        conteudo = []
        conteudo.append("RELATÓRIO DE PROCESSOS")
        conteudo.append("=" * 50)
        conteudo.append(f"Data de geração: {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}")
        conteudo.append(f"Total de processos: {len(dados)}")
        conteudo.append("=" * 50)
        conteudo.append("")

        # Adiciona os dados de cada processo
        for i, processo in enumerate(dados, 1):
            conteudo.append(f"PROCESSO {i:03d}")
            conteudo.append("-" * 20)
            conteudo.append(f"Data de Registro: {processo['data_registro']}")
            conteudo.append(f"Número do Processo: {processo['numero_processo']}")

            # Converte a sigla da secretaria para o nome completo
            secretaria_sigla = processo['secretaria']
            secretaria_completa = secretarias_dict.get(secretaria_sigla, secretaria_sigla)
            conteudo.append(f"Secretaria: {secretaria_completa}")

            conteudo.append(f"Número da Licitação: {processo['numero_licitacao']}")
            conteudo.append(f"Situação: {processo['situacao']}")
            conteudo.append(f"Modalidade: {processo['modalidade']}")
            if processo['descricao']:
                conteudo.append(f"Descrição: {processo['descricao']}")
            conteudo.append("")

        # Salva o arquivo
        with open(arquivo_destino, 'w', encoding='utf-8') as arquivo:
            arquivo.write('\n'.join(conteudo))

        # Pergunta se o usuário deseja abrir o arquivo
        abrir = messagebox.askyesno("Exportação Concluída",
                                    f"Arquivo TXT exportado com sucesso!\n\n{arquivo_destino}\n\nDeseja abrir o arquivo agora?")
        if abrir:
            try:
                os.startfile(arquivo_destino)  # Windows
            except AttributeError:
                try:
                    subprocess.run(["open", arquivo_destino])  # macOS
                except FileNotFoundError:
                    subprocess.run(["xdg-open", arquivo_destino])  # Linux
            except Exception as e:
                messagebox.showerror("Erro", f"Não foi possível abrir o arquivo:\n{str(e)}")

    except Exception as e:
        print(f"[ERRO] Exportação TXT: {e}")
        messagebox.showerror("Erro", f"Erro ao exportar para TXT: {e}")


def ajustar_todas_colunas(usar_larguras_personalizadas=True):
    """Ajusta as larguras das colunas da tabela com base no estado da janela e no conteúdo.

    Adapta a visualização da tabela dependendo se a janela está maximizada ou não,
    aplicando larguras personalizadas ou calculando-as com base no conteúdo.

    Args:
        usar_larguras_personalizadas (bool): Se True, usa as larguras predefinidas em larguras_fixas.
                                           Se False, calcula as larguras com base no conteúdo.
    """
    try:
        # Determina se a janela está maximizada
        estado = janela.state()
        largura_janela = janela.winfo_width()
        maximizada = estado == 'zoomed' or largura_janela > 1300

        # Calcula a largura disponível para a tabela
        largura_disponivel = tabela.winfo_width() - 30
        if largura_disponivel <= 0:
            # Se a tabela ainda não foi renderizada, use a largura da janela como aproximação
            largura_disponivel = largura_janela - 50
            if largura_disponivel <= 0:
                largura_disponivel = 800  # Valor padrão seguro

        # Obtém a fonte atual para cálculos de largura
        font = tkinter.font.Font(root=janela)

        if usar_larguras_personalizadas:
            # Modo 1: Usar larguras predefinidas
            if maximizada:
                # Quando maximizada, mostra todas as colunas exceto descrição
                colunas_visiveis = [col for col in cabecalhos if col != 'descricao']

                # Define larguras específicas para cada coluna quando maximizada
                for col in colunas_visiveis:
                    # Define larguras específicas para quando a janela está maximizada
                    if col == 'data_registro':
                        largura = 140  # Aumentado de 120 para 140
                    elif col == 'numero_processo':
                        largura = 140
                    elif col == 'secretaria':
                        largura = 180
                    elif col == 'numero_licitacao':
                        largura = 130
                    elif col == 'modalidade':
                        largura = 120
                    elif col == 'situacao':
                        largura = 100
                    elif col == 'data_inicio':
                        largura = 110
                    elif col == 'data_entrega':
                        largura = 110
                    elif col == 'entregue_por':
                        largura = 130
                    elif col == 'recebimento':
                        largura = 140  # Aumentado para 140
                    elif col == 'devolvido_a':
                        largura = 150  # Aumentado de 130 para 150
                    elif col == 'contratado':
                        largura = 130
                    else:
                        largura = larguras_fixas.get(col, 100)

                    # Aplica larguras mínimas para colunas específicas
                    if col == 'secretaria':
                        largura = max(largura, 160)

                    if col in ['data_inicio', 'data_entrega']:
                        largura = max(largura, 120)

                    # Limita a largura entre 50 e 300 pixels
                    largura = max(50, min(largura, 300))
                    tabela.column(col, width=largura, stretch=True)

                # Configura a coluna de descrição (reduzida de 300 para 200)
                tabela.column('descricao', width=200, stretch=True)

            else:
                # Quando não maximizada, mostra apenas colunas essenciais
                for col in cabecalhos:
                    if col in colunas_visiveis_padrao:
                        tabela.column(col, width=larguras_fixas.get(col, 100), stretch=False)
                    else:
                        tabela.column(col, width=0, stretch=False)

                # Oculta a coluna de descrição
                tabela.column('descricao', width=0, stretch=False)

        else:
            # Modo 2: Calcular larguras com base no conteúdo
            for col in cabecalhos:
                # Oculta colunas não essenciais quando não maximizada
                if not maximizada and col not in colunas_visiveis_padrao:
                    tabela.column(col, width=0)
                    continue

                # Oculta a descrição quando não maximizada
                if col == 'descricao' and not maximizada:
                    continue

                # Calcula a largura máxima necessária para o conteúdo
                max_width = font.measure(cabecalhos[col])

                # Verifica o conteúdo de cada célula na coluna
                for item in tabela.get_children():
                    valor = str(tabela.set(item, col))
                    item_width = font.measure(valor)
                    max_width = max(max_width, item_width)

                # Adiciona padding
                largura_final = max_width + 24

                # Ajusta proporcionalmente quando maximizada
                if maximizada and col != 'descricao':
                    colunas_visiveis = [c for c in cabecalhos if c != 'descricao']
                    soma_larguras = sum(font.measure(cabecalhos[c]) + 24 for c in colunas_visiveis)
                    if soma_larguras > 0:
                        proporcao = (max_width + 24) / soma_larguras
                        largura_final = int(largura_disponivel * proporcao)

                # Garante largura mínima para secretaria
                if col == 'secretaria':
                    largura_final = max(largura_final, 160)

                # Limita a largura entre 50 e 300 pixels
                largura_final = max(50, min(largura_final, 300))
                tabela.column(col, width=largura_final)

        # Atualiza a interface
        tabela.update_idletasks()

    except Exception as e:
        print(f"[ERRO] Falha ao ajustar colunas: {e}")
        # Garante que a tabela permaneça utilizável mesmo em caso de erro
        for col in cabecalhos:
            tabela.column(col, width=100, stretch=True)


def ordenar_coluna(coluna):
    try:
        # Alterna a ordem da coluna clicada
        ordem_colunas_reversa[coluna] = not ordem_colunas_reversa[coluna]

        # Limpa setas de todas as colunas antes de atualizar
        for col in tabela['columns']:
            tabela.heading(col, text=cabecalhos[col])

        # Atualiza o cabeçalho da coluna ordenada com a seta
        seta = '▼' if ordem_colunas_reversa[coluna] else '▲'
        tabela.heading(coluna, text=f"{cabecalhos[coluna]} {seta}")

        # Coleta os itens para ordenação
        items = []
        for item in tabela.get_children(''):
            valor = tabela.set(item, coluna)

            # Converte datas para formato ordenável
            if coluna in ['data_inicio', 'data_entrega', 'data_registro']:
                if valor and valor != "não preenchido":
                    try:
                        data_obj = datetime.strptime(valor, "%d/%m/%Y")
                        valor = data_obj.strftime("%Y%m%d")
                    except ValueError:
                        pass

            items.append((valor, item))

        # Ordena os itens
        items.sort(reverse=ordem_colunas_reversa[coluna])

        # Reposiciona os itens na tabela
        for index, (val, item) in enumerate(items):
            tabela.move(item, '', index)

    except Exception as e:
        print(f"[ERRO] Falha ao ordenar coluna '{coluna}': {e}")
        messagebox.showerror("Erro", f"Não foi possível ordenar a coluna {cabecalhos.get(coluna, coluna)}")


# 2. Vincule corretamente o evento ao cabeçalho


def center_window(window, width=860, height=650):
    """Centraliza uma janela na tela com dimensões específicas.

    Calcula a posição para centralizar a janela na tela, garantindo que
    as dimensões não excedam o tamanho da tela disponível, e aplica
    a geometria calculada à janela.

    Args:
        window (Tk/Toplevel): A janela a ser centralizada.
        width (int): Largura desejada para a janela. Padrão é 860.
        height (int): Altura desejada para a janela. Padrão é 650.
    """
    try:
        # Obtém as dimensões da tela
        screen_width = window.winfo_screenwidth()
        screen_height = window.winfo_screenheight()

        # Garante que as dimensões não excedam o tamanho da tela
        width = min(width, screen_width - 20)  # Margem de 20 pixels
        height = min(height, screen_height - 40)  # Margem de 40 pixels

        # Calcula as coordenadas para centralizar
        x = (screen_width - width) // 2
        y = (screen_height - height) // 6  # Posiciona um pouco acima do centro

        # Garante que as coordenadas não sejam negativas
        x = max(0, x)
        y = max(0, y)

        # Aplica a geometria calculada à janela
        window.geometry(f"{width}x{height}+{x}+{y}")
    except Exception as e:
        print(f"[ERRO] Falha ao centralizar janela: {e}")
        # Em caso de erro, tenta definir uma geometria padrão segura
        try:
            window.geometry("800x600+10+10")
        except:
            pass
    window.maxsize(screen_width, screen_height)


def ativar_botao_atualizar():
    """Ativa o botão Atualizar somente quando há seleção válida na tabela"""
    try:
        itens_selecionados = tabela.selection()
        if itens_selecionados:
            processo = tabela.item(itens_selecionados[0])['values']
            if processo and len(processo) > 1:
                numero_original = str(processo[1]).strip().upper()
                if numero_original:
                    botao_cadastrar.config(
                        text="Atualizar",
                        state='normal',
                        command=lambda: atualizar_processo(numero_original)
                    )
    except:
        pass


# --- INTERFACE GRÁFICA ---

janela = tk.Tk()
janela.title("MeuGestor")
try:
    w, h, xi, yi = carregar_tamanho_janela('principal', 860, 650)
    if xi is not None and yi is not None:
        janela.geometry(f"{w}x{h}+{xi}+{yi}")
    else:
        center_window(janela, w, h)
except Exception:
    center_window(janela)
try:
    janela.resizable(True, True)
except Exception:
    pass
try:
    def _close_root():
        try:
            janela.update_idletasks()
            salvar_tamanho_janela('principal', janela.winfo_width(), janela.winfo_height(), janela.winfo_x(),
                                  janela.winfo_y())
        except Exception:
            pass
        janela.destroy()


    janela.protocol("WM_DELETE_WINDOW", _close_root)
except Exception:
    pass
try:
    bind_persist_geometry('principal', janela)
except Exception:
    pass
try:
    bind_entry_word_delete(janela)
except Exception:
    pass
try:
    bind_text_word_delete(janela)
except Exception:
    pass
try:
    setup_global_text_shortcuts(janela)
    setup_global_entry_shortcuts(janela)
    setup_context_menu(janela)
except Exception:
    pass


# Ajustar janela principal para tocar topo e base da tela
def ajustar_altura_principal():
    try:
        try:
            if existe_config_janela('principal'):
                return
        except Exception:
            pass
        janela.update_idletasks()

        # Dimensões totais da tela
        screen_w = janela.winfo_screenwidth()
        screen_h = janela.winfo_screenheight()

        # Obtém a área de trabalho (exclui barra de tarefas) via API do Windows
        SPI_GETWORKAREA = 0x0030
        rect = wintypes.RECT()
        work_ok = False
        try:
            res = ctypes.windll.user32.SystemParametersInfoW(SPI_GETWORKAREA, 0, ctypes.byref(rect), 0)
            work_ok = bool(res)
        except Exception:
            work_ok = False

        if work_ok:
            work_left, work_top, work_right, work_bottom = rect.left, rect.top, rect.right, rect.bottom
            work_width = max(work_right - work_left, 1)
            work_height = max(work_bottom - work_top, 1)
        else:
            # Fallback: usa tela inteira caso API falhe
            work_left, work_top = 0, 0
            work_width, work_height = screen_w, screen_h

        # Calcula largura mínima necessária para exibir todo o frame_cadastro
        try:
            req_width = max(frame_cadastro.winfo_reqwidth(), janela.winfo_width() or 0)
        except Exception:
            req_width = janela.winfo_width() or 860

        # Margem de segurança nas laterais
        base_width = req_width + 20
        # Não ultrapassar a área de trabalho
        width = min(max(base_width, 860), work_width)
        x = work_left + (work_width - width) // 2
        y = work_top

        # Ajusta altura para caber na área de trabalho
        desired_height = work_height
        janela.geometry(f"{width}x{desired_height}+{x}+{y}")
        janela.update_idletasks()

        # Compensa decoração para não ultrapassar limite inferior da área de trabalho
        top_y = janela.winfo_rooty()
        bottom_y = top_y + janela.winfo_height()
        overshoot_bottom = bottom_y - (work_top + work_height)
        if overshoot_bottom != 0:
            adjusted_height = desired_height - overshoot_bottom
            if adjusted_height > 0:
                janela.geometry(f"{width}x{adjusted_height}+{x}+{y}")
                janela.update_idletasks()
        # Garante que a janela não seja menor do que o necessário

    except Exception as e:
        print(f"[AVISO] Não foi possível ajustar altura da janela principal: {e}")


# Agenda ajuste após carregamento inicial dos widgets
janela.after(0, ajustar_altura_principal)

# Exibir lembretes iniciais (apenas os que começam por 'Lembrar')
try:
    janela.after(500, mostrar_lembretes_iniciais)
except Exception as e:
    print(f"[ERRO] Agendando lembretes iniciais: {e}")

selecionar_todos_var = tk.BooleanVar()
situacao_var = tk.StringVar(value="Em Andamento")
toggle_var = tk.BooleanVar(value=False)

checkbox_selecionar_todos = tk.Checkbutton(janela, text="Selecionar Todos", variable=selecionar_todos_var)

# Estatísticas globais
registros_concluidos = 0
registros_andamento = 0
registros_editados = 0
registros_apagados = 0
registros_exportados = 0
registros_importados = 0
registros_restaurados = 0
lembrete_ids = {}  # Dicionário para armazenar IDs dos lembretes


def ajuste_atrasado():
    ajustar_todas_colunas(usar_larguras_personalizadas=True)


def configurar_ajuste_colunas():
    ajuste_timer = None

    def on_configure(event):
        nonlocal ajuste_timer
        if ajuste_timer:
            janela.after_cancel(ajuste_timer)
        ajuste_timer = janela.after(50, ajustar_todas_colunas)

    janela.bind('<Configure>', on_configure)


configurar_ajuste_colunas()
# Configuração de estilo
style = ttk.Style()
style.theme_use("clam")
style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#607D8B", foreground="white")
style.configure("Treeview", font=("Segoe UI", 10), rowheight=26)
style.map("Treeview", background=[("selected", "#B0BEC5")])
style.configure("TButton", font=("Segoe UI", 10, "bold"))
style.configure("TLabel", font=("Segoe UI", 10))
style.configure("TLabelframe", font=("Segoe UI", 11, "bold"))
style.configure("TLabelframe.Label", font=("Segoe UI", 10, "bold"), foreground="#37474F")
style.map("Treeview.Heading", background=[("active", "#455A64")], foreground=[("active", "white")])
style.map("Treeview", foreground=[("selected", "black")])
janela.configure(bg="#ECEFF1")

# Variáveis
situacao_var = tk.StringVar(value="Em Andamento")

# Frame de cadastro
frame_cadastro = tk.LabelFrame(janela, text="Novo Processo", padx=10, pady=5, bg="#ECEFF1", fg="#37474F")
frame_cadastro.pack(fill="x", expand=True, padx=10, pady=0)

label_estatisticas = tk.Label(frame_cadastro,
                              text="Carregando estatísticas...",
                              font=("Segoe UI", 9),  # <-- fonte maior
                              bg="#ECEFF1",
                              fg="#37474F",
                              justify=tk.LEFT)

# Campos do formulário


tk.Label(frame_cadastro, text="Número do contrato:*", bg="#ECEFF1").grid(row=0, column=0, sticky=tk.W)
entrada_numero = tk.Entry(frame_cadastro, font=("Segoe UI", 10))  # Remover width=30
entrada_numero.grid(row=0, column=1, sticky="ew", padx=(4, 0), pady=5)

# tk.Número da Licitação
tk.Label(frame_cadastro, text="Número da Licitação:", bg="#ECEFF1").grid(row=0, column=2, sticky=tk.W, padx=(10, 0))
entrada_licitacao = tk.Entry(frame_cadastro, font=("Segoe UI", 10))
entrada_licitacao.grid(row=0, column=3, sticky=tk.W + tk.E, padx=(4, 0), pady=5)

lista_modalidades = [
    "Dispensa", "Convite", "Tomada de Preços", "Concorrência Pública",
    "Concurso", "Leilão", "Pregão", "Inexigibilidade",
    "Ata de Registro de Preço", "Comparação de Preços",
    "Regras Próprias de Organismos Internacionais",
    "Regime Diferenciado de Contratações (RDC)",
    "Chamada Pública", "Não se Aplica", "Diálogo Competitivo"
]

tk.Label(frame_cadastro, text="Recebimento:", bg="#ECEFF1").grid(row=2, column=0, sticky=tk.W)
entrada_recebimento = tk.Entry(frame_cadastro, font=("Segoe UI", 10))
entrada_recebimento.grid(row=2, column=1, sticky=tk.W + tk.E, padx=(4, 0), pady=5)
entrada_recebimento.bind("<KeyRelease>", formatar_data_hora)
entrada_recebimento.bind("<FocusOut>", lambda e: checar_data_entry(entrada_recebimento))

# Secretaria
tk.Label(frame_cadastro, text="Secretaria:*", bg="#ECEFF1").grid(row=1, column=0, sticky=tk.W)
entrada_secretaria = AutocompleteEntry(
    frame_cadastro,
    secretarias_formatadas,
    font=("Segoe UI", 10)
)
entrada_secretaria.set_listbox_properties(
    x_offset=0,  # Ajuste fino horizontal
    y_offset=5,  # Ajuste fino vertical
    width=350,  # Largura suficiente para mostrar sigla + nome
    max_height=14
)
entrada_secretaria.grid(row=1, column=1, sticky=tk.W + tk.E, padx=(4, 0), pady=5)
entrada_secretaria.delete(0, tk.END)

# Ativar seleção pelo Enter, Tab e clique do mouse
entrada_secretaria.bind("<Return>", entrada_secretaria.on_enter)
entrada_secretaria.bind("<Tab>", entrada_secretaria.on_tab)
entrada_secretaria.bind("<ButtonRelease-1>", entrada_secretaria.on_listbox_click)

# Modalidade
tk.Label(frame_cadastro, text="Modalidade:", bg="#ECEFF1").grid(row=1, column=2, sticky=tk.W, padx=(10, 0))
entrada_modalidade = AutocompleteEntry(
    frame_cadastro,
    modalidades_licitacao,
    font=("Segoe UI", 10)
)
entrada_modalidade.set_listbox_properties(
    x_offset=0,
    y_offset=5,
    width=250,
    max_height=15
)
entrada_modalidade.grid(row=1, column=3, sticky=tk.W + tk.E, padx=(4, 0), pady=5)
entrada_modalidade.delete(0, tk.END)

# Ativar seleção pelo Enter, Tab e clique do mouse
entrada_modalidade.bind("<Return>", entrada_modalidade.on_enter)
entrada_modalidade.bind("<Tab>", entrada_modalidade.on_tab)
entrada_modalidade.bind("<ButtonRelease-1>", entrada_modalidade.on_listbox_click)

# Devolução
tk.Label(frame_cadastro, text="Devolução:", bg="#ECEFF1").grid(row=2, column=2, sticky=tk.W, padx=(10, 0))
entrada_devolucao = tk.Entry(frame_cadastro, font=("Segoe UI", 10))
entrada_devolucao.grid(row=2, column=3, sticky=tk.W + tk.E, padx=(4, 0), pady=5)
entrada_devolucao.bind("<KeyRelease>", formatar_data_hora)
entrada_devolucao.bind("<FocusOut>", lambda e: checar_data_entry(entrada_devolucao))

# Para Entregue por

tk.Label(frame_cadastro, text="Entregue por:", bg="#ECEFF1").grid(row=3, column=0, sticky=tk.W)
entrada_entregue_por = AutocompleteEntry(frame_cadastro, nomes_autocomplete, font=("Segoe UI", 10))
entrada_entregue_por.grid(row=3, column=1, sticky=tk.W + tk.E, padx=(4, 0), pady=5)
entrada_entregue_por.set_listbox_properties(
    x_offset=-1,  # Ajuste horizontal (positivo = direita, negativo = esquerda)
    y_offset=1  # Ajuste vertical (positivo = abaixo, negativo = acima)
)

# Devolvido a
tk.Label(frame_cadastro, text="Devolvido a:", bg="#ECEFF1").grid(row=3, column=2, sticky=tk.W, padx=(10, 0))
entrada_devolvido_a = AutocompleteEntry(frame_cadastro, nomes_autocomplete, font=("Segoe UI", 10))
entrada_devolvido_a.grid(row=3, column=3, sticky=tk.W + tk.E, padx=(4, 0), pady=5)
entrada_devolvido_a.set_listbox_properties(
    x_offset=-1,  # Ajuste horizontal
    y_offset=1  # Ajuste vertical
)

# Situação
tk.Label(frame_cadastro, text="Situação:", bg="#ECEFF1").grid(row=4, column=0, sticky=tk.W)
frame_situacao = tk.Frame(frame_cadastro, bg="#ECEFF1")
frame_situacao.grid(row=4, column=1, sticky=tk.W, padx=10, pady=5)
tk.Radiobutton(frame_situacao, text="Em Andamento", variable=situacao_var, value="Em Andamento", bg="#ECEFF1",
               font=("Segoe UI", 10), command=ativar_botao_atualizar).pack(side=tk.LEFT)
tk.Radiobutton(frame_situacao, text="Concluído", variable=situacao_var, value="Concluído", bg="#ECEFF1",
               font=("Segoe UI", 10), command=ativar_botao_atualizar).pack(side=tk.LEFT, padx=5)

# Contratado
tk.Label(frame_cadastro, text="Contratado:", bg="#ECEFF1", font=("Segoe UI", 10)).grid(row=4, column=2, sticky=tk.W,
                                                                                       padx=(10, 0))
entrada_contratado = AutocompleteEntry(frame_cadastro, nomes_contratado, font=("Segoe UI", 10))
entrada_contratado.grid(row=4, column=3, sticky=tk.W + tk.E, padx=(4, 0), pady=5)
entrada_contratado.bind("<KeyRelease>", atualizar_campo_contratado)

# Configurar a posição e o tamanho da listbox do campo Contratado
entrada_contratado.set_listbox_properties(
    x_offset=0,  # Ajuste horizontal (em pixels)
    y_offset=2,  # Ajuste vertical (em pixels)
    width=235  # Largura da listbox em pixels (ajuste este valor conforme necessário)
)


# Sobrescrever o comportamento da tecla seta para baixo para o campo Contratado
def contratado_arrow_down(event):
    # Mostrar sugestões apenas se já houver texto digitado
    valor = entrada_contratado.get().strip().upper()
    if valor:  # tk.Só mostra sugestões se já tiver algo digitado
        valor_sem_acento = remover_acentos(valor)
        matches = [
            item for item in nomes_contratado
            if valor_sem_acento in remover_acentos(str(item).upper())
        ]
        if matches:
            entrada_contratado.show_suggestions(matches)
    else:
        # Se o campo estiver vazio, esconde as sugestões
        entrada_contratado.hide_suggestions()
    return "break"


entrada_contratado.bind("<Down>", contratado_arrow_down)


def configurar_binds_autocomplete():
    campos_autocomplete = [
        entrada_secretaria,
        entrada_modalidade,
        entrada_entregue_por,
        entrada_devolvido_a,
        entrada_contratado
    ]

    for campo in campos_autocomplete:
        campo.bind("<Return>", campo.on_enter)
        campo.bind("<Tab>", campo.on_tab)
        campo.bind("<ButtonRelease-1>", campo.on_listbox_click)


# Chamar esta função uma vez após criar a interface
configurar_binds_autocomplete()

entrada_numero.bind("<Button-1>", ativar_edicao_campo)
entrada_secretaria.bind("<Button-1>", ativar_edicao_campo)
entrada_numero.bind("<Button-1>", ativar_edicao_campo)
entrada_secretaria.bind("<Button-1>", ativar_edicao_campo)
entrada_licitacao.bind("<Button-1>", ativar_edicao_campo)
entrada_modalidade.bind("<Button-1>", ativar_edicao_campo)
entrada_recebimento.bind("<Button-1>", ativar_edicao_campo)
entrada_devolucao.bind("<Button-1>", ativar_edicao_campo)
entrada_entregue_por.bind("<Button-1>", ativar_edicao_campo)
entrada_devolvido_a.bind("<Button-1>", ativar_edicao_campo)
entrada_contratado.bind("<Button-1>", ativar_edicao_campo)


def configurar_tab_ordem():
    """
    Configura a ordem de foco (Tab e Shift+Tab) de todos os campos da interface,
    mantendo o suporte ao autocomplete quando aplicável.
    """

    def mover_foco(event, anterior=None, proximo=None):
        """
        Move o foco para o widget anterior (Shift+Tab) ou próximo (Tab).
        Detecta automaticamente a direção conforme o estado da tecla Shift.
        """
        if event.state & 0x0001:  # SHIFT pressionado → foco reverso
            if anterior:
                anterior.focus_set()
        else:  # Tab normal
            if proximo:
                proximo.focus_set()
        return "break"

    # ============================================================
    # 🔹 Ordem de foco ajustada — inclui o Checkbutton "Selecionar Todos"
    # ============================================================
    ordem_foco = [
        (toggle_btn, None, entrada_numero),  # Novo primeiro elemento
        (entrada_numero, toggle_btn, entrada_licitacao),
        (entrada_licitacao, entrada_numero, entrada_secretaria),
        (entrada_secretaria, entrada_licitacao, entrada_modalidade),
        (entrada_modalidade, entrada_secretaria, entrada_recebimento),
        (entrada_recebimento, entrada_modalidade, entrada_devolucao),
        (entrada_devolucao, entrada_recebimento, entrada_entregue_por),
        (entrada_entregue_por, entrada_devolucao, entrada_devolvido_a),
        (entrada_devolvido_a, entrada_entregue_por, frame_situacao.winfo_children()[0]),
    ]

    # ============================================================
    # 🔹 Aplica binds unificados de Tab e Shift+Tab
    # ============================================================
    for campo, anterior, proximo in ordem_foco:
        # Armazena o próximo e o anterior configurado para uso por widgets especiais (ex.: AutocompleteEntry)
        try:
            setattr(campo, 'tab_prev', anterior)
            setattr(campo, 'tab_next', proximo)
        except Exception:
            pass
        campo.unbind("<Tab>")
        campo.unbind("<Shift-Tab>")

        campo.bind("<Tab>", lambda e, a=anterior, p=proximo: mover_foco(e, a, p))
        campo.bind("<Shift-Tab>", lambda e, a=anterior, p=proximo: mover_foco(e, a, p))

        # Se for um AutocompleteEntry, mantém suporte à listbox
        if hasattr(campo, "on_tab"):
            campo.bind(
                "<Tab>",
                lambda e, f=campo, a=anterior, p=proximo:
                f.on_tab(e) if f.listbox else mover_foco(e, a, p)
            )
            campo.bind("<Shift-Tab>", lambda e, f=campo, a=anterior, p=proximo: mover_foco(e, a, p))

    # ===========================================
    # CAMPOS ENTREGUE POR / DEVOLVIDO A — AJUSTES
    # ===========================================
    entrada_entregue_por.bind(
        "<Tab>",
        lambda e: entrada_entregue_por.on_tab(e)
        if entrada_entregue_por.listbox
        else (entrada_devolvido_a.focus_set() or "break")
    )
    entrada_entregue_por.bind(
        "<Return>",
        lambda e: entrada_entregue_por.on_enter(e)
        if entrada_entregue_por.listbox
        else (entrada_devolvido_a.focus_set() or "break")
    )

    entrada_devolvido_a.bind(
        "<Tab>",
        lambda e: entrada_devolvido_a.on_tab(e)
        if entrada_devolvido_a.listbox
        else (frame_situacao.winfo_children()[0].focus_set() or "break")
    )
    entrada_devolvido_a.bind(
        "<Return>",
        lambda e: entrada_devolvido_a.on_enter(e)
        if entrada_devolvido_a.listbox
        else (frame_situacao.winfo_children()[0].focus_set() or "break")
    )

    # ===========================================
    # CAMPOS DE SITUAÇÃO tk.E CONTRATADO
    # ===========================================
    frame_situacao.winfo_children()[0].bind(
        "<Tab>", lambda e: frame_situacao.winfo_children()[1].focus_set() or "break"
    )
    frame_situacao.winfo_children()[1].bind(
        "<Tab>", lambda e: entrada_contratado.focus_set() or "break"
    )

    entrada_contratado.bind(
        "<Tab>",
        lambda e: entrada_contratado.on_tab(e)
        if entrada_contratado.listbox
        else (entrada_descricao.focus_set() or "break")
    )

    # ===========================================
    # BOTÕES tk.E CAMPOS FINAIS
    # ===========================================
    botao_lembrete.bind("<Tab>", lambda e: check_lembrete.focus_set() or "break")
    check_lembrete.bind("<Tab>", lambda e: botao_cadastrar.focus_set() or "break")

    botao_cadastrar.bind("<Tab>", lambda e: botao_limpar.focus_set() or "break")
    botao_limpar.bind("<Tab>", lambda e: botao_editar.focus_set() or "break")
    botao_editar.bind("<Tab>", lambda e: botao_excluir.focus_set() or "break")
    botao_excluir.bind("<Tab>", lambda e: botao_exportar.focus_set() or "break")
    botao_exportar.bind("<Tab>", lambda e: botao_exportar_txt.focus_set() or "break")
    botao_exportar_txt.bind("<Tab>", lambda e: botao_exportar_excel.focus_set() or "break")
    botao_exportar_excel.bind("<Tab>", lambda e: botao_banco_dados.focus_set() or "break")
    botao_banco_dados.bind("<Tab>", lambda e: botao_restaurar.focus_set() or "break")

    # ===========================================
    # 🔹 Ordem reversa (Shift+Tab)
    # ===========================================
    # Corrigir: Shift+Tab em "Número" deve focar em "Selecionar Todos"
    entrada_numero.bind("<Shift-Tab>", lambda e: toggle_btn.focus_set() or "break")
    toggle_btn.bind("<Tab>", lambda e: entrada_numero.focus_set() or "break")
    # E Shift+Tab em "Selecionar Todos" também volta para "Limpar Filtros"
    toggle_btn.bind("<Shift-Tab>", lambda e: botao_limpar_filtros.focus_set() or "break")

    entrada_licitacao.bind("<Shift-Tab>", lambda e: entrada_numero.focus_set() or "break")
    entrada_secretaria.bind("<Shift-Tab>", lambda e: entrada_licitacao.focus_set() or "break")
    entrada_modalidade.bind("<Shift-Tab>", lambda e: entrada_secretaria.focus_set() or "break")
    entrada_recebimento.bind("<Shift-Tab>", lambda e: entrada_modalidade.focus_set() or "break")
    entrada_devolucao.bind("<Shift-Tab>", lambda e: entrada_recebimento.focus_set() or "break")
    entrada_entregue_por.bind("<Shift-Tab>", lambda e: entrada_devolucao.focus_set() or "break")
    entrada_devolvido_a.bind("<Shift-Tab>", lambda e: entrada_entregue_por.focus_set() or "break")

    frame_situacao.winfo_children()[0].bind(
        "<Shift-Tab>", lambda e: entrada_devolvido_a.focus_set() or "break"
    )
    frame_situacao.winfo_children()[1].bind(
        "<Shift-Tab>", lambda e: frame_situacao.winfo_children()[0].focus_set() or "break"
    )
    entrada_contratado.bind(
        "<Shift-Tab>", lambda e: frame_situacao.winfo_children()[1].focus_set() or "break"
    )
    entrada_descricao.bind("<Shift-Tab>", lambda e: entrada_contratado.focus_set() or "break")

    botao_lembrete.bind("<Shift-Tab>", lambda e: entrada_descricao.focus_set() or "break")
    check_lembrete.bind("<Shift-Tab>", lambda e: botao_lembrete.focus_set() or "break")

    botao_cadastrar.bind("<Shift-Tab>", lambda e: check_lembrete.focus_set() or "break")
    botao_limpar.bind("<Shift-Tab>", lambda e: botao_cadastrar.focus_set() or "break")
    botao_editar.bind("<Shift-Tab>", lambda e: botao_limpar.focus_set() or "break")
    botao_excluir.bind("<Shift-Tab>", lambda e: botao_editar.focus_set() or "break")
    botao_exportar.bind("<Shift-Tab>", lambda e: botao_excluir.focus_set() or "break")
    botao_exportar_txt.bind("<Shift-Tab>", lambda e: botao_exportar.focus_set() or "break")
    botao_exportar_excel.bind("<Shift-Tab>", lambda e: botao_exportar_txt.focus_set() or "break")
    botao_banco_dados.bind("<Shift-Tab>", lambda e: botao_exportar_excel.focus_set() or "break")
    botao_restaurar.bind("<Shift-Tab>", lambda e: botao_banco_dados.focus_set() or "break")


def sair_texto_tab(event):
    botao_lembrete.focus_set()
    return "break"


# Defina a função ANTES de usá-la no bind
def ativar_botao_ao_modificar_observacoes(event):
    """Ativa o botão Atualizar quando o texto das observações é modificado"""
    try:
        item_selecionado = tabela.focus()
        if item_selecionado:
            processo = tabela.item(item_selecionado)['values']
            if len(processo) > 1:  # Verifica se há dados suficientes
                numero_original = str(processo[1]).strip().upper()
                botao_cadastrar.config(
                    text="Atualizar",
                    state='normal',
                    command=lambda: atualizar_processo(numero_original)
                )
    except (IndexError, KeyError):
        # Se não há processo selecionado, não ativa o botão
        pass


# Criar o frame para lembretes
frame_lembretes = tk.Frame(frame_cadastro, bg="#ECEFF1")
frame_lembretes.grid(row=5, column=0, columnspan=3, pady=(15, 0), sticky=tk.W)
# Configurar as colunas do frame_cadastro para expandir

# Alterar as linhas 4122-4125 para:
frame_cadastro.grid_columnconfigure(0, weight=0)
try:
    col_width_px = int(frame_cadastro.winfo_fpixels('7.5c') - frame_cadastro.winfo_fpixels('1c'))
except Exception:
    col_width_px = 280
left_width_px = max(int(col_width_px * 0.9), 1)
frame_cadastro.grid_columnconfigure(1, weight=1, minsize=left_width_px)
frame_cadastro.grid_columnconfigure(2, weight=0)
right_width_px = left_width_px
frame_cadastro.grid_columnconfigure(3, weight=1, minsize=right_width_px)

tk.Label(frame_cadastro, text="Observações:", bg="#ECEFF1", font=("Segoe UI", 10)).grid(row=5, column=0, sticky=tk.NW,
                                                                                        pady=(10, 0))
# Frame com sombreamento muito sutil para as Observações (50% reduzido)
frame_obs = tk.Frame(frame_cadastro, bg="white", relief="ridge", bd=1)
frame_obs.grid(row=5, column=1, columnspan=3, sticky=tk.W + tk.E + tk.N + tk.S, padx=0, pady=(4, 0))

# Caixa de texto de Observações dentro do frame com borda e scrollbar vertical
scrollbar_descricao = tk.Scrollbar(frame_obs)
scrollbar_descricao.pack(side=tk.RIGHT, fill=tk.Y)

entrada_descricao = tk.Text(frame_obs, width=90, height=4, font=("Segoe UI", 10), wrap=tk.WORD, bg="white")
entrada_descricao.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
entrada_descricao.configure(yscrollcommand=scrollbar_descricao.set)
scrollbar_descricao.config(command=entrada_descricao.yview)

entrada_descricao.bind("<KeyRelease>", lambda event: atualizar_cor_botao_lembrete_checkbox())
entrada_descricao.bind("<Button-1>", ativar_edicao_campo)
try:
    enhance_text_controls(entrada_descricao)
except Exception:
    pass

# Criar o frame para lembretes logo após as observações
frame_lembretes = tk.Frame(frame_cadastro, bg="#ECEFF1")
frame_lembretes.grid(row=5, column=0, columnspan=3, pady=(35, 0),
                     sticky=tk.W)  # Posicionado logo abaixo das observações

lembrete_var = tk.BooleanVar(value=False)
botao_lembrete = tk.Button(
    frame_lembretes,
    text="Lembretes",
    command=abrir_lembretes,
    bg="#ECEFF1",  # Fundo transparente (mesmo do frame)
    fg="#37474F",  # Texto cinza escuro (padrão)
    font=("Segoe UI", 9),  # Fonte normal (sem negrito)
    relief=tk.FLAT,  # Sem relevo
    bd=0,  # Sem borda
    activebackground="#ECEFF1",  # Mantém fundo transparente ao clicar
    activeforeground="#263238",  # Texto mais escuro ao clicar
    highlightthickness=0  # Remove contorno completamente
)
botao_lembrete.pack(side=tk.LEFT, padx=(10, 5))
atualizar_cor_botao_lembrete()  # Atualiza a cor do botão na inicialização


# Destaque suave no hover do botão "Lembretes"
def _hover_destacar_botao_lembrete(event=None):
    try:
        botao_lembrete.config(bg="#DDE7F0", relief=tk.RAISED, bd=1)
    except Exception:
        pass


def _hover_remover_destaque_botao_lembrete(event=None):
    try:
        # Se o botão estiver com foco, mantenha o relevo; apenas restaure o fundo
        if botao_lembrete == botao_lembrete.focus_get():
            botao_lembrete.config(bg="#ECEFF1", relief=tk.RAISED, bd=1)
        else:
            botao_lembrete.config(bg="#ECEFF1", relief=tk.FLAT, bd=0)
    except Exception:
        pass


# Ajusta o fundo ao clicar e adiciona bindings de hover
botao_lembrete.configure(activebackground="#DDE7F0")
botao_lembrete.bind("<Enter>", _hover_destacar_botao_lembrete)
botao_lembrete.bind("<Leave>", _hover_remover_destaque_botao_lembrete)


# Destaque por foco via teclado (TAB/SHIFT+TAB)
def _focus_destacar_botao_lembrete(event=None):
    try:
        botao_lembrete.config(bg="#DDE7F0", relief=tk.RAISED, bd=1)
    except Exception:
        pass


def _focus_remover_destaque_botao_lembrete(event=None):
    try:
        # Ao perder o foco, volta ao estado padrão
        botao_lembrete.config(bg="#ECEFF1", relief=tk.FLAT, bd=0)
    except Exception:
        pass


botao_lembrete.bind("<FocusIn>", _focus_destacar_botao_lembrete)
botao_lembrete.bind("<FocusOut>", _focus_remover_destaque_botao_lembrete)

check_lembrete = tk.Checkbutton(frame_lembretes, variable=lembrete_var,
                                bg="#ECEFF1", activebackground="#ECEFF1",
                                selectcolor="#C8E6C9",
                                highlightthickness=0, highlightbackground="#ECEFF1", highlightcolor="#ECEFF1",
                                command=lambda: [toggle_lembrete(), atualizar_cor_botao_lembrete_checkbox()])
check_lembrete.pack(side=tk.LEFT, padx=(5, 10))  # Adicione esta linha


# Destaque em verde claro para o checkbox de lembrete (hover/foco)
def _highlight_check_lembrete_on(event=None):
    try:
        check_lembrete.config(bg="#DFF2E1", activebackground="#DFF2E1", selectcolor="#4CAF50")
    except Exception:
        pass


def _highlight_check_lembrete_off(event=None):
    try:
        check_lembrete.config(bg="#ECEFF1", activebackground="#ECEFF1", selectcolor="#C8E6C9")
    except Exception:
        pass


check_lembrete.bind("<Enter>", _highlight_check_lembrete_on)
check_lembrete.bind("<Leave>", _highlight_check_lembrete_off)
check_lembrete.bind("<FocusIn>", _highlight_check_lembrete_on)
check_lembrete.bind("<FocusOut>", _highlight_check_lembrete_off)

# Mantenha o frame_botoes existente abaixo (expandindo horizontalmente)
frame_botoes_container = tk.Frame(frame_cadastro, bg="#ECEFF1")
frame_botoes_container.grid(row=6, column=0, columnspan=4, pady=10, sticky="we")


def sair_texto_tab(event):
    event.widget.tk_focusNext().focus()
    return "break"


entrada_descricao.bind("<Tab>", sair_texto_tab)


def sair_texto_shift_tab(event):
    event.widget.tk_focusPrev().focus()
    return "break"


entrada_descricao.bind("<Shift-Tab>", sair_texto_shift_tab)

# Removido o bind que ativava o botão ao modificar observações
entrada_descricao.bind("<Button-1>", ativar_edicao_campo)

frame_botoes = tk.Frame(frame_botoes_container, bg="#ECEFF1")
frame_botoes.pack(anchor="center", fill="x")  # Centraliza a barra e permite expansão horizontal

# Frame interno para centralizar o grupo de botões
row_frame = tk.Frame(frame_botoes, bg="#ECEFF1")
row_frame.pack(anchor="center")


# Espaçadores flexíveis para manter espaçamento igual entre botões e nas extremidades
def add_spacer(parent, min_width=8, expand=False):
    """Adiciona um espaçamento horizontal fixo entre botões."""
    s = tk.Frame(parent, bg="#ECEFF1", width=min_width)
    s.pack(side=tk.LEFT, fill='x', expand=expand)
    return s


estilo_botao_padrao = {
    "bg": BUTTON_PRIMARY_BG,
    "fg": BUTTON_PRIMARY_FG,
    "activebackground": BUTTON_PRIMARY_ACTIVE_BG,
    "activeforeground": BUTTON_PRIMARY_ACTIVE_FG,
    "font": ("Segoe UI", 10, "bold"),
    "relief": tk.FLAT,
    "bd": 0,
    "highlightthickness": 2,
    "highlightbackground": BUTTON_PRIMARY_HIGHLIGHT
}
botao_cadastrar = tk.Button(row_frame, text="Cadastrar", command=cadastrar_processo, width=7, **estilo_botao_padrao)
botao_cadastrar.pack(side=tk.LEFT, padx=0, ipadx=3)
add_spacer(row_frame, min_width=4)

botao_limpar = tk.Button(row_frame, text="Limpar", command=limpar_campos, width=5, **estilo_botao_padrao)
botao_limpar.pack(side=tk.LEFT, padx=0, ipadx=3)
add_spacer(row_frame, min_width=4)

botao_editar = tk.Button(row_frame, text="Editar", command=editar_processo, width=5, **estilo_botao_padrao)
botao_editar.pack(side=tk.LEFT, padx=0, ipadx=3)
add_spacer(row_frame, min_width=4)

estilo_excluir = dict(estilo_botao_padrao)
estilo_excluir.update({
    "bg": BUTTON_DANGER_BG,
    "fg": BUTTON_DANGER_FG,
    "activebackground": BUTTON_DANGER_ACTIVE_BG,
    "activeforeground": BUTTON_DANGER_ACTIVE_FG,
    "highlightbackground": BUTTON_DANGER_HIGHLIGHT
})
botao_excluir = tk.Button(row_frame, text="Excluir", command=excluir_processo, width=6, **estilo_excluir)
botao_excluir.pack(side=tk.LEFT, padx=0, ipadx=3)
add_spacer(row_frame, min_width=4)

botao_exportar = tk.Button(row_frame, text="Exportar PDF", command=exportar_pdf, width=10, **estilo_botao_padrao)
botao_exportar.pack(side=tk.LEFT, padx=0, ipadx=4)
add_spacer(row_frame, min_width=4)

botao_exportar_txt = tk.Button(row_frame, text="Exportar TXT", command=exportar_txt, width=10, **estilo_botao_padrao)
botao_exportar_txt.pack(side=tk.LEFT, padx=0, ipadx=4)
add_spacer(row_frame, min_width=4)
botao_exportar_txt.bind("<Return>", lambda event: botao_exportar_txt.invoke())

botao_exportar_excel = tk.Button(row_frame, text="Exportar Excel", command=exportar_excel, width=11,
                                 **estilo_botao_padrao)
botao_exportar_excel.pack(side=tk.LEFT, padx=0, ipadx=4)
add_spacer(row_frame, min_width=4)
botao_exportar_excel.bind("<Return>", lambda event: botao_exportar_excel.invoke())

botao_banco_dados = tk.Button(row_frame, text="Banco de Dados", command=abrir_janela_banco_dados, width=12,
                              **estilo_botao_padrao)
botao_banco_dados.pack(side=tk.LEFT, padx=0, ipadx=4)
add_spacer(row_frame, min_width=4)

botao_restaurar = tk.Button(
    row_frame, text="Restaurar Registros",
    command=abrir_janela_restaurar,
    width=14,
    **estilo_botao_padrao
)
botao_restaurar.pack(side=tk.LEFT, padx=0, ipadx=5)
add_spacer(row_frame, min_width=4)

# Botão ATALHOS removido conforme solicitação do usuário
# Mantendo apenas a funcionalidade do F1

# botão Enter
botao_cadastrar.bind("<Return>", lambda event: cadastrar_processo())
botao_limpar.bind("<Return>", lambda event: botao_limpar.invoke())
botao_editar.bind("<Return>", lambda event: botao_editar.invoke())
botao_excluir.bind("<Return>", lambda event: botao_excluir.invoke())
botao_exportar.bind("<Return>", lambda event: botao_exportar.invoke())
botao_banco_dados.bind("<Return>", lambda event: botao_banco_dados.invoke())
botao_restaurar.bind("<Return>", lambda event: botao_restaurar.invoke())

# Frame de busca

container_busca = tk.Frame(janela, bg="#ECEFF1")
container_busca.pack(fill="both")

frame_busca = tk.LabelFrame(container_busca, text="Buscar Processos", padx=10, pady=10, bg="#ECEFF1", fg="#37474F")
frame_busca.pack(fill="x", expand=True, pady=2)

# Permite que os campos de busca se expandam quando a janela é maximizada
try:
    frame_busca.grid_columnconfigure(0, weight=2)  # Buscar
    frame_busca.grid_columnconfigure(1, weight=3)  # Secretaria
    frame_busca.grid_columnconfigure(2, weight=1)  # Situação
    frame_busca.grid_columnconfigure(3, weight=3)  # Modalidade
    frame_busca.grid_columnconfigure(4, weight=0)  # Botão Buscar
    frame_busca.grid_columnconfigure(5, weight=0)  # Botão Limpar (se existir)
except Exception:
    pass

# Primeira linha: rótulos
tk.Label(frame_busca, text="Buscar:", bg="#ECEFF1").grid(row=0, column=0, sticky=tk.W, padx=0)
tk.Label(frame_busca, text="Secretaria:", bg="#ECEFF1").grid(row=0, column=1, sticky=tk.W, padx=0)
tk.Label(frame_busca, text="Situação:", bg="#ECEFF1").grid(row=0, column=2, sticky=tk.W, padx=0)
tk.Label(frame_busca, text="Modalidade:", bg="#ECEFF1").grid(row=0, column=3, sticky=tk.W, padx=0)

# Segunda linha: caixas de texto/autocomplete
entrada_busca = tk.Entry(frame_busca, width=22, font=("Segoe UI", 10))
entrada_busca.grid(row=1, column=0, sticky=tk.E + tk.W, padx=5, pady=0)

# Ativa busca ao pressionar Enter na caixa de texto Buscar
entrada_busca.bind("<Return>", lambda event: buscar_processos())


# Quando o texto de busca é apagado, recarrega a tabela automaticamente
def on_busca_keyrelease(event):
    try:
        texto = entrada_busca.get().strip()
        if texto == "":
            # Se houver filtros ativos, respeita-os usando buscar_processos();
            # caso contrário, lista todos os registros
            if (entrada_filtro_secretaria.get().strip() or
                    entrada_filtro_situacao.get().strip() or
                    entrada_filtro_modalidade.get().strip()):
                buscar_processos()
            else:
                listar_processos()
            return "break"
    except Exception as e:
        print(f"[ERRO] on_busca_keyrelease: {e}")
    return None


entrada_busca.bind("<KeyRelease>", on_busca_keyrelease)

# Campos de filtro da busca
entrada_filtro_secretaria = AutocompleteEntry(
    frame_busca, secretarias_formatadas, width=33, font=("Segoe UI", 10)
)
entrada_filtro_secretaria.set_listbox_properties(x_offset=0, y_offset=5, width=280, max_height=10)
entrada_filtro_secretaria.completion_list = secretarias_formatadas
entrada_filtro_secretaria.grid(row=1, column=1, sticky=tk.E + tk.W, padx=1, pady=0)

entrada_filtro_situacao = AutocompleteEntry(
    frame_busca, ["Em Andamento", "Concluído"], width=14, font=("Segoe UI", 10)
)
entrada_filtro_situacao.set_listbox_properties(x_offset=0, y_offset=5, width=120, max_height=2)
entrada_filtro_situacao.completion_list = ["Em Andamento", "Concluído"]
entrada_filtro_situacao.grid(row=1, column=2, sticky=tk.E + tk.W, padx=5, pady=0)

entrada_filtro_modalidade = AutocompleteEntry(
    frame_busca,
    modalidades_licitacao,  # Use a mesma lista ordenada
    font=("Segoe UI", 10)
)
entrada_filtro_modalidade.set_listbox_properties(x_offset=0, y_offset=5, width=250, max_height=15)
entrada_filtro_modalidade.completion_list = modalidades_licitacao
entrada_filtro_modalidade.grid(row=1, column=3, sticky=tk.E + tk.W, padx=5, pady=0)

botao_buscar = tk.Button(frame_busca, text="Buscar", command=buscar_processos,
                         bg=BUTTON_PRIMARY_BG, fg=BUTTON_PRIMARY_FG,
                         activebackground=BUTTON_PRIMARY_ACTIVE_BG, activeforeground=BUTTON_PRIMARY_ACTIVE_FG, width=6,
                         font=("Segoe UI", 10, "bold"), relief=tk.FLAT, bd=0, highlightthickness=1,
                         highlightbackground=BUTTON_PRIMARY_HIGHLIGHT)
botao_buscar.grid(row=1, column=4, padx=4, sticky=tk.W)


def enter_busca_ordenar(event):
    """Realiza a busca ou ordenação conforme o campo de busca em foco."""
    busca = entrada_busca.get().strip()
    secretaria = entrada_filtro_secretaria.get().strip()
    situacao = entrada_filtro_situacao.get().strip()
    modalidade = entrada_filtro_modalidade.get().strip()

    # Verifica se algum campo de busca está preenchido
    if busca or secretaria or situacao or modalidade:
        # Ativa o botão "Buscar" (ou chama a função associada ao botão)
        botao_buscar.invoke()  # Substitua 'botao_buscar' pelo nome da sua variável do botão Buscar

    else:
        # Realiza a ordenação conforme o foco do campo
        widget = event.widget
        if widget == entrada_filtro_secretaria:
            ordenar_coluna('secretaria')
        elif widget == entrada_filtro_situacao:
            ordenar_coluna('situacao')
        elif widget == entrada_filtro_modalidade:
            ordenar_coluna('modalidade')

    # Avança o foco para o próximo campo de entrada
    widget.tk_focusNext().focus()
    return "break"


# Bind especial para os filtros
entrada_filtro_secretaria.bind("<Return>", enter_busca_ordenar)
entrada_filtro_situacao.bind("<Return>", enter_busca_ordenar)

# O campo "Buscar" continua apenas buscando normalmente
entrada_busca.bind("<Return>", lambda event: buscar_processos())


def enter_como_tab(event):
    widget = event.widget

    # Se for um AutocompleteEntry com listbox aberta → insere item
    if isinstance(widget, AutocompleteEntry) and widget.listbox:
        return widget.on_enter(event)

    # Caso contrário, simula o comportamento do TAB conforme configurado
    if widget == entrada_numero:
        entrada_licitacao.focus_set()
    elif widget == entrada_licitacao:
        entrada_secretaria.focus_set()
    elif widget == entrada_secretaria:
        entrada_modalidade.focus_set()
    elif widget == entrada_modalidade:
        entrada_recebimento.focus_set()  # Garante que vá para Recebimento
    elif widget == entrada_recebimento:
        entrada_devolucao.focus_set()
    elif widget == entrada_devolucao:
        entrada_entregue_por.focus_set()
    elif widget == entrada_entregue_por:
        entrada_devolvido_a.focus_set()
    elif widget == entrada_devolvido_a:
        frame_situacao.winfo_children()[0].focus_set()
    elif widget == frame_situacao.winfo_children()[0]:
        frame_situacao.winfo_children()[1].focus_set()
    elif widget == frame_situacao.winfo_children()[1]:
        entrada_contratado.focus_set()
    elif widget == entrada_contratado:
        entrada_descricao.focus_set()

    return "break"


# Lista de campos de texto que tk.NÃO são de busca
campos_texto = [
    entrada_numero, entrada_licitacao, entrada_secretaria, entrada_modalidade,
    entrada_recebimento, entrada_devolucao, entrada_entregue_por, entrada_devolvido_a
]

for widget in campos_texto:
    widget.bind("<Return>", enter_como_tab)

# Para o campo de observações (Text), Enter não deve ser Tab, mas Tab já está tratado acima

botao_limpar_filtros = tk.Button(frame_busca, text="Limpar", command=limpar_filtros,
                                 bg=BUTTON_PRIMARY_BG, fg=BUTTON_PRIMARY_FG,
                                 activebackground=BUTTON_PRIMARY_ACTIVE_BG, activeforeground=BUTTON_PRIMARY_ACTIVE_FG,
                                 width=6,
                                 font=("Segoe UI", 10, "bold"), relief=tk.FLAT, bd=0, highlightthickness=1,
                                 highlightbackground=BUTTON_PRIMARY_HIGHLIGHT)
botao_limpar_filtros.grid(row=1, column=5, padx=(5, 4))

# Configuração dos botões (no código de inicialização)
botao_buscar.config(command=buscar_processos)
botao_limpar_filtros.config(command=limpar_filtros)

campos_ordem = [
    entrada_numero,
    entrada_licitacao,
    entrada_secretaria,
    entrada_modalidade,
    entrada_recebimento,
    entrada_devolucao,
    entrada_entregue_por,
    entrada_devolvido_a,
    entrada_contratado,
    entrada_descricao  # Observações (Text)
]


def focar_proximo_campo(event):
    widget = event.widget

    # Definir a ordem explicitamente
    if widget == entrada_numero:
        entrada_licitacao.focus_set()
    elif widget == entrada_licitacao:
        entrada_secretaria.focus_set()
    elif widget == entrada_secretaria:
        entrada_modalidade.focus_set()
    elif widget == entrada_modalidade:
        entrada_recebimento.focus_set()  # Garante que vá para Recebimento
    elif widget == entrada_recebimento:
        entrada_devolucao.focus_set()
    elif widget == entrada_devolucao:
        entrada_entregue_por.focus_set()
    elif widget == entrada_entregue_por:
        entrada_devolvido_a.focus_set()
    elif widget == entrada_devolvido_a:
        frame_situacao.winfo_children()[0].focus_set()
    elif widget == frame_situacao.winfo_children()[0]:
        frame_situacao.winfo_children()[1].focus_set()
    elif widget == frame_situacao.winfo_children()[1]:
        entrada_contratado.focus_set()
    elif widget == entrada_contratado:
        entrada_descricao.focus_set()

    return "break"


# Configuração do Enter nos campos de busca
entrada_busca.bind("<Return>", lambda event: buscar_processos())
entrada_filtro_secretaria.bind("<Return>", lambda event: buscar_processos())
entrada_filtro_situacao.bind("<Return>", lambda event: buscar_processos())


# Adicione esses bindings após a criação dos campos de busca
def manter_foco_apos_escape(event):
    # Limpa filtros mas mantém o foco no mesmo campo
    try:
        limpar_filtros(preservar_ordem=True)
    except Exception:
        pass
    try:
        widget = event.widget
        widget.focus_set()
        if isinstance(widget, tk.Entry):
            widget.icursor(tk.END)
    except Exception:
        pass
    return "break"


entrada_busca.bind("<Escape>", manter_foco_apos_escape)
entrada_filtro_secretaria.bind("<Escape>", manter_foco_apos_escape)
entrada_filtro_situacao.bind("<Escape>", manter_foco_apos_escape)
entrada_filtro_modalidade.bind("<Escape>", manter_foco_apos_escape)


def foco_para_botao_buscar(event):
    botao_buscar.focus_set()
    return "break"


def foco_para_filtro_modalidade(event):
    entrada_filtro_modalidade.focus_set()
    return "break"


# Shift+Tab no Limpar volta para Buscar
botao_limpar_filtros.bind("<Shift-Tab>", foco_para_botao_buscar)
# Shift+Tab no Buscar volta para Modalidade
botao_buscar.bind("<Shift-Tab>", foco_para_filtro_modalidade)

# Frame da tabela
frame_lista = tk.LabelFrame(janela, text="Processos Cadastrados", padx=10, pady=5, bg="#ECEFF1", fg="#37474F")
frame_lista.pack(fill="both", expand=True, padx=10, pady=0)

toggle_var = tk.BooleanVar(value=False)


# Função para selecionar todos os registros com Ctrl+A

def selecionar_todos_ctrl_a(event=None):
    """Alterna entre selecionar todos e remover seleção dos registros da tabela"""
    if tabela.get_children():  # Verifica se há registros na tabela
        # Se já há itens selecionados, remove a seleção
        if tabela.selection():
            tabela.selection_remove(tabela.get_children())  # Remove seleção de todos
            toggle_var.set(False)  # Desmarca o checkbox "Selecionar Todos"
        else:
            # Se não há seleção, seleciona todos
            tabela.selection_set(tabela.get_children())  # Seleciona todos os itens
            toggle_var.set(True)  # Marca o checkbox "Selecionar Todos"
        return "break"  # Impede que o evento se propague


def toggle_selecionar_todos():
    if toggle_var.get():  # Se estiver marcado
        tabela.selection_set(tabela.get_children())  # Seleciona todos os itens
    else:  # Se estiver desmarcado
        tabela.selection_remove(tabela.get_children())  # Desmarca todos os itens


# Frame do topo da lista
frame_topo_lista = tk.Frame(frame_lista, bg="#ECEFF1")
# Aumenta o espaço abaixo do topo da lista para separar do cabeçalho da tabela
frame_topo_lista.pack(fill="x", padx=0, pady=(0, 8))

# Botão de seleção (Checkbutton) antes do texto "Selecionar todos"
toggle_btn = tk.Checkbutton(
    frame_topo_lista,
    variable=toggle_var,
    command=toggle_selecionar_todos,
    indicatoron=True,
    width=2,
    bg="#ECEFF1",
    selectcolor="#C8E6C9",  # Verde claro ao marcar
    activebackground="#ECEFF1",
    bd=0,
    highlightthickness=0, highlightbackground="#ECEFF1", highlightcolor="#ECEFF1"
)
toggle_btn.pack(side=tk.LEFT, padx=0)


# Destaque em verde claro para o checkbox "Selecionar Todos" (hover/foco)
def _highlight_toggle_on(event=None):
    try:
        toggle_btn.config(bg="#DFF2E1", activebackground="#DFF2E1", selectcolor="#4CAF50")
    except Exception:
        pass


def _highlight_toggle_off(event=None):
    try:
        toggle_btn.config(bg="#ECEFF1", activebackground="#ECEFF1", selectcolor="#C8E6C9")
    except Exception:
        pass


toggle_btn.bind("<Enter>", _highlight_toggle_on)
toggle_btn.bind("<Leave>", _highlight_toggle_off)
toggle_btn.bind("<FocusIn>", _highlight_toggle_on)
toggle_btn.bind("<FocusOut>", _highlight_toggle_off)

# Label com o texto "Selecionar todos"
label_selecionar_todos = tk.Label(
    frame_topo_lista,
    text="Selecionar Todos",
    font=("Segoe UI", 8),
    bg="#ECEFF1",
    fg="#706E6E"
)
label_selecionar_todos.pack(side=tk.LEFT, padx=(0, 20))

# Estatísticas ao lado de "Selecionar Todos"
label_concluidos = tk.Label(frame_topo_lista, text="Concluídos: 0", font=("Segoe UI", 8),
                            bg="#ECEFF1", fg="#37474F")
label_concluidos.pack(side=tk.LEFT, padx=(0, 15))

label_andamento = tk.Label(frame_topo_lista, text="Em Andamento: 0", font=("Segoe UI", 8),
                           bg="#ECEFF1", fg="#37474F")
label_andamento.pack(side=tk.LEFT, padx=(0, 15))

label_editados = tk.Label(frame_topo_lista, text="Editados: 0", font=("Segoe UI", 8),
                          bg="#ECEFF1", fg="#37474F")
label_editados.pack(side=tk.LEFT, padx=(0, 15))

label_apagados = tk.Label(frame_topo_lista, text="Apagados: 0", font=("Segoe UI", 8),
                          bg="#ECEFF1", fg="#37474F")
label_apagados.pack(side=tk.LEFT, padx=(0, 0))

# python
label_exportados = tk.Label(frame_topo_lista, text="Exportações: 0", font=("Segoe UI", 8),
                            bg="#ECEFF1", fg="#37474F")
label_exportados.pack(side=tk.LEFT, padx=(0, 15))

label_importados = tk.Label(frame_topo_lista, text="Importações: 0", font=("Segoe UI", 8),
                            bg="#ECEFF1", fg="#37474F")
label_importados.pack(side=tk.LEFT, padx=(0, 15))

label_restaurados = tk.Label(frame_topo_lista, text="Restaurados: 0", font=("Segoe UI", 8),
                             bg="#ECEFF1", fg="#37474F")
label_restaurados.pack(side=tk.LEFT, padx=(0, 0))


# Funções de foco personalizado para ciclo de TAB
def foco_para_toggle_btn(event):
    toggle_btn.focus_set()
    return "break"


def destacar_toggle_btn(event):
    toggle_btn.config(selectcolor="#3DA19C")  # cor do quadrado


def remover_destaque_toggle_btn(event):
    toggle_btn.config(selectcolor="#95CCC9")  # cor original do quadrado


toggle_btn.bind("<FocusIn>", destacar_toggle_btn)
toggle_btn.bind("<FocusOut>", remover_destaque_toggle_btn)


def foco_para_entrada_numero(event):
    entrada_numero.focus_set()
    return "break"


def foco_para_botao_limpar_filtros(event):
    botao_limpar_filtros.focus_set()
    return "break"


# Bind do TAB no botão Limpar Filtros (busca)
botao_limpar_filtros.bind("<Tab>", foco_para_toggle_btn)

# Bind do TAB no botão Selecionar Todos
toggle_btn.bind("<Tab>", foco_para_entrada_numero)

# Bind do TAB reverso (Shift+Tab) para voltar do Selecionar Todos para Limpar Filtros
toggle_btn.bind("<Shift-Tab>", foco_para_botao_limpar_filtros)

# Bind do TAB reverso (Shift+Tab) para voltar do Nº do contrato para Selecionar Todos
entrada_numero.bind("<Shift-Tab>", foco_para_toggle_btn)

frame_tabela = tk.Frame(frame_lista, bg="#ECEFF1")
frame_tabela.pack(fill="both", expand=True)

scrollbar_y = tk.Scrollbar(frame_tabela)
scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

scrollbar_x = tk.Scrollbar(frame_tabela, orient=tk.HORIZONTAL)
scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)

# Configuração da tabela
colunas = (
    'data_registro',
    'numero_processo',
    'secretaria',
    'numero_licitacao',
    'situacao',
    'modalidade',
    'data_inicio',
    'data_entrega',
    'entregue_por',
    'devolvido_a',
    'contratado',
    'descricao'
)

ordem_colunas_reversa = {col: False for col in colunas}

cabecalhos = {
    'data_registro': 'Data Registro',
    'numero_processo': "Nº Processo",
    'secretaria': 'Secretaria',
    'numero_licitacao': "Nº Licitação",
    'situacao': 'Situação',
    'modalidade': 'Modalidade',
    'data_inicio': 'Recebimento',
    'data_entrega': 'Devolução',
    'entregue_por': 'Entregue por',
    'devolvido_a': 'Devolvido para',
    'contratado': 'Contratado',
    'descricao': 'Descrição'
}

# Defina estas variáveis ANTES de serem usadas nas funções
colunas_visiveis_padrao = [
    'data_registro', 'numero_processo', 'secretaria', 'numero_licitacao',
    'situacao', 'modalidade', 'data_inicio', 'data_entrega',
    'entregue_por', 'devolvido_a', 'contratado', 'descricao'
]

larguras_fixas = {
    'data_registro': 108,
    'numero_processo': 130,
    'secretaria': 95,
    'numero_licitacao': 115,
    'modalidade': 104,
    'situacao': 96,
    'data_inicio': 90,
    'data_entrega': 80,
    'entregue_por': 120,
    'devolvido_a': 120,
    'contratado': 120

}

larguras_padrao = {
    'descricao': 250  # Apenas para a coluna que não será mostrada
}

tabela = ttk.Treeview(
    frame_tabela,
    columns=colunas,
    show="headings",
    yscrollcommand=scrollbar_y.set,
    xscrollcommand=scrollbar_x.set,
    style="Treeview"
)

# Garantir seleção estendida para suportar Shift+Seta
tabela.configure(selectmode='extended')

tooltip = ToolTip(tabela)

ultima_linha_tooltip = None  # variável global para controlar a linha atual do tooltip

for col in colunas:
    tabela.heading(col, text=cabecalhos[col], command=lambda c=col: ordenar_coluna(c))


# Atualiza o estado do checkbox "Selecionar Todos" conforme a seleção na tabela
def atualizar_estado_selecionar_todos(event=None):
    try:
        total = len(tabela.get_children())
        selecionados = len(tabela.selection())
        marcado = total > 0 and selecionados == total
        toggle_var.set(marcado)
        if marcado:
            checkbox_selecionar_todos.select()
        else:
            checkbox_selecionar_todos.deselect()
    except Exception:
        pass


# Bind para refletir seleção no checkbox
tabela.bind('<<TreeviewSelect>>', atualizar_estado_selecionar_todos)


# --- Seleção com Shift+Seta para cima/baixo ---
def _set_anchor_on_click(event):
    try:
        # Ignora cliques fora das linhas
        region = tabela.identify_region(event.x, event.y)
        if region not in ("cell", "tree"):
            return
        row = tabela.identify_row(event.y)
        if row:
            tabela.focus(row)
            # Armazena a âncora para seleção de faixas
            setattr(tabela, 'anchor_item', row)
    except Exception:
        pass


def _shift_select(direction):
    try:
        items = tabela.get_children("")
        if not items:
            return "break"

        cur = tabela.focus()
        if not cur:
            sel = tabela.selection()
            cur = sel[0] if sel else items[0]

        idx = items.index(cur)
        if direction == 'up':
            new_idx = max(0, idx - 1)
        else:
            new_idx = min(len(items) - 1, idx + 1)

        new_item = items[new_idx]

        anchor = getattr(tabela, 'anchor_item', None)
        if anchor is None:
            anchor = cur
            setattr(tabela, 'anchor_item', anchor)

        a_idx = items.index(anchor)
        start = min(a_idx, new_idx)
        end = max(a_idx, new_idx)

        tabela.selection_set(items[start:end + 1])
        tabela.focus(new_item)
        tabela.see(new_item)
        return "break"
    except Exception:
        return "break"


# Binds de teclado e mouse para seleção estendida
tabela.bind('<Button-1>', _set_anchor_on_click, add='+')
tabela.bind('<Shift-Up>', lambda e: _shift_select('up'))
tabela.bind('<Shift-Down>', lambda e: _shift_select('down'))


# Adicione esta função para tratar o ENTER em botões
def ativar_botao(event):
    event.widget.invoke()


# Configuração dos botões com binding de ENTER
botoes = [
    (botao_cadastrar, cadastrar_processo),
    (botao_limpar, limpar_campos),
    (botao_editar, editar_processo),
    (botao_excluir, lambda: botao_excluir.invoke()),
    (botao_exportar, exportar_pdf),
    (botao_exportar_txt, exportar_txt),
    (botao_exportar_excel, exportar_excel),
    (botao_banco_dados, abrir_janela_banco_dados),
    (botao_restaurar, abrir_janela_restaurar),
    (botao_buscar, buscar_processos),
    (botao_limpar_filtros, limpar_filtros)
]

# Aplica os bindings em todos os botões
for botao, comando in botoes:
    botao.bind('<Return>', ativar_botao)
    botao.bind('<KP_Enter>', ativar_botao)  # Para o Enter do teclado numérico


def mostrar_tooltip_descricao(event):
    global ultima_linha_tooltip

    # Evita mostrar tooltip sobre o cabeçalho da tabela, mas não oculta
    region = tabela.identify_region(event.x, event.y)
    if region not in ("cell", "tree"):
        return

    # Identifica a linha sob o mouse
    item_id = tabela.identify_row(event.y)

    if item_id:
        descricao = tabela.set(item_id, 'descricao')
        situacao = tabela.set(item_id, 'situacao')

        if situacao == "Em Andamento" and descricao and descricao.strip():
            texto = descricao.strip()
            if len(texto) > 300:
                texto = texto[:300] + "..."

            # Atualiza a linha atual
            if item_id != ultima_linha_tooltip:
                # Remove destaque da linha anterior (se houver)
                if ultima_linha_tooltip is not None:
                    prev_sit = tabela.set(ultima_linha_tooltip, 'situacao')
                    prev_tag = 'concluido' if prev_sit == 'Concluído' else 'andamento' if prev_sit == 'Em Andamento' else ''
                    tabela.item(ultima_linha_tooltip, tags=(prev_tag,))
                ultima_linha_tooltip = item_id

            # Mostra apenas o conteúdo digitado em Observações
            tooltip.show(f"{texto}")
            # Aplica destaque visual à linha com tooltip
            try:
                tabela.item(item_id, tags=('hover_tooltip',))
            except Exception:
                pass
        else:
            # Se mudou para uma linha que não é "Em Andamento", oculta o tooltip
            if ultima_linha_tooltip is not None:
                tooltip.hide()
                # Restaura tag original pela situação
                try:
                    prev_sit = tabela.set(ultima_linha_tooltip, 'situacao')
                    prev_tag = 'concluido' if prev_sit == 'Concluído' else 'andamento' if prev_sit == 'Em Andamento' else ''
                    tabela.item(ultima_linha_tooltip, tags=(prev_tag,))
                except Exception:
                    pass
                ultima_linha_tooltip = None
    else:
        # Se saiu de uma linha "Em Andamento", oculta o tooltip
        if ultima_linha_tooltip is not None:
            tooltip.hide()
            try:
                prev_sit = tabela.set(ultima_linha_tooltip, 'situacao')
                prev_tag = 'concluido' if prev_sit == 'Concluído' else 'andamento' if prev_sit == 'Em Andamento' else ''
                tabela.item(ultima_linha_tooltip, tags=(prev_tag,))
            except Exception:
                pass
            ultima_linha_tooltip = None


def on_leave(event):
    global ultima_linha_tooltip
    tooltip.hide()
    try:
        if ultima_linha_tooltip is not None:
            prev_sit = tabela.set(ultima_linha_tooltip, 'situacao')
            prev_tag = 'concluido' if prev_sit == 'Concluído' else 'andamento' if prev_sit == 'Em Andamento' else ''
            tabela.item(ultima_linha_tooltip, tags=(prev_tag,))
    except Exception:
        pass
    ultima_linha_tooltip = None


# binds
tabela.bind('<Motion>', mostrar_tooltip_descricao)
tabela.bind('<Leave>', on_leave)


# Função para copiar valores
def copiar_numero_processo():
    item = tabela.focus()
    if not item:
        messagebox.showwarning("Aviso", "Selecione um registro.")
        return
    valores = tabela.item(item, "values")
    valor = valores[1]  # nº do contrato
    janela.clipboard_clear()
    janela.clipboard_append(valor)
    pyperclip.copy(valor)
    messagebox.showinfo("Copiado", f"Nº do contrato copiado: {valor}")


def copiar_numero_licitacao():
    item = tabela.focus()
    if not item:
        messagebox.showwarning("Aviso", "Selecione um registro.")
        return
    valores = tabela.item(item, "values")
    valor = valores[3]  # nº da licitação
    janela.clipboard_clear()
    janela.clipboard_append(valor)
    pyperclip.copy(valor)
    messagebox.showinfo("Copiado", f"Nº da licitação copiado: {valor}")


# Menu de contexto
menu_tabela = tk.Menu(janela, tearoff=0)
menu_tabela.add_command(label="Copiar Nº do Contrato", command=copiar_numero_processo)
menu_tabela.add_command(label="Copiar Nº da Licitação", command=copiar_numero_licitacao)


def mostrar_menu_tabela(event):
    item = tabela.identify_row(event.y)
    if item:
        tabela.selection_set(item)
        menu_tabela.tk_popup(event.x_root, event.y_root)


tabela.bind("<Button-3>", mostrar_menu_tabela)


def atualizar_estatisticas():
    global registros_concluidos, registros_andamento, registros_editados, registros_apagados
    global registros_exportados, registros_importados, registros_restaurados

    label_concluidos.config(text=f"Concluídos: {registros_concluidos}")
    label_andamento.config(text=f"Em Andamento: {registros_andamento}")
    label_editados.config(text=f"Editados: {registros_editados}")
    label_apagados.config(text=f"Apagados: {registros_apagados}")
    label_exportados.config(text=f"Exportações: {registros_exportados}")
    label_importados.config(text=f"Importações: {registros_importados}")
    label_restaurados.config(text=f"Restaurados: {registros_restaurados}")

    janela.after(5000, atualizar_estatisticas)


def contar_registros():
    global registros_concluidos, registros_andamento

    # Conta registros concluídos e em andamento
    cursor.execute("SELECT COUNT(*) FROM trabalhos_realizados WHERE situacao = 'Concluído'")
    registros_concluidos = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM trabalhos_realizados WHERE situacao = 'Em Andamento'")
    registros_andamento = cursor.fetchone()[0]

    atualizar_estatisticas()


# Inicia as estatísticas
contar_registros()


def listar_processos():
    for row in tabela.get_children():
        tabela.delete(row)

    cursor.execute('''
        SELECT data_registro, numero_processo, secretaria, numero_licitacao,
               situacao, modalidade, data_inicio, data_entrega,
               entregue_por, devolvido_a, contratado, descricao
        FROM trabalhos_realizados
        ORDER BY data_registro DESC
    ''')

    for row in cursor.fetchall():
        valores = list(row)
        valores[0] = formatar_data_hora_str(valores[0])

        # Formatar datas para dd/mm/aaaa
        if valores[0]:
            try:
                valores[0] = DateUtils.para_exibicao(str(valores[0]))
            except Exception:
                pass
        if valores[6]:
            try:
                # Garantir que a data de recebimento esteja no formato dia/mês/ano
                if "-" in valores[6]:
                    try:
                        valores[6] = DateUtils.para_exibicao(str(valores[6]))
                    except Exception:
                        pass
            except Exception:
                pass
        if valores[7]:
            try:
                # Garantir que a data de devolução esteja no formato dia/mês/ano
                if "-" in valores[7]:
                    try:
                        valores[7] = DateUtils.para_exibicao(str(valores[7]))
                    except Exception:
                        pass
            except Exception:
                pass

        # Tratar valores None - substituir por string vazia
        for i in range(len(valores)):
            if valores[i] is None:
                valores[i] = ""

        tag = 'concluido' if valores[4] == 'Concluído' else 'andamento' if valores[4] == 'Em Andamento' else ''
        tabela.insert("", "end", values=valores, tags=(tag,))

    # Após repopular a tabela, posiciona a rolagem no topo
    try:
        tabela.yview_moveto(0)
    except Exception:
        pass


# GitHUb

def mostrar_guia_atalhos():
    """Exibe uma janela com todos os atalhos de teclado disponíveis no programa"""
    janela_atalhos = tk.Toplevel(janela)
    janela_atalhos.title("Guia de Atalhos de Teclado")
    try:
        w, h = carregar_tamanho_janela('guia_atalhos', 500, 400)
        center_window(janela_atalhos, w, h)
    except Exception:
        janela_atalhos.geometry("500x400")
    janela_atalhos.resizable(True, True)
    janela_atalhos.transient(janela)
    janela_atalhos.grab_set()
    # Manter a janela sempre em primeiro plano
    janela_atalhos.attributes("-topmost", True)
    # Garantir foco para que o ESC funcione imediatamente
    try:
        janela_atalhos.focus_force()
    except Exception:
        pass

    def _fechar():
        try:
            janela_atalhos.update_idletasks()
            salvar_tamanho_janela('guia_atalhos', janela_atalhos.winfo_width(), janela_atalhos.winfo_height(),
                                  janela_atalhos.winfo_x(), janela_atalhos.winfo_y())
        except Exception:
            pass
        janela_atalhos.destroy()

    janela_atalhos.bind("<Escape>", lambda e: _fechar())
    try:
        janela_atalhos.protocol("WM_DELETE_WINDOW", _fechar)
    except Exception:
        pass

    # Estilo para o título
    titulo_style = {"font": ("Segoe UI", 12, "bold"), "pady": 10, "bg": "#ECEFF1"}

    # Estilo para as categorias
    categoria_style = {"font": ("Segoe UI", 10, "bold"), "anchor": "w", "bg": "#ECEFF1", "pady": 5}

    # Estilo para os atalhos
    atalho_style = {"font": ("Segoe UI", 9), "anchor": "w", "bg": "#ECEFF1"}

    # Frame principal
    frame_principal = tk.Frame(janela_atalhos, bg="#ECEFF1", padx=20, pady=10)
    frame_principal.pack(fill="both", expand=True)

    # Título
    tk.Label(frame_principal, text="Guia de Atalhos de Teclado", **titulo_style).pack(fill="x")

    # Frame para os atalhos com scrollbar
    frame_scroll = tk.Frame(frame_principal, bg="#ECEFF1")
    frame_scroll.pack(fill="both", expand=True, pady=10)

    # Adicionar scrollbar
    scrollbar = tk.Scrollbar(frame_scroll)
    scrollbar.pack(side="right", fill="y")

    # Canvas para conter os atalhos
    canvas = tk.Canvas(frame_scroll, bg="#ECEFF1", yscrollcommand=scrollbar.set)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar.config(command=canvas.yview)

    # Frame dentro do canvas para os atalhos
    frame_atalhos = tk.Frame(canvas, bg="#ECEFF1")
    canvas.create_window((0, 0), window=frame_atalhos, anchor="nw")

    # Configurar rolagem com a roda do mouse
    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    canvas.bind_all("<MouseWheel>", _on_mousewheel)

    # Lista de atalhos por categoria
    atalhos = {
        "Navegação": [
            ("Tab", "Avançar para o próximo campo"),
            ("Shift+Tab", "Voltar para o campo anterior"),
            ("Esc", "Limpar campos")
        ],
        "Seleção": [
            ("Ctrl+A", "Selecionar todo o texto em um campo"),
            ("Ctrl+T", "Selecionar todos os registros da tabela")
        ],
        "Edição": [
            ("Ctrl+X", "Recortar texto"),
            ("Ctrl+C", "Copiar texto"),
            ("Ctrl+V", "Colar texto"),
            ("Ctrl+BackSpace", "Apagar palavra anterior"),
            ("Ctrl+Delete", "Apagar palavra seguinte")
        ],
        "Busca e Salvamento": [
            ("Ctrl+F", "Buscar"),
            ("Ctrl+S", "Salvar")
        ],
        "Funções (F1–F12)": [
            ("F1", "Mostrar guia de atalhos"),
            ("F2", "Cadastrar"),
            ("F3", "Editar"),
            ("F4", "Exportar PDF"),
            ("F5", "Exportar TXT"),
            ("F6", "Exportar Excel"),
            ("F7", "Banco de Dados"),
            ("F8", "Exportar Banco"),
            ("F9", "Importar Banco"),
            ("F10", "Detalhes do Lembrete"),
            ("F11", "Restaurar Registros Excluídos"),
            ("F12", "Visualizar Registro Excluído")
        ],
        "Tabela": [
            ("Double-Click", "Visualizar processo selecionado")
        ]
    }

    # Adicionar atalhos ao frame
    row = 0
    for categoria, lista_atalhos in atalhos.items():
        # Adicionar categoria
        tk.Label(frame_atalhos, text=categoria, **categoria_style).grid(row=row, column=0, columnspan=2, sticky="w")
        row += 1

        # Adicionar atalhos
        for atalho, descricao in lista_atalhos:
            tk.Label(frame_atalhos, text=atalho, width=15, **atalho_style).grid(row=row, column=0, sticky="w",
                                                                                padx=(20, 0))
            tk.Label(frame_atalhos, text=descricao, **atalho_style).grid(row=row, column=1, sticky="w")
            row += 1

        # Espaço entre categorias
        tk.Label(frame_atalhos, text="", bg="#ECEFF1").grid(row=row, column=0)
        row += 1

    estilo_botao_padrao = {
        "bg": BUTTON_PRIMARY_BG,
        "fg": BUTTON_PRIMARY_FG,
        "activebackground": BUTTON_PRIMARY_ACTIVE_BG,
        "activeforeground": BUTTON_PRIMARY_ACTIVE_FG,
        "font": ("Segoe UI", 10, "bold"),
        "relief": tk.FLAT,
        "bd": 0,
        "highlightthickness": 2,
        "highlightbackground": BUTTON_PRIMARY_HIGHLIGHT
    }
    btn_fechar = tk.Button(frame_principal, text="Fechar", command=janela_atalhos.destroy, width=8,
                           **estilo_botao_padrao)
    btn_fechar.pack(pady=10)

    # Configurar o canvas para scroll
    frame_atalhos.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

    janela_atalhos.update_idletasks()


def backup_git():
    try:
        os.chdir(os.path.dirname(os.path.abspath(__file__)))

        # Verifica se há mudanças antes de tentar commit
        status = subprocess.run(
            ["git", "status", "--porcelain"],
            capture_output=True, text=True, check=False
        )

        if not status.stdout.strip():
            print("[INFO] Nenhuma mudança para backup")
            return

        subprocess.run(["git", "add", "."], check=True)

        result_commit = subprocess.run(
            ["git", "commit", "-m", f"Backup automático {datetime.now().strftime('%Y-%m-%d %H:%M')}"],
            capture_output=True, text=True
        )

        if "nothing to commit" in result_commit.stdout:
            print("[INFO] Nada para commitar")
            return

        result_push = subprocess.run(
            ["git", "push", "origin", "master"],
            capture_output=True, text=True
        )

        print("[INFO] Backup Git concluído com sucesso")

    except subprocess.CalledProcessError as e:
        print(f"[ERRO] Falha no backup Git: {e}")
    except Exception as e:
        print(f"[ERRO] Erro inesperado no backup Git: {e}")


def backup_git_thread():
    thread = threading.Thread(target=backup_git)
    thread.daemon = True
    thread.start()


# Backup Git configurado

# Vincular eventos
tabela.bind('<Motion>', mostrar_tooltip_descricao)
tabela.bind('<Leave>', on_leave)

for col in colunas:
    tabela.heading(col, text=cabecalhos[col], command=lambda c=col: ordenar_coluna(c))
    tabela.column(col, width=120, anchor="center")

tabela.pack(fill="both", expand=True)
tabela.tag_configure('concluido', background="#C8E6C9")  # Verde claro
tabela.tag_configure('andamento', background="#FFCDD2")  # Vermelho claro
tabela.tag_configure('hover_tooltip', background=ROW_TOOLTIP_HOVER_BG)

scrollbar_y.config(command=tabela.yview)
scrollbar_x.config(command=tabela.xview)

# Vincula duplo clique na tabela para visualização
tabela.bind("<Double-1>", lambda e: visualizar_processo())

# Vincula duplo clique nos campos para edição
for widget in [entrada_numero, entrada_secretaria, entrada_licitacao, entrada_modalidade,
               entrada_recebimento, entrada_devolucao, entrada_entregue_por, entrada_devolvido_a, entrada_descricao]:
    widget.bind("<Double-1>", ativar_edicao_campo)

# Vincula ESC globalmente para limpar campos
janela.bind("<Escape>", lambda e: limpar_campos())

# Configura o botão Limpar
botao_limpar.config(command=limpar_campos)

janela.bind("<Control-a>", selecionar_todos_ctrl_a)
janela.bind("<Control-A>", selecionar_todos_ctrl_a)


# =============================
# Atalhos globais de edição
# =============================
def _widget_textual_focado():
    try:
        w = janela.focus_get()
        return w if isinstance(w, (tk.Entry, tk.Text)) else None
    except Exception:
        return None


def ctrl_selecionar_texto(event=None):
    w = _widget_textual_focado()
    if not w:
        return
    try:
        if isinstance(w, tk.Entry):
            w.selection_range(0, tk.END)
        else:  # tk.Text
            w.tag_add('sel', '1.0', 'end-1c')
        return "break"
    except Exception:
        return


def _indice_inicio_palavra_anterior(texto, pos):
    i = pos
    while i > 0 and texto[i - 1].isspace():
        i -= 1
    while i > 0 and not texto[i - 1].isspace():
        i -= 1
    return i


def _indice_fim_proxima_palavra(texto, pos):
    n = len(texto)
    i = pos
    while i < n and texto[i].isspace():
        i += 1
    while i < n and not texto[i].isspace():
        i += 1
    return i


def ctrl_backspace(event=None):
    w = _widget_textual_focado()
    if not w:
        return
    try:
        if isinstance(w, tk.Entry):
            pos = w.index(INSERT)
            texto = w.get()
            ini = _indice_inicio_palavra_anterior(texto, pos)
            w.delete(ini, pos)
            w.icursor(ini)
        else:  # tk.Text
            linha, col = map(int, w.index(INSERT).split('.'))
            prefixo = w.get(f"{linha}.0", f"{linha}.{col}")
            ini_col = _indice_inicio_palavra_anterior(prefixo, len(prefixo))
            w.delete(f"{linha}.{ini_col}", f"{linha}.{col}")
            w.mark_set(INSERT, f"{linha}.{ini_col}")
        return "break"
    except Exception:
        return


def ctrl_delete(event=None):
    w = _widget_textual_focado()
    if not w:
        return
    try:
        if isinstance(w, tk.Entry):
            pos = w.index(INSERT)
            texto = w.get()
            fim = _indice_fim_proxima_palavra(texto, pos)
            w.delete(pos, fim)
        else:  # tk.Text
            linha, col = map(int, w.index(INSERT).split('.'))
            linha_texto = w.get(f"{linha}.0", f"{linha}.end-1c")
            fim_col = _indice_fim_proxima_palavra(linha_texto, col)
            w.delete(f"{linha}.{col}", f"{linha}.{fim_col}")
        return "break"
    except Exception:
        return


def ctrl_copy(event=None):
    w = _widget_textual_focado()
    if not w:
        return
    w.event_generate('<<Copy>>')
    return "break"


def ctrl_cut(event=None):
    w = _widget_textual_focado()
    if not w:
        return
    w.event_generate('<<Cut>>')
    return "break"


def ctrl_paste(event=None):
    w = _widget_textual_focado()
    if not w:
        return
    w.event_generate('<<Paste>>')
    return "break"


def ctrl_salvar(event=None):
    try:
        if str(botao_cadastrar['state']) != 'disabled':
            botao_cadastrar.invoke()
            return "break"
    except Exception:
        pass


def ctrl_buscar(event=None):
    try:
        entrada_busca.focus_set()
        return "break"
    except Exception:
        return


def piscar_botao(btn, vezes=2, intervalo=120):
    try:
        original_bg = str(btn.cget("bg"))

        def _to_rgb(hx):
            hx = hx.strip()
            if hx.startswith('#'):
                hx = hx[1:]
            if len(hx) == 3:
                hx = ''.join(c * 2 for c in hx)
            r = int(hx[0:2], 16)
            g = int(hx[2:4], 16)
            b = int(hx[4:6], 16)
            return r, g, b

        def _to_hex(r, g, b):
            return f"#{r:02x}{g:02x}{b:02x}"

        def _lighten(hx, factor=0.35):
            try:
                r, g, b = _to_rgb(hx)
                r = min(255, int(r + (255 - r) * factor))
                g = min(255, int(g + (255 - g) * factor))
                b = min(255, int(b + (255 - b) * factor))
                return _to_hex(r, g, b)
            except Exception:
                return original_bg

        alt_bg = _lighten(original_bg)

        def _step(i=0):
            if i >= vezes * 2:
                btn.config(bg=original_bg)
                return
            btn.config(bg=alt_bg if i % 2 == 0 else original_bg)
            btn.after(intervalo, lambda: _step(i + 1))

        _step()
    except Exception:
        pass


def abrir_detalhes_lembrete(event=None):
    try:
        if not (janela_lembretes_aberta and janela_lembretes_aberta.winfo_exists()):
            abrir_lembretes()
        win = janela_lembretes_aberta
        if not (win and win.winfo_exists()):
            return

        def _find_button_text(widget, text):
            for child in widget.winfo_children():
                try:
                    if isinstance(child, tk.Button) and str(child.cget("text")).strip().lower() == text.lower():
                        return child
                except Exception:
                    pass
                res = _find_button_text(child, text)
                if res:
                    return res
            return None

        btn = _find_button_text(win, "Visualizar")
        if btn:
            piscar_botao(btn)
            btn.invoke()
        else:
            win.focus_set()
    except Exception:
        pass


def abrir_visualizar_registro_excluido(event=None):
    try:
        abrir_janela_restaurar()

        def _find_restaurar_window():
            for w in janela.winfo_children():
                try:
                    if isinstance(w, tk.Toplevel) and str(w.title()).strip() == "Restaurar Registro Excluído":
                        return w
                except Exception:
                    pass
            return None

        win = _find_restaurar_window()
        if not (win and win.winfo_exists()):
            return

        def _find_button_text(widget, text):
            for child in widget.winfo_children():
                try:
                    if isinstance(child, tk.Button) and str(child.cget("text")).strip().lower() == text.lower():
                        return child
                except Exception:
                    pass
                res = _find_button_text(child, text)
                if res:
                    return res
            return None

        btn = _find_button_text(win, "Visualizar")
        if btn:
            piscar_botao(btn)
            btn.invoke()
        else:
            win.focus_set()
    except Exception:
        pass


def abrir_exportar_banco_f(event=None):
    try:
        abrir_janela_banco_dados()

        def _find_window():
            for w in janela.winfo_children():
                try:
                    if isinstance(w, tk.Toplevel) and "Gerenciamento do Banco de Dados" in str(w.title()):
                        return w
                except Exception:
                    pass
            return None

        win = _find_window()
        if not win:
            return

        def _find_button_text(widget, text):
            for child in widget.winfo_children():
                try:
                    if isinstance(child, tk.Button) and str(child.cget("text")).strip() == text:
                        return child
                except Exception:
                    pass
                res = _find_button_text(child, text)
                if res:
                    return res
            return None

        btn = _find_button_text(win, "📤 Exportar Banco")
        if btn:
            piscar_botao(btn)
            btn.invoke()
    except Exception:
        pass


def abrir_importar_banco_f(event=None):
    try:
        abrir_janela_banco_dados()

        def _find_window():
            for w in janela.winfo_children():
                try:
                    if isinstance(w, tk.Toplevel) and "Gerenciamento do Banco de Dados" in str(w.title()):
                        return w
                except Exception:
                    pass
            return None

        win = _find_window()
        if not win:
            return

        def _find_button_text(widget, text):
            for child in widget.winfo_children():
                try:
                    if isinstance(child, tk.Button) and str(child.cget("text")).strip() == text:
                        return child
                except Exception:
                    pass
                res = _find_button_text(child, text)
                if res:
                    return res
            return None

        btn = _find_button_text(win, "📥 Importar Banco")
        if btn:
            piscar_botao(btn)
            btn.invoke()
    except Exception:
        pass


# Binds globais, com prioridade para widgets de texto
janela.bind("<Control-a>", ctrl_selecionar_texto)
janela.bind("<Control-A>", ctrl_selecionar_texto)
janela.bind("<Control-BackSpace>", ctrl_backspace)
janela.bind("<Control-Delete>", ctrl_delete)
janela.bind("<Control-c>", ctrl_copy)
janela.bind("<Control-C>", ctrl_copy)
janela.bind("<Control-x>", ctrl_cut)
janela.bind("<Control-X>", ctrl_cut)
# Removido bind de Ctrl+V para evitar colagens duplicadas; usar padrão do Tk
# janela.bind("<Control-v>", ctrl_paste)
# janela.bind("<Control-V>", ctrl_paste)
janela.bind("<Control-s>", ctrl_salvar)
janela.bind("<Control-S>", ctrl_salvar)
janela.bind("<Control-f>", ctrl_buscar)
janela.bind("<Control-F>", ctrl_buscar)
janela.bind("<Control-t>", selecionar_todos_ctrl_a)
janela.bind("<Control-T>", selecionar_todos_ctrl_a)
janela.bind("<F1>", lambda event: mostrar_guia_atalhos())
janela.bind("<F2>", lambda e: (piscar_botao(botao_cadastrar), botao_cadastrar.invoke()))
janela.bind("<F3>", lambda e: (piscar_botao(botao_editar), botao_editar.invoke()))
janela.bind("<F4>", lambda e: (piscar_botao(botao_exportar), botao_exportar.invoke()))
janela.bind("<F5>", lambda e: (piscar_botao(botao_exportar_txt), botao_exportar_txt.invoke()))
janela.bind("<F6>", lambda e: (piscar_botao(botao_exportar_excel), botao_exportar_excel.invoke()))
janela.bind("<F7>", lambda e: (piscar_botao(botao_banco_dados), botao_banco_dados.invoke()))
janela.bind("<F8>", abrir_exportar_banco_f)
janela.bind("<F9>", abrir_importar_banco_f)
janela.bind("<F10>", abrir_detalhes_lembrete)
janela.bind("<F11>", lambda e: (piscar_botao(botao_restaurar), botao_restaurar.invoke()))
janela.bind("<F12>", abrir_visualizar_registro_excluido)


def verificar_mudancas_e_backup():
    """Verifica se há mudanças no repositório Git e inicia o backup se necessário.

    Muda para o diretório do projeto, verifica o status do Git para detectar
    alterações não commitadas, e inicia o processo de backup em uma thread
    separada se houver mudanças.
    """
    try:
        # Caminho do projeto Git
        caminho_projeto = os.path.dirname(os.path.abspath(__file__))

        # Verifica se o diretório existe
        if not os.path.isdir(caminho_projeto):
            print(f"[ERRO] Diretório do projeto não encontrado: {caminho_projeto}")
            return

        # Salva o diretório atual para restaurar depois
        diretorio_atual = os.getcwd()

        try:
            # Muda para o diretório do projeto
            os.chdir(caminho_projeto)

            # Verifica se há um repositório Git
            if not os.path.isdir(os.path.join(caminho_projeto, ".git")):
                print(f"[AVISO] Repositório Git não encontrado em: {caminho_projeto}")
                return

            # Verifica o status do Git
            result_status = subprocess.run(
                ["git", "status", "--porcelain"],
                capture_output=True,
                text=True,
                check=False  # tk.Não levanta exceção se o comando falhar
            )

            # Verifica se há mudanças
            if result_status.returncode == 0 and result_status.stdout.strip():
                print("[INFO] Mudanças detectadas. Iniciando backup Git...")
                backup_git_thread()
            else:
                print("[INFO] Nenhuma mudança detectada. Backup Git não necessário.")

        finally:
            # Restaura o diretório original
            os.chdir(diretorio_atual)

    except FileNotFoundError:
        print("[ERRO] Git não encontrado. Verifique se o Git está instalado e no PATH.")
    except PermissionError:
        print(f"[ERRO] Sem permissão para acessar o diretório: {caminho_projeto}")
    except Exception as e:
        print(f"[ERRO] Falha ao verificar mudanças Git: {e}")


def ao_fechar():
    """Função executada quando a janela principal é fechada.

    Salva as larguras das colunas da tabela e fecha a aplicação de forma segura.
    """
    try:
        salvar_larguras_colunas(tabela, colunas)
        print("[INFO] Configurações de colunas salvas")
        try:
            janela.update_idletasks()
            salvar_tamanho_janela('principal', janela.winfo_width(), janela.winfo_height(), janela.winfo_x(),
                                  janela.winfo_y())
        except Exception:
            pass
        janela.destroy()
        print("[INFO] Aplicação encerrada pelo usuário")
    except Exception as e:
        print(f"[ERRO] Falha ao fechar aplicação: {e}")
        # Tenta fechar a janela mesmo em caso de erro
        try:
            janela.destroy()
        except:
            pass


janela.protocol("WM_DELETE_WINDOW", ao_fechar)

# ================================================================
# 1. Inicialização dos Processos e Configurações
# ================================================================
# Listar processos e ordenar a coluna 'situacao'
listar_processos()
ordenar_coluna('situacao')

# Verifica se houve mudanças e inicia o backup em thread, apenas se necessário
verificar_mudancas_e_backup()
nomes_autocomplete = carregar_nomes_autocomplete()

# ================================================================
# 2. Configuração de Campos de Ordem e Foco
# ================================================================
# Vincula a ação de pressionar ENTER aos campos de ordem
for campo in campos_ordem[:-1]:
    campo.bind("<Return>", focar_proximo_campo)

# Configura a ordem das abas e corrige o foco no campo "modalidade_recebimento"
configurar_tab_ordem()

# ================================================================
# 3. Inicialização da Interface Tkinter
# ================================================================
# Verifica promessas do dia antes de iniciar a interface
verificar_promessas_do_dia()

# Inicia a interface gráfica Tkinter
janela.mainloop()

# ================================================================
# 4. Fechamento da Conexão com o Banco de Dados
# ================================================================
# Fecha a conexão com o banco ao sair da aplicação
conn.close()
